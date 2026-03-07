"""Unit tests for _build_condition_check_xlsx."""
import sys
import os
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from worker.main import _build_condition_check_xlsx


def _make_rows():
    return [
        {
            "filename": "doc1.pdf",
            "company_name": "테스트 주식회사",
            "method": "pymupdf",
            "pages": 5,
            "elapsed_s": 2.3,
            "parse_warning": "JSON 응답 일부를 복구했습니다.",
            "conditions": [
                {"result": True, "evidence": "매출 증가 추세"},
                {"result": False, "evidence": "영업이익 부족"},
            ],
        },
        {
            "filename": "doc2.pdf",
            "company_name": "Sample Corp",
            "method": "nova_hybrid",
            "pages": 12,
            "elapsed_s": 4.1,
            "error": "PDF 손상",
            "conditions": [],
        },
        {
            "filename": "doc3.pdf",
            "company_name": "정상 기업",
            "method": "pymupdf",
            "pages": 3,
            "elapsed_s": 1.5,
            "conditions": [
                {"result": True, "evidence": "조건 충족 근거"},
                {"result": True, "evidence": "두 번째 조건 충족"},
            ],
        },
    ]


def test_xlsx_creates_file(tmp_path):
    rows = _make_rows()
    conditions = ["매출 성장률 10% 이상", "영업이익률 5% 이상"]
    xlsx_path = _build_condition_check_xlsx(tmp_path, rows, conditions)
    assert xlsx_path.exists()
    assert xlsx_path.suffix == ".xlsx"
    assert xlsx_path.stat().st_size > 0


def test_xlsx_has_correct_sheets(tmp_path):
    from openpyxl import load_workbook

    rows = _make_rows()
    conditions = ["조건A", "조건B"]
    xlsx_path = _build_condition_check_xlsx(tmp_path, rows, conditions)

    wb = load_workbook(str(xlsx_path))
    assert "조건 검사 결과" in wb.sheetnames
    assert "요약" in wb.sheetnames


def test_xlsx_correct_row_count(tmp_path):
    from openpyxl import load_workbook

    rows = _make_rows()
    conditions = ["조건A", "조건B"]
    xlsx_path = _build_condition_check_xlsx(tmp_path, rows, conditions)

    wb = load_workbook(str(xlsx_path))
    ws = wb["조건 검사 결과"]
    # 1 header row + 3 data rows = 4
    assert ws.max_row == 4


def test_xlsx_correct_column_count(tmp_path):
    from openpyxl import load_workbook

    rows = _make_rows()
    conditions = ["조건A", "조건B"]
    xlsx_path = _build_condition_check_xlsx(tmp_path, rows, conditions)

    wb = load_workbook(str(xlsx_path))
    ws = wb["조건 검사 결과"]
    # 7 base + 2 conditions * 2 (result+evidence) = 11
    assert ws.max_column == 11


def test_xlsx_pass_fail_values(tmp_path):
    from openpyxl import load_workbook

    rows = _make_rows()
    conditions = ["조건A", "조건B"]
    xlsx_path = _build_condition_check_xlsx(tmp_path, rows, conditions)

    wb = load_workbook(str(xlsx_path))
    ws = wb["조건 검사 결과"]
    assert ws.cell(row=2, column=6).value == "JSON 응답 일부를 복구했습니다."
    # Row 2 (first data row), col 8 (first result column)
    assert ws.cell(row=2, column=8).value == "✓ 충족"
    assert ws.cell(row=2, column=9).value == "매출 증가 추세"
    assert ws.cell(row=2, column=10).value == "✗ 미충족"


def test_xlsx_summary_sheet(tmp_path):
    from openpyxl import load_workbook

    rows = _make_rows()
    conditions = ["조건A", "조건B"]
    xlsx_path = _build_condition_check_xlsx(tmp_path, rows, conditions)

    wb = load_workbook(str(xlsx_path))
    ws = wb["요약"]
    assert ws.cell(row=1, column=2).value == 3  # total files
    assert ws.cell(row=2, column=2).value == 1  # warning files
    assert ws.cell(row=3, column=2).value == 1  # error files
    assert ws.cell(row=4, column=2).value == 2  # conditions count


def test_xlsx_empty_rows(tmp_path):
    """Should handle empty results gracefully."""
    xlsx_path = _build_condition_check_xlsx(tmp_path, [], ["조건1"])
    assert xlsx_path.exists()


def test_xlsx_many_conditions(tmp_path):
    """Should handle many conditions without error."""
    conditions = [f"조건 {i}" for i in range(20)]
    rows = [{
        "filename": "test.pdf",
        "company_name": "Test",
        "method": "pymupdf",
        "pages": 1,
        "elapsed_s": 0.5,
        "conditions": [{"result": i % 2 == 0, "evidence": f"근거 {i}"} for i in range(20)],
    }]
    xlsx_path = _build_condition_check_xlsx(tmp_path, rows, conditions)
    assert xlsx_path.exists()
    assert xlsx_path.stat().st_size > 0
