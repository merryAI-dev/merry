"""
기업별 JSON 결과를 CSV로 내보내기.

batch_process.py 가 생성한 results/{company}/*.json 을 읽어서:
  - results/financial_summary.csv  : 연도별 재무 지표 (IS/BS/CF)
  - results/cap_table_summary.csv  : 주주 현황
  - results/doc_inventory.csv      : 문서 목록 + 처리 상태

사용법:
    python3 scripts/export_to_csv.py results/
    python3 scripts/export_to_csv.py results/ --output exports/
"""

import argparse
import csv
import json
import logging
import re
import sys
from pathlib import Path
from typing import Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ─── 표준재무제표증명 전용 텍스트 파서 ─────────────────────────────────────────
# 국세청 표준재무제표는 좌우 2단 구조
# IS left  col[2]=코드, col[3]=금액  /  IS right col[6]=코드, col[7]=금액
# BS left  col[2]=코드, col[3]=금액  /  BS right col[7]=코드, col[8]=금액

_IS_LEFT_MAP  = {"001": "revenue", "129": "operating_income"}
_IS_RIGHT_MAP = {"219": "net_income", "220": "net_income", "225": "net_income"}
_BS_LEFT_MAP  = {"228": "total_assets"}
_BS_RIGHT_MAP = {"333": "total_liabilities", "382": "total_equity"}

# 제목 감지: 국세청 양식은 글자 사이 공백 삽입 ("재 무 상 태 표")
_IS_TITLE_RE = re.compile(r"손\s*익\s*계\s*산|손익계산")
_BS_TITLE_RE = re.compile(r"재\s*무\s*상\s*태\s*표|재무상태표|대차대조표")


def _parse_amount(s: str) -> Optional[float]:
    if not s:
        return None
    cleaned = re.sub(r"[,\s원]", "", str(s).strip())
    try:
        v = float(cleaned)
        return v if v != 0 else None
    except ValueError:
        return None


def _extract_year_from_content(content: str) -> Optional[str]:
    m = re.search(r"(20\d{2})년\s*\d{1,2}월\s*\d{1,2}일\s*부터", content)
    return m.group(1) if m else None


def _table_title(rows: list) -> str:
    return " ".join(str(c) for r in rows[:2] for c in r if c)


def _extract_from_row(row: list, code_map: dict, target: dict, left: bool):
    """좌(col2/3) 또는 우(col6/7 or col7/8) 컬럼에서 코드→지표 추출."""
    if left:
        code = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        amount_str = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
    else:
        # BS 우측: col7/8 / IS 우측: col6/7 — 길이로 구분
        if len(row) >= 9:
            code = str(row[7]).strip() if row[7] is not None else ""
            amount_str = str(row[8]).strip() if row[8] is not None else ""
        elif len(row) >= 8:
            code = str(row[6]).strip() if row[6] is not None else ""
            amount_str = str(row[7]).strip() if row[7] is not None else ""
        else:
            return

    if code and code in code_map and amount_str:
        key = code_map[code]
        if key not in target:
            target[key] = _parse_amount(amount_str)


def _parse_standard_financial_form(record: dict) -> dict:
    """표준재무제표증명에서 재무 지표 파싱."""
    content = record.get("structured_content", {})
    pages = content.get("pages", []) if isinstance(content, dict) else []

    metrics_is: dict = {}
    metrics_bs: dict = {}
    year = None

    for page in pages:
        for elem in page.get("elements", []):
            if elem.get("type") == "text":
                if not year:
                    year = _extract_year_from_content(elem.get("content", ""))

            if elem.get("type") == "table":
                rows = elem.get("content", {}).get("rows", [])
                if not rows:
                    continue

                title = _table_title(rows)
                is_is = bool(_IS_TITLE_RE.search(title))
                is_bs = bool(_BS_TITLE_RE.search(title))

                for row in rows:
                    if is_is:
                        _extract_from_row(row, _IS_LEFT_MAP,  metrics_is, left=True)
                        _extract_from_row(row, _IS_RIGHT_MAP, metrics_is, left=False)
                    if is_bs:
                        _extract_from_row(row, _BS_LEFT_MAP,  metrics_bs, left=True)
                        _extract_from_row(row, _BS_RIGHT_MAP, metrics_bs, left=False)

    return {
        "year": year,
        "is_metrics": metrics_is,
        "bs_metrics": metrics_bs,
        "parsed_from_text": True,
    }

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)


# ─── 재무 데이터 평탄화 ──────────────────────────────────────────────────────

def _flatten_financial_tables(company: str, filename: str, ft: dict) -> list:
    """financial_tables dict -> 연도별 행 리스트."""
    rows = []

    is_ = ft.get("income_statement", {})
    bs_ = ft.get("balance_sheet", {})
    cf_ = ft.get("cash_flow", {})

    # 연도 목록: IS 기준, 없으면 BS, CF 순
    years = (
        is_.get("years")
        or bs_.get("years")
        or cf_.get("years")
        or ["unknown"]
    )

    def _get(metrics: dict, key: str, idx: int) -> Optional[float]:
        vals = metrics.get(key, [])
        if idx < len(vals):
            return vals[idx]
        return None

    for i, year in enumerate(years):
        row = {
            "company_name": company,
            "source_file": filename,
            "year": year,
            # IS
            "revenue": _get(is_.get("metrics", {}), "revenue", i),
            "operating_income": _get(is_.get("metrics", {}), "operating_income", i),
            "net_income": _get(is_.get("metrics", {}), "net_income", i),
            # BS
            "total_assets": _get(bs_.get("metrics", {}), "total_assets", i),
            "total_liabilities": _get(bs_.get("metrics", {}), "total_liabilities", i),
            "total_equity": _get(bs_.get("metrics", {}), "total_equity", i),
            # CF
            "operating_cf": _get(cf_.get("metrics", {}), "operating_cf", i),
            "investing_cf": _get(cf_.get("metrics", {}), "investing_cf", i),
            "financing_cf": _get(cf_.get("metrics", {}), "financing_cf", i),
            # 메타
            "is_found": is_.get("found", False),
            "bs_found": bs_.get("found", False),
            "cf_found": cf_.get("found", False),
        }
        rows.append(row)

    return rows


def _flatten_cap_table(company: str, filename: str, ft: dict) -> list:
    """cap_table -> 주주별 행 리스트."""
    cap = ft.get("cap_table", {})
    if not cap.get("found"):
        return []

    rows = []
    total_shares = cap.get("total_shares", 0)
    for sh in cap.get("shareholders", []):
        rows.append({
            "company_name": company,
            "source_file": filename,
            "shareholder_name": sh.get("name"),
            "shares": sh.get("shares"),
            "ratio": sh.get("ratio"),
            "total_shares_issued": total_shares,
        })
    return rows


# ─── 결과 파일 수집 ──────────────────────────────────────────────────────────

def collect_records(results_dir: Path) -> list:
    """results/ 하위의 모든 .json 파일(에러 제외) 로드."""
    records = []
    for json_file in sorted(results_dir.rglob("*.json")):
        if json_file.name.endswith(".error.json"):
            continue
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                rec = json.load(f)
            records.append(rec)
        except (json.JSONDecodeError, OSError) as e:
            logger.warning(f"Skip {json_file}: {e}")
    return records


def collect_error_records(results_dir: Path) -> list:
    """실패한 .error.json 파일 수집."""
    records = []
    for json_file in sorted(results_dir.rglob("*.error.json")):
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                rec = json.load(f)
            records.append(rec)
        except (json.JSONDecodeError, OSError):
            pass
    return records


# ─── CSV 저장 헬퍼 ───────────────────────────────────────────────────────────

def _write_csv(path: Path, rows: list, fieldnames: list):
    if not rows:
        logger.info(f"  No data for {path.name}, skipping.")
        return
    with open(path, "w", newline="", encoding="utf-8-sig") as f:  # utf-8-sig: Excel 호환
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    logger.info(f"  Saved {len(rows)} rows -> {path}")


# ─── Excel 저장 ─────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="3182F6") if _OPENPYXL else None
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10) if _OPENPYXL else None
_BODY_FONT   = Font(size=10) if _OPENPYXL else None


def _xl_sheet(wb: "Workbook", title: str, rows: list, fieldnames: list):
    """Write rows to a new worksheet with styled header row."""
    ws = wb.create_sheet(title=title)

    # Header row
    for col, name in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for r_idx, row in enumerate(rows, 2):
        for col, name in enumerate(fieldnames, 1):
            val = row.get(name)
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.font = _BODY_FONT
            cell.alignment = Alignment(vertical="center")

    # Auto column width (max 40)
    for col in range(1, len(fieldnames) + 1):
        max_len = len(str(fieldnames[col - 1]))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def _write_xlsx(path: "Path", fin_rows: list, cap_rows: list, inventory_rows: list):
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _xl_sheet(wb, "Financial Summary", fin_rows, [
        "company_name", "year", "source_file",
        "revenue", "operating_income", "net_income",
        "total_assets", "total_liabilities", "total_equity",
        "operating_cf", "investing_cf", "financing_cf",
        "is_found", "bs_found", "cf_found",
    ])
    _xl_sheet(wb, "Cap Table", cap_rows, [
        "company_name", "shareholder_name", "shares",
        "ratio", "total_shares_issued", "source_file",
    ])
    _xl_sheet(wb, "Document Inventory", inventory_rows, [
        "company_name", "company_name_source", "filename",
        "doc_purpose", "doc_type", "status", "error",
        "total_pages", "processing_method", "processing_time_seconds",
        "cache_hit", "processed_at",
    ])

    wb.save(path)
    logger.info(f"  Saved Excel -> {path}")


# ─── 메인 익스포트 ───────────────────────────────────────────────────────────

def export(results_dir: Path, output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)

    records = collect_records(results_dir)
    error_records = collect_error_records(results_dir)

    logger.info(f"Loaded {len(records)} ok records, {len(error_records)} error records")

    # ── 1. 문서 인벤토리 (전체 현황) ──────────────────────────────────────────
    inventory_rows = []
    for rec in records + error_records:
        inventory_rows.append({
            "company_name": rec.get("company_name"),
            "company_name_source": rec.get("company_name_source"),
            "filename": rec.get("filename"),
            "doc_purpose": rec.get("doc_purpose"),
            "doc_type": rec.get("doc_type"),
            "status": rec.get("status"),
            "error": rec.get("error"),
            "total_pages": rec.get("total_pages"),
            "processing_method": rec.get("processing_method"),
            "processing_time_seconds": rec.get("processing_time_seconds"),
            "cache_hit": rec.get("cache_hit"),
            "processed_at": rec.get("processed_at"),
        })

    # company_name + filename 기준 정렬
    inventory_rows.sort(key=lambda r: (r.get("company_name") or "", r.get("filename") or ""))

    _write_csv(
        output_dir / "doc_inventory.csv",
        inventory_rows,
        fieldnames=[
            "company_name", "company_name_source", "filename",
            "doc_purpose", "doc_type", "status", "error",
            "total_pages", "processing_method", "processing_time_seconds",
            "cache_hit", "processed_at",
        ],
    )

    # ── 2. 재무 요약 (IS / BS / CF) ───────────────────────────────────────────
    fin_rows = []
    for rec in records:
        company = rec.get("company_name", "unknown")
        filename = rec.get("filename", "")

        # 표준재무제표증명 → 전용 텍스트 파서 우선 사용
        if rec.get("doc_purpose") == "financial_statement" and rec.get("structured_content"):
            parsed = _parse_standard_financial_form(rec)
            if parsed["is_metrics"] or parsed["bs_metrics"]:
                fin_rows.append({
                    "company_name": company,
                    "year": parsed["year"] or "unknown",
                    "source_file": filename,
                    "revenue": parsed["is_metrics"].get("revenue"),
                    "operating_income": parsed["is_metrics"].get("operating_income"),
                    "net_income": parsed["is_metrics"].get("net_income"),
                    "total_assets": parsed["bs_metrics"].get("total_assets"),
                    "total_liabilities": parsed["bs_metrics"].get("total_liabilities"),
                    "total_equity": parsed["bs_metrics"].get("total_equity"),
                    "operating_cf": None,
                    "investing_cf": None,
                    "financing_cf": None,
                    "is_found": bool(parsed["is_metrics"]),
                    "bs_found": bool(parsed["bs_metrics"]),
                    "cf_found": False,
                })
                continue

        ft = rec.get("financial_tables") or {}
        if not ft:
            continue
        rows = _flatten_financial_tables(company, filename, ft)
        fin_rows.extend(rows)

    fin_rows.sort(key=lambda r: (r.get("company_name") or "", str(r.get("year") or "")))

    _write_csv(
        output_dir / "financial_summary.csv",
        fin_rows,
        fieldnames=[
            "company_name", "year", "source_file",
            "revenue", "operating_income", "net_income",
            "total_assets", "total_liabilities", "total_equity",
            "operating_cf", "investing_cf", "financing_cf",
            "is_found", "bs_found", "cf_found",
        ],
    )

    # ── 3. Cap Table ─────────────────────────────────────────────────────────
    cap_rows = []
    for rec in records:
        ft = rec.get("financial_tables") or {}
        if not ft:
            continue
        company = rec.get("company_name", "unknown")
        filename = rec.get("filename", "")
        cap_rows.extend(_flatten_cap_table(company, filename, ft))

    cap_rows.sort(key=lambda r: (r.get("company_name") or "", str(r.get("ratio") or 0)))

    _write_csv(
        output_dir / "cap_table_summary.csv",
        cap_rows,
        fieldnames=[
            "company_name", "shareholder_name", "shares",
            "ratio", "total_shares_issued", "source_file",
        ],
    )

    # ── 4. 회사별 개별 CSV (선택: 재무 데이터 있는 기업만) ───────────────────
    companies = {}
    for row in fin_rows:
        companies.setdefault(row["company_name"], []).append(row)

    if companies:
        per_company_dir = output_dir / "per_company"
        per_company_dir.mkdir(exist_ok=True)
        for company, rows in companies.items():
            safe = company.replace("/", "_").replace(" ", "_")
            _write_csv(
                per_company_dir / f"{safe}_financial.csv",
                rows,
                fieldnames=[
                    "year", "revenue", "operating_income", "net_income",
                    "total_assets", "total_liabilities", "total_equity",
                    "operating_cf", "investing_cf", "financing_cf",
                    "source_file",
                ],
            )

    # ── 5. Consolidated Excel workbook ───────────────────────────────────────
    if _OPENPYXL:
        _write_xlsx(
            output_dir / "financial_report.xlsx",
            fin_rows=fin_rows,
            cap_rows=cap_rows,
            inventory_rows=inventory_rows,
        )
    else:
        logger.warning("openpyxl not installed — skipping Excel output")

    logger.info(f"\nExport complete -> {output_dir}")


def main():
    parser = argparse.ArgumentParser(description="Export batch results to CSV")
    parser.add_argument("results_dir", help="batch_process.py 출력 폴더")
    parser.add_argument("--output", default=None, help="CSV 저장 폴더 (기본: results_dir/)")
    args = parser.parse_args()

    results_dir = Path(args.results_dir).resolve()
    if not results_dir.is_dir():
        logger.error(f"Not a directory: {results_dir}")
        sys.exit(1)

    output_dir = Path(args.output).resolve() if args.output else results_dir

    export(results_dir=results_dir, output_dir=output_dir)


if __name__ == "__main__":
    main()
