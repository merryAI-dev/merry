#!/usr/bin/env python3
"""
투자 검토 엑셀 파일 분석 스크립트
IS요약, Cap Table, 투자조건을 파싱하여 JSON 출력
"""
import argparse
import json
import sys
from openpyxl import load_workbook

def find_cell_value(ws, search_terms, search_range=100):
    """시트에서 특정 텍스트가 포함된 셀의 인접 값을 찾음"""
    for row in range(1, search_range):
        for col in range(1, 20):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell_str = str(cell.value).strip()
                for term in search_terms:
                    if term in cell_str:
                        # 오른쪽 셀 값 반환
                        right_val = ws.cell(row=row, column=col+1).value
                        if right_val is not None:
                            return right_val
                        # 아래쪽 셀 값 반환
                        below_val = ws.cell(row=row+1, column=col).value
                        if below_val is not None:
                            return below_val
    return None

def parse_investment_terms(wb):
    """투자조건 시트 파싱"""
    terms = {}
    sheet_names = ['투자조건체크', '투자조건', 'Investment Terms', '조건']
    
    ws = None
    for name in sheet_names:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    
    if ws:
        terms['investment_amount'] = find_cell_value(ws, ['투자금액', '투자원금', 'Investment Amount'])
        terms['price_per_share'] = find_cell_value(ws, ['투자단가', '주당가격', 'Price per Share'])
        terms['shares'] = find_cell_value(ws, ['투자주식수', '취득주식수', 'Shares'])
        terms['security_type'] = find_cell_value(ws, ['투자유형', '증권종류', 'RCPS', 'Security Type'])
        terms['investment_date'] = find_cell_value(ws, ['투자일', '납입일', 'Investment Date'])
        terms['exit_date'] = find_cell_value(ws, ['회수예정일', 'Exit Date', '만기일'])
    
    return terms

def parse_income_statement(wb):
    """IS요약/손익추정 시트 파싱"""
    is_data = {'company': {}, 'reviewer': {}}
    sheet_names = ['IS요약', '손익추정', 'IS', 'Income Statement', '5. IS요약']
    
    ws = None
    for name in wb.sheetnames:
        for search in sheet_names:
            if search in name:
                ws = wb[name]
                break
        if ws:
            break
    
    if ws:
        # 연도별 순이익 찾기
        for row in range(1, 50):
            for col in range(1, 20):
                cell = ws.cell(row=row, column=col)
                if cell.value and '당기순이익' in str(cell.value):
                    # 해당 행의 연도별 값 추출
                    for c in range(col+1, col+10):
                        header = ws.cell(row=1, column=c).value or ws.cell(row=2, column=c).value
                        val = ws.cell(row=row, column=c).value
                        if header and val:
                            year = str(header).replace('년', '').replace('E', '').strip()
                            if year.isdigit():
                                is_data['company'][year] = val
                    break
        
        # 심사역 제시 찾기
        for row in range(1, 50):
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                if cell.value and '심사역' in str(cell.value):
                    # 해당 섹션의 당기순이익 찾기
                    for r in range(row, row+20):
                        check = ws.cell(row=r, column=col).value
                        if check and '당기순이익' in str(check):
                            for c in range(col+1, col+10):
                                header = ws.cell(row=1, column=c).value or ws.cell(row=2, column=c).value
                                val = ws.cell(row=r, column=c).value
                                if header and val:
                                    year = str(header).replace('년', '').replace('E', '').strip()
                                    if year.isdigit():
                                        is_data['reviewer'][year] = val
                            break
                    break
    
    return is_data

def parse_cap_table(wb):
    """Cap Table 파싱"""
    cap_data = {'total_shares': None, 'shareholders': []}
    sheet_names = ['Cap Table', 'CapTable', '주주현황', '지분구조']
    
    ws = None
    for name in wb.sheetnames:
        for search in sheet_names:
            if search.lower() in name.lower():
                ws = wb[name]
                break
        if ws:
            break
    
    if ws:
        cap_data['total_shares'] = find_cell_value(ws, ['총발행주식수', '발행주식총수', 'Total Shares'])
    
    return cap_data

def analyze_valuation_file(filepath):
    """메인 분석 함수"""
    wb = load_workbook(filepath, data_only=True)
    
    result = {
        'sheets': wb.sheetnames,
        'investment_terms': parse_investment_terms(wb),
        'income_statement': parse_income_statement(wb),
        'cap_table': parse_cap_table(wb)
    }
    
    wb.close()
    return result

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='투자 검토 엑셀 파일 분석')
    parser.add_argument('filepath', help='분석할 엑셀 파일 경로')
    parser.add_argument('--output', '-o', help='출력 JSON 파일 경로')
    args = parser.parse_args()
    
    try:
        result = analyze_valuation_file(args.filepath)
        output = json.dumps(result, ensure_ascii=False, indent=2, default=str)
        
        if args.output:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(output)
            print(f"분석 완료: {args.output}")
        else:
            print(output)
    except Exception as e:
        print(f"오류: {e}", file=sys.stderr)
        sys.exit(1)
