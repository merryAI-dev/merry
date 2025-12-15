#!/usr/bin/env python3
"""
Exit 프로젝션 엑셀 생성 스크립트
PER 기반 시나리오별 수익률 분석
"""
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 스타일 정의
BLUE_FONT = Font(color="0000FF", bold=False)
BOLD_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INPUT_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
RESULT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def format_cell(ws, row, col, value=None, font=None, fill=None, number_format=None, alignment=None):
    """셀 포맷팅 헬퍼"""
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    if alignment:
        cell.alignment = alignment
    cell.border = THIN_BORDER
    return cell

def generate_exit_projection(
    investment_amount,
    price_per_share,
    shares,
    total_shares,
    net_income_company,
    net_income_reviewer,
    target_year,
    company_name,
    per_multiples,
    investment_year=None,
    output_path=None
):
    """Exit 프로젝션 엑셀 생성"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Exit 프로젝션"
    
    # 컬럼 너비 설정
    col_widths = [20, 18, 18, 18, 18, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    row = 1
    
    # === 제목 ===
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    title_cell = ws.cell(row=row, column=1, value=f"{company_name} {target_year}년 Exit 프로젝션")
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    row += 2
    
    # === 투자 조건 섹션 ===
    format_cell(ws, row, 1, "투자 조건", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    
    # 투자금액
    format_cell(ws, row, 1, "투자금액")
    inv_cell = format_cell(ws, row, 2, investment_amount, BLUE_FONT, INPUT_FILL, '#,##0"원"')
    row += 1
    
    # 투자단가
    format_cell(ws, row, 1, "투자단가")
    pps_cell = format_cell(ws, row, 2, price_per_share, BLUE_FONT, INPUT_FILL, '#,##0"원"')
    row += 1
    
    # 투자주식수
    format_cell(ws, row, 1, "투자주식수")
    shares_cell = format_cell(ws, row, 2, shares, BLUE_FONT, INPUT_FILL, '#,##0"주"')
    row += 1
    
    # 총 발행주식수
    format_cell(ws, row, 1, "총 발행주식수 (투자 후)")
    total_cell = format_cell(ws, row, 2, total_shares, BLUE_FONT, INPUT_FILL, '#,##0"주"')
    row += 1
    
    # 지분율
    format_cell(ws, row, 1, "지분율")
    ownership_row = row
    format_cell(ws, row, 2, f"={shares_cell.coordinate}/{total_cell.coordinate}", number_format='0.00%')
    row += 1
    
    # 투자기간
    inv_year = investment_year or datetime.now().year
    holding_period = target_year - inv_year
    format_cell(ws, row, 1, "투자기간")
    period_cell = format_cell(ws, row, 2, holding_period, BLUE_FONT, INPUT_FILL, '0.0"년"')
    row += 2
    
    # === 순이익 가정 섹션 ===
    format_cell(ws, row, 1, f"{target_year}년 당기순이익 가정", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    
    # 회사제시
    format_cell(ws, row, 1, "회사제시")
    ni_company_cell = format_cell(ws, row, 2, net_income_company, BLUE_FONT, INPUT_FILL, '#,##0"원"')
    row += 1
    
    # 심사역제시
    format_cell(ws, row, 1, "심사역제시")
    ni_reviewer_cell = format_cell(ws, row, 2, net_income_reviewer, BLUE_FONT, INPUT_FILL, '#,##0"원"')
    row += 2
    
    # === Exit 분석 - 회사제시 ===
    company_start_row = row
    format_cell(ws, row, 1, "Exit 분석 - 회사제시 기준", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    
    # 헤더
    headers = ["PER 배수", "기업가치", "주당가치", "회수금액", "멀티플", "IRR"]
    for col, h in enumerate(headers, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1
    
    # PER별 계산 - 회사제시
    for per in per_multiples:
        per_cell = format_cell(ws, row, 1, per, BLUE_FONT, INPUT_FILL, '0"x"')
        # 기업가치 = 순이익 * PER
        ev_formula = f"={ni_company_cell.coordinate}*{per_cell.coordinate}"
        ev_cell = format_cell(ws, row, 2, ev_formula, number_format='#,##0"원"')
        # 주당가치 = 기업가치 / 총주식수
        sp_formula = f"={ev_cell.coordinate}/{total_cell.coordinate}"
        sp_cell = format_cell(ws, row, 3, sp_formula, number_format='#,##0"원"')
        # 회수금액 = 주당가치 * 투자주식수
        rec_formula = f"={sp_cell.coordinate}*{shares_cell.coordinate}"
        rec_cell = format_cell(ws, row, 4, rec_formula, number_format='#,##0"원"')
        # 멀티플 = 회수금액 / 투자금액
        mult_formula = f"={rec_cell.coordinate}/{inv_cell.coordinate}"
        format_cell(ws, row, 5, mult_formula, fill=RESULT_FILL, number_format='0.00"x"')
        # IRR = (멀티플)^(1/기간) - 1
        irr_formula = f"=POWER({rec_cell.coordinate}/{inv_cell.coordinate},1/{period_cell.coordinate})-1"
        format_cell(ws, row, 6, irr_formula, fill=RESULT_FILL, number_format='0.0%')
        row += 1
    
    row += 1
    
    # === Exit 분석 - 심사역제시 ===
    format_cell(ws, row, 1, "Exit 분석 - 심사역제시 기준", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    
    # 헤더
    for col, h in enumerate(headers, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1
    
    # PER별 계산 - 심사역제시
    for per in per_multiples:
        per_cell = format_cell(ws, row, 1, per, BLUE_FONT, INPUT_FILL, '0"x"')
        ev_formula = f"={ni_reviewer_cell.coordinate}*{per_cell.coordinate}"
        ev_cell = format_cell(ws, row, 2, ev_formula, number_format='#,##0"원"')
        sp_formula = f"={ev_cell.coordinate}/{total_cell.coordinate}"
        sp_cell = format_cell(ws, row, 3, sp_formula, number_format='#,##0"원"')
        rec_formula = f"={sp_cell.coordinate}*{shares_cell.coordinate}"
        rec_cell = format_cell(ws, row, 4, rec_formula, number_format='#,##0"원"')
        mult_formula = f"={rec_cell.coordinate}/{inv_cell.coordinate}"
        format_cell(ws, row, 5, mult_formula, fill=RESULT_FILL, number_format='0.00"x"')
        irr_formula = f"=POWER({rec_cell.coordinate}/{inv_cell.coordinate},1/{period_cell.coordinate})-1"
        format_cell(ws, row, 6, irr_formula, fill=RESULT_FILL, number_format='0.0%')
        row += 1
    
    row += 2
    
    # === 범례 ===
    format_cell(ws, row, 1, "범례", BOLD_FONT)
    row += 1
    legend_cell = ws.cell(row=row, column=1, value="파란색 텍스트")
    legend_cell.font = BLUE_FONT
    ws.cell(row=row, column=2, value="= 입력값 (수정 가능)")
    row += 1
    ws.cell(row=row, column=1, value="노란색 배경").fill = INPUT_FILL
    ws.cell(row=row, column=2, value="= 핵심 가정")
    row += 1
    ws.cell(row=row, column=1, value="녹색 배경").fill = RESULT_FILL
    ws.cell(row=row, column=2, value="= 결과값")
    
    # 저장
    if output_path is None:
        output_path = f"{company_name}_{target_year}_Exit_프로젝션.xlsx"
    
    wb.save(output_path)
    print(f"✅ Exit 프로젝션 생성 완료: {output_path}")
    return output_path

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Exit 프로젝션 엑셀 생성')
    parser.add_argument('--investment_amount', type=float, required=True, help='투자금액 (원)')
    parser.add_argument('--price_per_share', type=float, required=True, help='투자단가 (원/주)')
    parser.add_argument('--shares', type=int, required=True, help='투자주식수')
    parser.add_argument('--total_shares', type=int, required=True, help='총 발행주식수')
    parser.add_argument('--net_income_company', type=float, required=True, help='회사제시 당기순이익')
    parser.add_argument('--net_income_reviewer', type=float, required=True, help='심사역제시 당기순이익')
    parser.add_argument('--target_year', type=int, required=True, help='Exit 목표연도')
    parser.add_argument('--company_name', type=str, required=True, help='회사명')
    parser.add_argument('--per_multiples', type=str, default="7,8,10", help='PER 배수들 (콤마 구분)')
    parser.add_argument('--investment_year', type=int, default=None, help='투자연도')
    parser.add_argument('--output', '-o', type=str, default=None, help='출력 파일 경로')
    
    args = parser.parse_args()
    per_list = [int(x.strip()) for x in args.per_multiples.split(',')]
    
    generate_exit_projection(
        investment_amount=args.investment_amount,
        price_per_share=args.price_per_share,
        shares=args.shares,
        total_shares=args.total_shares,
        net_income_company=args.net_income_company,
        net_income_reviewer=args.net_income_reviewer,
        target_year=args.target_year,
        company_name=args.company_name,
        per_multiples=per_list,
        investment_year=args.investment_year,
        output_path=args.output
    )
