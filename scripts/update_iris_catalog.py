#!/usr/bin/env python3
"""
Update IRIS+ catalog JSON from the official IRIS+ Catalog of Metrics Excel file.

Default source URL:
https://s3.amazonaws.com/giin-web-assets/iris/assets/files/iris/IRIS 5.3c Catalog of Metrics.xlsx.zip

This script preserves existing policy_theme_mapping and sdg_info from the current
data/iris_plus_catalog.json when available.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from zipfile import ZipFile

import httpx
import openpyxl

DEFAULT_URL = (
    "https://s3.amazonaws.com/giin-web-assets/iris/assets/files/iris/"
    "IRIS 5.3c Catalog of Metrics.xlsx.zip"
)

DEFAULT_SHEET_KEYWORD = "Catalog of Metrics"


def _slugify(text: str) -> str:
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "general"


def _normalize_primary_category(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    if value.lower().replace("-", "") == "crosscategory":
        return "Cross-Category"
    return value


def _parse_sdg_number(header: str) -> Optional[int]:
    match = re.match(r"SDG\s+(\d+):", header)
    if match:
        return int(match.group(1))
    return None


def _download_zip(url: str) -> bytes:
    with httpx.Client(timeout=60.0, follow_redirects=True) as client:
        response = client.get(url)
        response.raise_for_status()
        return response.content


def _extract_xlsx_from_zip(zip_bytes: bytes, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    with ZipFile(BytesIO(zip_bytes)) as zf:
        xlsx_files = [name for name in zf.namelist() if name.lower().endswith(".xlsx")]
        if not xlsx_files:
            raise ValueError("No .xlsx file found inside the zip.")
        target = xlsx_files[0]
        zf.extract(target, output_dir)
        return output_dir / target


def _load_existing_metadata(
    path: Path,
) -> Tuple[Dict[str, Any], Dict[str, Any], Dict[str, Any], Dict[str, Any]]:
    if not path.exists():
        return {}, {}, {}, {}
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    policy_mapping = data.get("policy_theme_mapping", {}) or {}
    sdg_info = data.get("sdg_info", {}) or {}

    existing_metrics = {}
    for category in data.get("categories", []):
        for subcategory in category.get("subcategories", []):
            for metric in subcategory.get("metrics", []):
                code = metric.get("code")
                if code:
                    existing_metrics[code] = metric
    return data, policy_mapping, sdg_info, existing_metrics


def _resolve_sheet(workbook: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    for name in workbook.sheetnames:
        if DEFAULT_SHEET_KEYWORD in name:
            return workbook[name]
    return workbook[workbook.sheetnames[0]]


def _collect_headers(sheet: openpyxl.worksheet.worksheet.Worksheet) -> List[Optional[str]]:
    return list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))


def _build_category_skeleton(existing: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    default_categories = [
        ("environmental", "Environmental Impact", "환경 임팩트"),
        ("social", "Social Impact", "사회 임팩트"),
        ("governance", "Governance & Ethics", "거버넌스 및 윤리"),
    ]
    category_map = {}
    existing_categories = {c.get("id"): c for c in existing.get("categories", []) if c.get("id")}
    for cat_id, cat_name, cat_name_kr in default_categories:
        existing_entry = existing_categories.get(cat_id, {})
        existing_name = existing_entry.get("name", cat_name)
        existing_name_kr = existing_entry.get("name_kr")
        if not existing_name_kr or existing_name_kr == existing_name:
            resolved_name_kr = cat_name_kr
        else:
            resolved_name_kr = existing_name_kr
        category_map[cat_id] = {
            "id": cat_id,
            "name": existing_name,
            "name_kr": resolved_name_kr,
            "subcategories": {},
        }
    return category_map


def _classify_category(
    env_flag: bool,
    social_flag: bool,
    primary_category: Optional[str],
) -> str:
    env_primary = {
        "Agriculture",
        "Biodiversity",
        "Biodiversity & Ecosystems",
        "Climate",
        "Energy",
        "Land",
        "Oceans and Coastal Zones",
        "Pollution",
        "Waste",
        "Waste & Circularity",
        "Water",
    }
    soc_primary = {
        "Diversity & Inclusion",
        "Education",
        "Employment",
        "Financial Services",
        "Health",
        "Infrastructure",
        "Real Estate",
    }

    if env_flag and not social_flag:
        return "environmental"
    if social_flag and not env_flag:
        return "social"
    if primary_category in env_primary:
        return "environmental"
    if primary_category in soc_primary:
        return "social"
    if env_flag and social_flag:
        return "social"
    return "governance"


def _build_catalog(
    xlsx_path: Path,
    existing_catalog: Dict[str, Any],
    policy_mapping: Dict[str, Any],
    sdg_info: Dict[str, Any],
    existing_metrics: Dict[str, Any],
) -> Dict[str, Any]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = _resolve_sheet(workbook)
    headers = _collect_headers(sheet)
    header_index = {h: i for i, h in enumerate(headers) if h}

    required = ["ID", "Metric Name", "Definition"]
    for key in required:
        if key not in header_index:
            raise ValueError(f"Missing required column: {key}")

    theme_start = header_index.get("Impact Category & Theme")
    theme_end = header_index.get("Social and/or Environmental Focus")
    theme_cols: List[Tuple[int, str]] = []
    if theme_start is not None and theme_end is not None and theme_end > theme_start:
        theme_cols = [
            (i, headers[i])
            for i in range(theme_start + 1, theme_end)
            if headers[i]
        ]

    sdg_cols = [
        (i, h) for i, h in enumerate(headers)
        if isinstance(h, str) and h.startswith("SDG ") and ":" in h
    ]

    category_map = _build_category_skeleton(existing_catalog)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = row[header_index["ID"]]
        if not code:
            continue
        code = str(code).strip()
        if not code:
            continue

        name = row[header_index["Metric Name"]] or ""
        name = str(name).strip()
        definition = row[header_index["Definition"]] or ""
        footnote = row[header_index.get("Footnote", -1)] if "Footnote" in header_index else None

        description = str(definition).strip()
        if footnote:
            description = f"{description} {str(footnote).strip()}".strip()

        reporting_format = row[header_index.get("Reporting Format", -1)] if "Reporting Format" in header_index else None
        unit = str(reporting_format).strip() if reporting_format else ""

        primary_category = _normalize_primary_category(
            row[header_index.get("Primary Impact Category", -1)]
            if "Primary Impact Category" in header_index
            else None
        )
        section = row[header_index.get("Section", -1)] if "Section" in header_index else None
        subsection = row[header_index.get("Subsection", -1)] if "Subsection" in header_index else None

        env_flag = bool(row[header_index.get("Environmental", -1)]) if "Environmental" in header_index else False
        social_flag = bool(row[header_index.get("Social", -1)]) if "Social" in header_index else False

        sdgs: List[int] = []
        for idx, header in sdg_cols:
            if row[idx]:
                sdg_num = _parse_sdg_number(header)
                if sdg_num:
                    sdgs.append(sdg_num)
        sdgs = sorted(set(sdgs))

        themes = [theme for idx, theme in theme_cols if row[idx]]

        keywords: List[str] = []
        for value in [primary_category, section, subsection]:
            if value:
                keywords.append(str(value))
        keywords.extend([str(t) for t in themes])
        keywords = sorted({k.strip() for k in keywords if k and str(k).strip()})

        existing_metric = existing_metrics.get(code, {})
        name_kr = existing_metric.get("name_kr") or name
        keywords_kr = existing_metric.get("keywords_kr") or keywords

        metric_entry = {
            "code": code,
            "name": name,
            "name_kr": name_kr,
            "description": description,
            "unit": unit,
            "sdgs": sdgs,
            "keywords_kr": keywords_kr,
        }

        category_id = _classify_category(env_flag, social_flag, primary_category)
        category = category_map[category_id]

        subcategory_name = primary_category or str(subsection or section or "General")
        subcategory_name = subcategory_name.strip() if subcategory_name else "General"
        subcategory_id = _slugify(subcategory_name)

        subcategories = category["subcategories"]
        if subcategory_id not in subcategories:
            subcategories[subcategory_id] = {
                "id": subcategory_id,
                "name": subcategory_name,
                "name_kr": subcategory_name,
                "metrics": [],
            }

        subcategories[subcategory_id]["metrics"].append(metric_entry)

    # Convert subcategories dicts to lists and sort entries
    categories: List[Dict[str, Any]] = []
    for category in category_map.values():
        subcategories = list(category["subcategories"].values())
        for sub in subcategories:
            sub["metrics"] = sorted(sub["metrics"], key=lambda m: m["code"])
        category["subcategories"] = sorted(subcategories, key=lambda s: s["name"])
        categories.append(category)

    categories = sorted(categories, key=lambda c: c["id"])

    return {
        "version": "5.3c",
        "last_updated": datetime.now(timezone.utc).isoformat(),
        "source": "GIIN IRIS+ Catalog (https://iris.thegiin.org) - IRIS+ 5.3c Catalog of Metrics",
        "categories": categories,
        "policy_theme_mapping": policy_mapping,
        "sdg_info": sdg_info,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Update IRIS+ catalog JSON from official Excel.")
    parser.add_argument("--url", default=DEFAULT_URL, help="URL to the IRIS+ catalog zip")
    parser.add_argument("--zip-path", default="", help="Path to a local zip file")
    parser.add_argument("--xlsx-path", default="", help="Path to a local Excel file")
    parser.add_argument(
        "--output",
        default="data/iris_plus_catalog.json",
        help="Output JSON path",
    )
    parser.add_argument(
        "--existing",
        default="",
        help="Existing JSON path to preserve mappings/labels (defaults to output path)",
    )
    parser.add_argument(
        "--workdir",
        default="temp/iris_plus_download",
        help="Working directory for downloads/extraction",
    )
    args = parser.parse_args()

    output_path = Path(args.output)
    workdir = Path(args.workdir)

    existing_path = Path(args.existing) if args.existing else output_path
    existing_catalog, policy_mapping, sdg_info, existing_metrics = _load_existing_metadata(existing_path)

    xlsx_path: Optional[Path] = None
    if args.xlsx_path:
        xlsx_path = Path(args.xlsx_path)
    elif args.zip_path:
        zip_path = Path(args.zip_path)
        if not zip_path.exists():
            raise FileNotFoundError(f"Zip not found: {zip_path}")
        zip_bytes = zip_path.read_bytes()
        xlsx_path = _extract_xlsx_from_zip(zip_bytes, workdir)
    else:
        zip_bytes = _download_zip(args.url)
        xlsx_path = _extract_xlsx_from_zip(zip_bytes, workdir)

    if not xlsx_path or not xlsx_path.exists():
        raise FileNotFoundError("Failed to resolve Excel path.")

    catalog = _build_catalog(
        xlsx_path=xlsx_path,
        existing_catalog=existing_catalog,
        policy_mapping=policy_mapping,
        sdg_info=sdg_info,
        existing_metrics=existing_metrics,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)

    print(f"Updated catalog written to {output_path}")


if __name__ == "__main__":
    main()
