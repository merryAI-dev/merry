#!/usr/bin/env python3
"""
고급 Exit 프로젝션 엑셀 생성 스크립트
- 부분 매각 시나리오 (2단계 Exit)
- 할인율 적용 NPV 계산
- 복합 시나리오 분석
"""
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 스타일 정의
BLUE_FONT = Font(color="0000FF", bold=False)
BOLD_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INPUT_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
RESULT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
SCENARIO_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def format_cell(ws, row, col, value=None, font=None, fill=None, number_format=None, alignment=None):
    """셀 포맷팅 헬퍼"""
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    if alignment:
        cell.alignment = alignment
    cell.border = THIN_BORDER
    return cell

def generate_advanced_exit_projection(
    investment_amount,
    price_per_share,
    shares,
    total_shares,
    net_income_2029,
    net_income_2030,
    company_name,
    per_multiples,
    partial_exit_ratio=0.5,
    discount_rate=0.10,
    investment_year=2025,
    output_path=None
):
    """
    고급 Exit 프로젝션 생성

    시나리오:
    1. 전체 매각 (2029년)
    2. 부분 매각 (2029년 50% + 2030년 50%)
    3. 할인율 적용 (NPV)
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Advanced Exit 프로젝션"

    # 컬럼 너비 설정
    for i in range(1, 15):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions['A'].width = 25

    row = 1

    # === 제목 ===
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    title_cell = ws.cell(row=row, column=1, value=f"{company_name} Exit 프로젝션 분석 (2029-2030)")
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    row += 2

    # === 투자 조건 섹션 ===
    format_cell(ws, row, 1, "투자 조건", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    format_cell(ws, row, 1, "투자금액")
    inv_cell = format_cell(ws, row, 2, investment_amount, BLUE_FONT, INPUT_FILL, '#,##0"원"')
    row += 1

    format_cell(ws, row, 1, "투자단가")
    pps_cell = format_cell(ws, row, 2, price_per_share, BLUE_FONT, INPUT_FILL, '#,##0"원"')
    row += 1

    format_cell(ws, row, 1, "투자주식수")
    shares_cell = format_cell(ws, row, 2, shares, BLUE_FONT, INPUT_FILL, '#,##0"주"')
    row += 1

    format_cell(ws, row, 1, "총 발행주식수")
    total_cell = format_cell(ws, row, 2, total_shares, BLUE_FONT, INPUT_FILL, '#,##0"주"')
    row += 1

    format_cell(ws, row, 1, "지분율")
    format_cell(ws, row, 2, f"={shares_cell.coordinate}/{total_cell.coordinate}", number_format='0.00%')
    row += 1

    format_cell(ws, row, 1, "투자연도")
    inv_year_cell = format_cell(ws, row, 2, investment_year, BLUE_FONT, INPUT_FILL, '0')
    row += 2

    # === 가정 섹션 ===
    format_cell(ws, row, 1, "분석 가정", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    format_cell(ws, row, 1, "2029년 당기순이익")
    ni_2029_cell = format_cell(ws, row, 2, net_income_2029, BLUE_FONT, INPUT_FILL, '#,##0"원"')
    row += 1

    format_cell(ws, row, 1, "2030년 당기순이익")
    ni_2030_cell = format_cell(ws, row, 2, net_income_2030, BLUE_FONT, INPUT_FILL, '#,##0"원"')
    row += 1

    format_cell(ws, row, 1, "부분 매각 비율 (1차)")
    partial_cell = format_cell(ws, row, 2, partial_exit_ratio, BLUE_FONT, INPUT_FILL, '0%')
    row += 1

    format_cell(ws, row, 1, "할인율")
    discount_cell = format_cell(ws, row, 2, discount_rate, BLUE_FONT, INPUT_FILL, '0%')
    row += 2

    # ==========================================
    # 시나리오 1: 2029년 전체 매각
    # ==========================================
    scenario1_start = row
    format_cell(ws, row, 1, "【시나리오 1】 2029년 전체 매각", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    # 헤더
    headers = ["PER", "기업가치", "주당가치", "회수금액", "멀티플", "투자기간", "IRR"]
    for col, h in enumerate(headers, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1

    s1_data_rows = []
    for per in per_multiples:
        start_col_row = row
        per_cell = format_cell(ws, row, 1, per, BLUE_FONT, INPUT_FILL, '0"x"')

        # 기업가치 = 순이익 * PER
        ev_formula = f"={ni_2029_cell.coordinate}*{per_cell.coordinate}"
        ev_cell = format_cell(ws, row, 2, ev_formula, number_format='#,##0"원"')

        # 주당가치
        sp_formula = f"={ev_cell.coordinate}/{total_cell.coordinate}"
        sp_cell = format_cell(ws, row, 3, sp_formula, number_format='#,##0"원"')

        # 회수금액 = 주당가치 * 투자주식수
        rec_formula = f"={sp_cell.coordinate}*{shares_cell.coordinate}"
        rec_cell = format_cell(ws, row, 4, rec_formula, number_format='#,##0"원"')

        # 멀티플
        mult_formula = f"={rec_cell.coordinate}/{inv_cell.coordinate}"
        mult_cell = format_cell(ws, row, 5, mult_formula, fill=RESULT_FILL, number_format='0.00"x"')

        # 투자기간
        period_formula = f"=2029-{inv_year_cell.coordinate}"
        period_cell = format_cell(ws, row, 6, period_formula, number_format='0"년"')

        # IRR
        irr_formula = f"=POWER({mult_cell.coordinate},1/{period_cell.coordinate})-1"
        irr_cell = format_cell(ws, row, 7, irr_formula, fill=RESULT_FILL, number_format='0.0%')

        s1_data_rows.append({
            'per': per_cell,
            'rec': rec_cell,
            'mult': mult_cell,
            'irr': irr_cell,
            'sp': sp_cell
        })
        row += 1

    row += 1

    # ==========================================
    # 시나리오 2: 부분 매각 (2029년 50% + 2030년 50%)
    # ==========================================
    scenario2_start = row
    format_cell(ws, row, 1, "【시나리오 2】 부분 매각 (2029년 50% + 2030년 50%)", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    row += 1

    # 헤더
    headers2 = ["PER", "2029 주당가치", "1차 회수액", "2030 주당가치", "2차 회수액", "총 회수액",
                "멀티플", "복합 IRR", "비고"]
    for col, h in enumerate(headers2, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1

    s2_data_rows = []
    for per in per_multiples:
        per_cell = format_cell(ws, row, 1, per, BLUE_FONT, INPUT_FILL, '0"x"')

        # 2029년 주당가치
        ev_2029 = f"={ni_2029_cell.coordinate}*{per_cell.coordinate}"
        sp_2029_formula = f"=({ev_2029})/{total_cell.coordinate}"
        sp_2029_cell = format_cell(ws, row, 2, sp_2029_formula, number_format='#,##0"원"')

        # 1차 회수액 (50% 매각)
        rec1_formula = f"={sp_2029_cell.coordinate}*{shares_cell.coordinate}*{partial_cell.coordinate}"
        rec1_cell = format_cell(ws, row, 3, rec1_formula, number_format='#,##0"원"')

        # 2030년 주당가치 (PER 동일 가정)
        ev_2030 = f"={ni_2030_cell.coordinate}*{per_cell.coordinate}"
        sp_2030_formula = f"=({ev_2030})/{total_cell.coordinate}"
        sp_2030_cell = format_cell(ws, row, 4, sp_2030_formula, number_format='#,##0"원"')

        # 2차 회수액 (나머지 50% 매각)
        rec2_formula = f"={sp_2030_cell.coordinate}*{shares_cell.coordinate}*(1-{partial_cell.coordinate})"
        rec2_cell = format_cell(ws, row, 5, rec2_formula, number_format='#,##0"원"')

        # 총 회수액
        total_rec_formula = f"={rec1_cell.coordinate}+{rec2_cell.coordinate}"
        total_rec_cell = format_cell(ws, row, 6, total_rec_formula, fill=SCENARIO_FILL, number_format='#,##0"원"')

        # 멀티플
        mult2_formula = f"={total_rec_cell.coordinate}/{inv_cell.coordinate}"
        mult2_cell = format_cell(ws, row, 7, mult2_formula, fill=RESULT_FILL, number_format='0.00"x"')

        # 복합 IRR (NPV 기반 근사)
        # 정확한 IRR은 XIRR 함수가 필요하지만, 간단한 근사치 사용
        # IRR ≈ ((총회수액/투자액)^(1/평균보유기간) - 1)
        # 평균보유기간 = (4년 * 50% + 5년 * 50%) = 4.5년
        avg_period = 4.5
        irr2_formula = f"=POWER({mult2_cell.coordinate},1/{avg_period})-1"
        irr2_cell = format_cell(ws, row, 8, irr2_formula, fill=RESULT_FILL, number_format='0.0%')

        # 비고
        format_cell(ws, row, 9, f"1차: 2029년 / 2차: 2030년")

        s2_data_rows.append({
            'per': per_cell,
            'total_rec': total_rec_cell,
            'mult': mult2_cell,
            'irr': irr2_cell
        })
        row += 1

    row += 1

    # ==========================================
    # 시나리오 3: 10% 할인율 적용 (NPV)
    # ==========================================
    scenario3_start = row
    format_cell(ws, row, 1, "【시나리오 3】 할인율 10% 적용 (현재가치)", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    # 서브 헤더
    format_cell(ws, row, 1, "3-A. 2029년 전체 매각 (할인)", BOLD_FONT, SCENARIO_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    headers3a = ["PER", "회수금액", "할인기간", "NPV", "멀티플 (NPV)", "IRR (NPV)"]
    for col, h in enumerate(headers3a, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1

    for i, per in enumerate(per_multiples):
        per_cell = format_cell(ws, row, 1, per, BLUE_FONT, INPUT_FILL, '0"x"')

        # 회수금액 (시나리오 1에서 참조)
        rec_ref = s1_data_rows[i]['rec'].coordinate
        format_cell(ws, row, 2, f"={rec_ref}", number_format='#,##0"원"')

        # 할인기간
        period_formula = f"=2029-{inv_year_cell.coordinate}"
        period_cell = format_cell(ws, row, 3, period_formula, number_format='0"년"')

        # NPV = 회수금액 / (1 + 할인율)^기간
        npv_formula = f"={rec_ref}/POWER(1+{discount_cell.coordinate},{period_cell.coordinate})"
        npv_cell = format_cell(ws, row, 4, npv_formula, fill=SCENARIO_FILL, number_format='#,##0"원"')

        # NPV 멀티플
        npv_mult_formula = f"={npv_cell.coordinate}/{inv_cell.coordinate}"
        npv_mult_cell = format_cell(ws, row, 5, npv_mult_formula, fill=RESULT_FILL, number_format='0.00"x"')

        # NPV IRR (조정된 수익률)
        npv_irr_formula = f"=POWER({npv_mult_cell.coordinate},1/{period_cell.coordinate})-1"
        format_cell(ws, row, 6, npv_irr_formula, fill=RESULT_FILL, number_format='0.0%')

        row += 1

    row += 1

    # 서브 시나리오 3-B: 부분 매각 NPV
    format_cell(ws, row, 1, "3-B. 부분 매각 (할인)", BOLD_FONT, SCENARIO_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    headers3b = ["PER", "1차 회수 NPV", "2차 회수 NPV", "총 NPV", "멀티플 (NPV)", "IRR (NPV)"]
    for col, h in enumerate(headers3b, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1

    for i, per in enumerate(per_multiples):
        per_cell = format_cell(ws, row, 1, per, BLUE_FONT, INPUT_FILL, '0"x"')

        # 1차 회수 NPV (2029년, 4년 할인)
        rec1_ref = s2_data_rows[i]['total_rec'].coordinate
        # 시나리오2에서 1차 회수액 재계산
        sp_2029_calc = f"=({ni_2029_cell.coordinate}*{per_cell.coordinate})/{total_cell.coordinate}"
        rec1_calc = f"=({sp_2029_calc})*{shares_cell.coordinate}*{partial_cell.coordinate}"
        npv1_formula = f"=({rec1_calc})/POWER(1+{discount_cell.coordinate},2029-{inv_year_cell.coordinate})"
        npv1_cell = format_cell(ws, row, 2, npv1_formula, number_format='#,##0"원"')

        # 2차 회수 NPV (2030년, 5년 할인)
        sp_2030_calc = f"=({ni_2030_cell.coordinate}*{per_cell.coordinate})/{total_cell.coordinate}"
        rec2_calc = f"=({sp_2030_calc})*{shares_cell.coordinate}*(1-{partial_cell.coordinate})"
        npv2_formula = f"=({rec2_calc})/POWER(1+{discount_cell.coordinate},2030-{inv_year_cell.coordinate})"
        npv2_cell = format_cell(ws, row, 3, npv2_formula, number_format='#,##0"원"')

        # 총 NPV
        total_npv_formula = f"={npv1_cell.coordinate}+{npv2_cell.coordinate}"
        total_npv_cell = format_cell(ws, row, 4, total_npv_formula, fill=SCENARIO_FILL, number_format='#,##0"원"')

        # NPV 멀티플
        npv_mult2_formula = f"={total_npv_cell.coordinate}/{inv_cell.coordinate}"
        npv_mult2_cell = format_cell(ws, row, 5, npv_mult2_formula, fill=RESULT_FILL, number_format='0.00"x"')

        # NPV IRR
        npv_irr2_formula = f"=POWER({npv_mult2_cell.coordinate},1/{avg_period})-1"
        format_cell(ws, row, 6, npv_irr2_formula, fill=RESULT_FILL, number_format='0.0%')

        row += 1

    row += 2

    # ==========================================
    # 요약 비교표
    # ==========================================
    format_cell(ws, row, 1, "【전체 시나리오 요약】", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    summary_headers = ["PER", "전체매각 멀티플", "전체매각 IRR", "부분매각 멀티플", "부분매각 IRR",
                      "NPV 멀티플", "전략 제안"]
    for col, h in enumerate(summary_headers, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"))
    row += 1

    for i, per in enumerate(per_multiples):
        format_cell(ws, row, 1, per, BOLD_FONT, number_format='0"x"')

        # 시나리오1 데이터
        s1_mult = s1_data_rows[i]['mult'].coordinate
        s1_irr = s1_data_rows[i]['irr'].coordinate
        format_cell(ws, row, 2, f"={s1_mult}", number_format='0.00"x"')
        format_cell(ws, row, 3, f"={s1_irr}", number_format='0.0%')

        # 시나리오2 데이터
        s2_mult = s2_data_rows[i]['mult'].coordinate
        s2_irr = s2_data_rows[i]['irr'].coordinate
        format_cell(ws, row, 4, f"={s2_mult}", number_format='0.00"x"')
        format_cell(ws, row, 5, f"={s2_irr}", number_format='0.0%')

        # NPV 비교는 수식으로 계산 복잡하므로 텍스트 안내
        format_cell(ws, row, 6, "시나리오3 참조")

        # 전략 제안
        if per == 20:
            strategy = "높은 PER: 전체 매각 고려"
        elif per == 10:
            strategy = "보수적: 부분 매각 추천"
        else:
            strategy = "적정 수익률 확보 가능"
        format_cell(ws, row, 7, strategy)

        row += 1

    row += 2

    # === 범례 및 설명 ===
    format_cell(ws, row, 1, "분석 설명", BOLD_FONT)
    row += 1
    ws.cell(row=row, column=1, value="• 시나리오 1: 2029년 전체 지분 매각")
    row += 1
    ws.cell(row=row, column=1, value="• 시나리오 2: 2029년 50% 매각 + 2030년 잔여 50% 매각")
    row += 1
    ws.cell(row=row, column=1, value="• 시나리오 3: 할인율 10% 적용 현재가치(NPV) 분석")
    row += 1
    ws.cell(row=row, column=1, value="• IRR: 연평균 수익률 / 멀티플: 총 수익 배수")
    row += 2

    # 범례
    format_cell(ws, row, 1, "범례", BOLD_FONT)
    row += 1
    legend_cell = ws.cell(row=row, column=1, value="파란색 텍스트")
    legend_cell.font = BLUE_FONT
    ws.cell(row=row, column=2, value="= 입력값 (수정 가능)")
    row += 1
    ws.cell(row=row, column=1, value="노란색 배경").fill = INPUT_FILL
    ws.cell(row=row, column=2, value="= 핵심 가정")
    row += 1
    ws.cell(row=row, column=1, value="녹색 배경").fill = RESULT_FILL
    ws.cell(row=row, column=2, value="= 주요 결과값 (IRR, 멀티플)")

    # 저장
    if output_path is None:
        output_path = f"{company_name}_Advanced_Exit_프로젝션.xlsx"

    wb.save(output_path)
    print(f"✅ 고급 Exit 프로젝션 생성 완료: {output_path}")
    print(f"   - 시나리오 1: 2029년 전체 매각")
    print(f"   - 시나리오 2: 부분 매각 (2029: 50% / 2030: 50%)")
    print(f"   - 시나리오 3: 할인율 {discount_rate*100}% 적용 NPV 분석")
    return output_path

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='고급 Exit 프로젝션 엑셀 생성')
    parser.add_argument('--investment_amount', type=float, required=True)
    parser.add_argument('--price_per_share', type=float, required=True)
    parser.add_argument('--shares', type=int, required=True)
    parser.add_argument('--total_shares', type=int, required=True)
    parser.add_argument('--net_income_2029', type=float, required=True)
    parser.add_argument('--net_income_2030', type=float, required=True)
    parser.add_argument('--company_name', type=str, required=True)
    parser.add_argument('--per_multiples', type=str, default="10,15,20")
    parser.add_argument('--partial_exit_ratio', type=float, default=0.5)
    parser.add_argument('--discount_rate', type=float, default=0.10)
    parser.add_argument('--investment_year', type=int, default=2025)
    parser.add_argument('--output', '-o', type=str, default=None)

    args = parser.parse_args()
    per_list = [int(x.strip()) for x in args.per_multiples.split(',')]

    generate_advanced_exit_projection(
        investment_amount=args.investment_amount,
        price_per_share=args.price_per_share,
        shares=args.shares,
        total_shares=args.total_shares,
        net_income_2029=args.net_income_2029,
        net_income_2030=args.net_income_2030,
        company_name=args.company_name,
        per_multiples=per_list,
        partial_exit_ratio=args.partial_exit_ratio,
        discount_rate=args.discount_rate,
        investment_year=args.investment_year,
        output_path=args.output
    )
