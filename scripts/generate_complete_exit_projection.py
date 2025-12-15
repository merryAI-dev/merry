#!/usr/bin/env python3
"""
완전판 Exit 프로젝션 엑셀 생성 스크립트
- SAFE 전환 시나리오 (밸류에이션 캡)
- 콜옵션 시나리오
- 부분 매각 시나리오
- 할인율 적용 NPV
"""
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 스타일 정의
BLUE_FONT = Font(color="0000FF", bold=False)
BOLD_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INPUT_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
RESULT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
SCENARIO_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SAFE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CALL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def format_cell(ws, row, col, value=None, font=None, fill=None, number_format=None, alignment=None):
    """셀 포맷팅 헬퍼"""
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    if alignment:
        cell.alignment = alignment
    cell.border = THIN_BORDER
    return cell

def generate_complete_exit_projection(
    investment_amount,
    price_per_share,
    shares,
    total_shares_before_safe,
    net_income_2029,
    net_income_2030,
    company_name,
    per_multiples,
    safe_amount=100000000,
    safe_valuation_cap=5000000000,
    call_option_price_multiplier=1.5,
    partial_exit_ratio=0.5,
    discount_rate=0.10,
    investment_year=2025,
    output_path=None
):
    """
    완전판 Exit 프로젝션 생성

    시나리오:
    1. 기본 Exit (SAFE 전환 전)
    2. SAFE 전환 후 Exit
    3. 콜옵션 행사 시나리오
    4. 부분 매각
    5. NPV 분석
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Complete Exit 분석"

    # 컬럼 너비 설정
    for i in range(1, 15):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions['A'].width = 28

    row = 1

    # === 제목 ===
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    title_cell = ws.cell(row=row, column=1, value=f"{company_name} Complete Exit 분석 (SAFE + 콜옵션 포함)")
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    row += 2

    # === 투자 조건 섹션 ===
    format_cell(ws, row, 1, "🔷 기본 투자 조건", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    format_cell(ws, row, 1, "투자금액")
    inv_cell = format_cell(ws, row, 2, investment_amount, BLUE_FONT, INPUT_FILL, '#,##0"원"')
    row += 1

    format_cell(ws, row, 1, "투자단가")
    pps_cell = format_cell(ws, row, 2, price_per_share, BLUE_FONT, INPUT_FILL, '#,##0"원"')
    row += 1

    format_cell(ws, row, 1, "투자주식수")
    shares_cell = format_cell(ws, row, 2, shares, BLUE_FONT, INPUT_FILL, '#,##0"주"')
    row += 1

    format_cell(ws, row, 1, "총 발행주식수 (SAFE 전환 전)")
    total_before_cell = format_cell(ws, row, 2, total_shares_before_safe, BLUE_FONT, INPUT_FILL, '#,##0"주"')
    row += 1

    format_cell(ws, row, 1, "지분율 (SAFE 전환 전)")
    format_cell(ws, row, 2, f"={shares_cell.coordinate}/{total_before_cell.coordinate}", number_format='0.00%')
    row += 1

    format_cell(ws, row, 1, "투자연도")
    inv_year_cell = format_cell(ws, row, 2, investment_year, BLUE_FONT, INPUT_FILL, '0')
    row += 2

    # === SAFE 조건 섹션 ===
    format_cell(ws, row, 1, "🔶 SAFE 투자 조건", HEADER_FONT, PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    format_cell(ws, row, 1, "SAFE 투자금액")
    safe_amount_cell = format_cell(ws, row, 2, safe_amount, BLUE_FONT, SAFE_FILL, '#,##0"원"')
    row += 1

    format_cell(ws, row, 1, "밸류에이션 캡")
    safe_cap_cell = format_cell(ws, row, 2, safe_valuation_cap, BLUE_FONT, SAFE_FILL, '#,##0"원"')
    row += 1

    # SAFE 전환 주식수 계산
    format_cell(ws, row, 1, "SAFE 전환 주식수 (계산)")
    # SAFE 주식수 = (SAFE 금액 / 밸류에이션 캡) * 총 발행주식수
    safe_shares_formula = f"=({safe_amount_cell.coordinate}/{safe_cap_cell.coordinate})*{total_before_cell.coordinate}"
    safe_shares_cell = format_cell(ws, row, 2, safe_shares_formula, fill=SAFE_FILL, number_format='#,##0"주"')
    row += 1

    format_cell(ws, row, 1, "총 발행주식수 (SAFE 전환 후)")
    total_after_safe_formula = f"={total_before_cell.coordinate}+{safe_shares_cell.coordinate}"
    total_after_cell = format_cell(ws, row, 2, total_after_safe_formula, fill=SAFE_FILL, number_format='#,##0"주"')
    row += 1

    format_cell(ws, row, 1, "희석 후 지분율")
    diluted_ownership_formula = f"={shares_cell.coordinate}/{total_after_cell.coordinate}"
    format_cell(ws, row, 2, diluted_ownership_formula, fill=SAFE_FILL, number_format='0.00%')
    row += 2

    # === 콜옵션 조건 ===
    format_cell(ws, row, 1, "🔸 콜옵션 조건", HEADER_FONT, PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    format_cell(ws, row, 1, "콜옵션 행사가 배수")
    call_mult_cell = format_cell(ws, row, 2, call_option_price_multiplier, BLUE_FONT, CALL_FILL, '0.0"x"')
    row += 1

    format_cell(ws, row, 1, "콜옵션 행사가 (주당)")
    call_price_formula = f"={pps_cell.coordinate}*{call_mult_cell.coordinate}"
    call_price_cell = format_cell(ws, row, 2, call_price_formula, fill=CALL_FILL, number_format='#,##0"원"')
    row += 1

    format_cell(ws, row, 1, "콜옵션 전체 행사 금액")
    call_total_formula = f"={call_price_cell.coordinate}*{shares_cell.coordinate}"
    call_total_cell = format_cell(ws, row, 2, call_total_formula, fill=CALL_FILL, number_format='#,##0"원"')
    row += 2

    # === 순이익 가정 ===
    format_cell(ws, row, 1, "📊 순이익 가정", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    format_cell(ws, row, 1, "2029년 당기순이익")
    ni_2029_cell = format_cell(ws, row, 2, net_income_2029, BLUE_FONT, INPUT_FILL, '#,##0"원"')
    row += 1

    format_cell(ws, row, 1, "2030년 당기순이익")
    ni_2030_cell = format_cell(ws, row, 2, net_income_2030, BLUE_FONT, INPUT_FILL, '#,##0"원"')
    row += 1

    format_cell(ws, row, 1, "부분 매각 비율 (1차)")
    partial_cell = format_cell(ws, row, 2, partial_exit_ratio, BLUE_FONT, INPUT_FILL, '0%')
    row += 1

    format_cell(ws, row, 1, "할인율")
    discount_cell = format_cell(ws, row, 2, discount_rate, BLUE_FONT, INPUT_FILL, '0%')
    row += 2

    # ==========================================
    # 시나리오 1: 2029년 전체 매각 (SAFE 전환 전)
    # ==========================================
    format_cell(ws, row, 1, "【시나리오 1】 2029년 전체 매각 (SAFE 전환 전)", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    headers = ["PER", "기업가치", "주당가치", "회수금액", "멀티플", "IRR"]
    for col, h in enumerate(headers, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1

    s1_data = []
    for per in per_multiples:
        per_cell = format_cell(ws, row, 1, per, BLUE_FONT, INPUT_FILL, '0"x"')
        ev_formula = f"={ni_2029_cell.coordinate}*{per_cell.coordinate}"
        ev_cell = format_cell(ws, row, 2, ev_formula, number_format='#,##0"원"')
        sp_formula = f"={ev_cell.coordinate}/{total_before_cell.coordinate}"
        sp_cell = format_cell(ws, row, 3, sp_formula, number_format='#,##0"원"')
        rec_formula = f"={sp_cell.coordinate}*{shares_cell.coordinate}"
        rec_cell = format_cell(ws, row, 4, rec_formula, number_format='#,##0"원"')
        mult_formula = f"={rec_cell.coordinate}/{inv_cell.coordinate}"
        mult_cell = format_cell(ws, row, 5, mult_formula, fill=RESULT_FILL, number_format='0.00"x"')
        period_formula = f"=2029-{inv_year_cell.coordinate}"
        irr_formula = f"=POWER({mult_cell.coordinate},1/({period_formula}))-1"
        irr_cell = format_cell(ws, row, 6, irr_formula, fill=RESULT_FILL, number_format='0.0%')

        s1_data.append({'per': per_cell, 'sp': sp_cell, 'rec': rec_cell, 'mult': mult_cell, 'irr': irr_cell})
        row += 1
    row += 1

    # ==========================================
    # 시나리오 2: SAFE 전환 후 2029년 전체 매각
    # ==========================================
    format_cell(ws, row, 1, "【시나리오 2】 SAFE 전환 후 2029년 전체 매각", HEADER_FONT, PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    format_cell(ws, row, 1, "희석 효과 반영: 총 주식수 증가", font=Font(italic=True, size=9))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    for col, h in enumerate(headers, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1

    s2_data = []
    for per in per_multiples:
        per_cell = format_cell(ws, row, 1, per, BLUE_FONT, INPUT_FILL, '0"x"')
        ev_formula = f"={ni_2029_cell.coordinate}*{per_cell.coordinate}"
        ev_cell = format_cell(ws, row, 2, ev_formula, number_format='#,##0"원"')
        # 희석 후 주당가치
        sp_formula = f"={ev_cell.coordinate}/{total_after_cell.coordinate}"
        sp_cell = format_cell(ws, row, 3, sp_formula, fill=SAFE_FILL, number_format='#,##0"원"')
        rec_formula = f"={sp_cell.coordinate}*{shares_cell.coordinate}"
        rec_cell = format_cell(ws, row, 4, rec_formula, fill=SAFE_FILL, number_format='#,##0"원"')
        mult_formula = f"={rec_cell.coordinate}/{inv_cell.coordinate}"
        mult_cell = format_cell(ws, row, 5, mult_formula, fill=RESULT_FILL, number_format='0.00"x"')
        period_formula = f"=2029-{inv_year_cell.coordinate}"
        irr_formula = f"=POWER({mult_cell.coordinate},1/({period_formula}))-1"
        irr_cell = format_cell(ws, row, 6, irr_formula, fill=RESULT_FILL, number_format='0.0%')

        s2_data.append({'per': per_cell, 'sp': sp_cell, 'rec': rec_cell, 'mult': mult_cell, 'irr': irr_cell})
        row += 1
    row += 1

    # ==========================================
    # 시나리오 3: 콜옵션 행사
    # ==========================================
    format_cell(ws, row, 1, "【시나리오 3】 콜옵션 행사", HEADER_FONT, PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    format_cell(ws, row, 1, "회사가 투자단가 × 1.5배로 주식 매입", font=Font(italic=True, size=9))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    headers_call = ["시점", "행사가 (주당)", "회수금액", "멀티플", "투자기간", "IRR"]
    for col, h in enumerate(headers_call, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1

    # 2029년 콜옵션 행사
    format_cell(ws, row, 1, "2029년")
    format_cell(ws, row, 2, f"={call_price_cell.coordinate}", fill=CALL_FILL, number_format='#,##0"원"')
    format_cell(ws, row, 3, f"={call_total_cell.coordinate}", fill=CALL_FILL, number_format='#,##0"원"')
    call_mult_formula = f"={call_total_cell.coordinate}/{inv_cell.coordinate}"
    format_cell(ws, row, 4, call_mult_formula, fill=RESULT_FILL, number_format='0.00"x"')
    format_cell(ws, row, 5, f"=2029-{inv_year_cell.coordinate}", number_format='0"년"')
    call_irr_formula = f"=POWER({call_total_cell.coordinate}/{inv_cell.coordinate},1/(2029-{inv_year_cell.coordinate}))-1"
    format_cell(ws, row, 6, call_irr_formula, fill=RESULT_FILL, number_format='0.0%')
    row += 2

    # ==========================================
    # 시나리오 4: 부분 매각 (SAFE 전환 후)
    # ==========================================
    format_cell(ws, row, 1, "【시나리오 4】 부분 매각 (2029년 50% + 2030년 50%)", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    row += 1

    format_cell(ws, row, 1, "SAFE 전환 후 희석된 상태에서 분할 매각", font=Font(italic=True, size=9))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    row += 1

    headers4 = ["PER", "2029 주당가치", "1차 회수액", "2030 주당가치", "2차 회수액", "총 회수액", "멀티플", "복합 IRR"]
    for col, h in enumerate(headers4, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1

    s4_data = []
    for per in per_multiples:
        per_cell = format_cell(ws, row, 1, per, BLUE_FONT, INPUT_FILL, '0"x"')

        # 2029년 주당가치 (SAFE 전환 후)
        ev_2029 = f"={ni_2029_cell.coordinate}*{per_cell.coordinate}"
        sp_2029_formula = f"=({ev_2029})/{total_after_cell.coordinate}"
        sp_2029_cell = format_cell(ws, row, 2, sp_2029_formula, fill=SAFE_FILL, number_format='#,##0"원"')

        # 1차 회수액
        rec1_formula = f"={sp_2029_cell.coordinate}*{shares_cell.coordinate}*{partial_cell.coordinate}"
        rec1_cell = format_cell(ws, row, 3, rec1_formula, number_format='#,##0"원"')

        # 2030년 주당가치 (SAFE 전환 후)
        ev_2030 = f"={ni_2030_cell.coordinate}*{per_cell.coordinate}"
        sp_2030_formula = f"=({ev_2030})/{total_after_cell.coordinate}"
        sp_2030_cell = format_cell(ws, row, 4, sp_2030_formula, fill=SAFE_FILL, number_format='#,##0"원"')

        # 2차 회수액
        rec2_formula = f"={sp_2030_cell.coordinate}*{shares_cell.coordinate}*(1-{partial_cell.coordinate})"
        rec2_cell = format_cell(ws, row, 5, rec2_formula, number_format='#,##0"원"')

        # 총 회수액
        total_rec_formula = f"={rec1_cell.coordinate}+{rec2_cell.coordinate}"
        total_rec_cell = format_cell(ws, row, 6, total_rec_formula, fill=SCENARIO_FILL, number_format='#,##0"원"')

        # 멀티플
        mult4_formula = f"={total_rec_cell.coordinate}/{inv_cell.coordinate}"
        mult4_cell = format_cell(ws, row, 7, mult4_formula, fill=RESULT_FILL, number_format='0.00"x"')

        # 복합 IRR (평균 보유기간 4.5년)
        avg_period = 4.5
        irr4_formula = f"=POWER({mult4_cell.coordinate},1/{avg_period})-1"
        irr4_cell = format_cell(ws, row, 8, irr4_formula, fill=RESULT_FILL, number_format='0.0%')

        s4_data.append({'mult': mult4_cell, 'irr': irr4_cell})
        row += 1
    row += 1

    # ==========================================
    # 시나리오 5: NPV 분석
    # ==========================================
    format_cell(ws, row, 1, "【시나리오 5】 할인율 10% 적용 NPV", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    # 5-A: 전체 매각 NPV
    format_cell(ws, row, 1, "5-A. 2029년 전체 매각 NPV (SAFE 전환 후)", BOLD_FONT, SCENARIO_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    headers5 = ["PER", "회수금액", "할인기간", "NPV", "멀티플 (NPV)", "IRR (NPV)"]
    for col, h in enumerate(headers5, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1

    for i, per in enumerate(per_multiples):
        per_cell = format_cell(ws, row, 1, per, BLUE_FONT, INPUT_FILL, '0"x"')
        rec_ref = s2_data[i]['rec'].coordinate
        format_cell(ws, row, 2, f"={rec_ref}", number_format='#,##0"원"')
        period_formula = f"=2029-{inv_year_cell.coordinate}"
        period_cell = format_cell(ws, row, 3, period_formula, number_format='0"년"')
        npv_formula = f"={rec_ref}/POWER(1+{discount_cell.coordinate},{period_cell.coordinate})"
        npv_cell = format_cell(ws, row, 4, npv_formula, fill=SCENARIO_FILL, number_format='#,##0"원"')
        npv_mult_formula = f"={npv_cell.coordinate}/{inv_cell.coordinate}"
        npv_mult_cell = format_cell(ws, row, 5, npv_mult_formula, fill=RESULT_FILL, number_format='0.00"x"')
        npv_irr_formula = f"=POWER({npv_mult_cell.coordinate},1/{period_cell.coordinate})-1"
        format_cell(ws, row, 6, npv_irr_formula, fill=RESULT_FILL, number_format='0.0%')
        row += 1
    row += 1

    # 5-B: 부분 매각 NPV
    format_cell(ws, row, 1, "5-B. 부분 매각 NPV", BOLD_FONT, SCENARIO_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    headers5b = ["PER", "1차 회수 NPV", "2차 회수 NPV", "총 NPV", "멀티플 (NPV)", "IRR (NPV)"]
    for col, h in enumerate(headers5b, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1

    for i, per in enumerate(per_multiples):
        per_cell = format_cell(ws, row, 1, per, BLUE_FONT, INPUT_FILL, '0"x"')

        # 1차 NPV
        sp_2029_calc = f"=({ni_2029_cell.coordinate}*{per_cell.coordinate})/{total_after_cell.coordinate}"
        rec1_calc = f"=({sp_2029_calc})*{shares_cell.coordinate}*{partial_cell.coordinate}"
        npv1_formula = f"=({rec1_calc})/POWER(1+{discount_cell.coordinate},2029-{inv_year_cell.coordinate})"
        npv1_cell = format_cell(ws, row, 2, npv1_formula, number_format='#,##0"원"')

        # 2차 NPV
        sp_2030_calc = f"=({ni_2030_cell.coordinate}*{per_cell.coordinate})/{total_after_cell.coordinate}"
        rec2_calc = f"=({sp_2030_calc})*{shares_cell.coordinate}*(1-{partial_cell.coordinate})"
        npv2_formula = f"=({rec2_calc})/POWER(1+{discount_cell.coordinate},2030-{inv_year_cell.coordinate})"
        npv2_cell = format_cell(ws, row, 3, npv2_formula, number_format='#,##0"원"')

        # 총 NPV
        total_npv_formula = f"={npv1_cell.coordinate}+{npv2_cell.coordinate}"
        total_npv_cell = format_cell(ws, row, 4, total_npv_formula, fill=SCENARIO_FILL, number_format='#,##0"원"')

        # NPV 멀티플
        npv_mult2_formula = f"={total_npv_cell.coordinate}/{inv_cell.coordinate}"
        npv_mult2_cell = format_cell(ws, row, 5, npv_mult2_formula, fill=RESULT_FILL, number_format='0.00"x"')

        # NPV IRR
        avg_period = 4.5
        npv_irr2_formula = f"=POWER({npv_mult2_cell.coordinate},1/{avg_period})-1"
        format_cell(ws, row, 6, npv_irr2_formula, fill=RESULT_FILL, number_format='0.0%')

        row += 1
    row += 2

    # ==========================================
    # 전체 시나리오 요약표
    # ==========================================
    format_cell(ws, row, 1, "【전체 시나리오 요약 비교】", HEADER_FONT, HEADER_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    row += 1

    summary_headers = ["PER", "S1: 기본\n(멀티플)", "S1: 기본\n(IRR)",
                      "S2: SAFE후\n(멀티플)", "S2: SAFE후\n(IRR)",
                      "S3: 콜옵션\n(멀티플)", "S3: 콜옵션\n(IRR)",
                      "S4: 부분매각\n(멀티플)", "S4: 부분매각\n(IRR)", "최적 전략"]
    for col, h in enumerate(summary_headers, 1):
        c = format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"))
        c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30
    row += 1

    for i, per in enumerate(per_multiples):
        format_cell(ws, row, 1, per, BOLD_FONT, number_format='0"x"')

        # S1
        format_cell(ws, row, 2, f"={s1_data[i]['mult'].coordinate}", number_format='0.00"x"')
        format_cell(ws, row, 3, f"={s1_data[i]['irr'].coordinate}", number_format='0.0%')

        # S2
        format_cell(ws, row, 4, f"={s2_data[i]['mult'].coordinate}", number_format='0.00"x"')
        format_cell(ws, row, 5, f"={s2_data[i]['irr'].coordinate}", number_format='0.0%')

        # S3 (콜옵션은 PER 무관, 첫번째만)
        if i == 0:
            format_cell(ws, row, 6, call_mult_formula, number_format='0.00"x"')
            format_cell(ws, row, 7, call_irr_formula, number_format='0.0%')
        else:
            format_cell(ws, row, 6, "동일")
            format_cell(ws, row, 7, "동일")

        # S4
        format_cell(ws, row, 8, f"={s4_data[i]['mult'].coordinate}", number_format='0.00"x"')
        format_cell(ws, row, 9, f"={s4_data[i]['irr'].coordinate}", number_format='0.0%')

        # 전략
        if per == per_multiples[-1]:  # 가장 높은 PER
            strategy = "S1 or S2: 높은 밸류 시 전체 매각"
        elif per == per_multiples[0]:  # 가장 낮은 PER
            strategy = "S3: 콜옵션 고려 or S4: 부분 매각"
        else:
            strategy = "S4: 부분 매각 균형 전략"
        format_cell(ws, row, 10, strategy)

        row += 1
    row += 2

    # ==========================================
    # SAFE 희석 효과 분석
    # ==========================================
    format_cell(ws, row, 1, "📉 SAFE 희석 효과 분석", HEADER_FONT, PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    headers_dilution = ["구분", "SAFE 전환 전", "SAFE 전환 후", "변화", "희석률"]
    for col, h in enumerate(headers_dilution, 1):
        format_cell(ws, row, col, h, BOLD_FONT, PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"))
    row += 1

    # 총 발행주식수
    format_cell(ws, row, 1, "총 발행주식수")
    format_cell(ws, row, 2, f"={total_before_cell.coordinate}", number_format='#,##0"주"')
    format_cell(ws, row, 3, f"={total_after_cell.coordinate}", number_format='#,##0"주"')
    format_cell(ws, row, 4, f"={total_after_cell.coordinate}-{total_before_cell.coordinate}", number_format='#,##0"주"')
    dilution_shares = f"=({total_after_cell.coordinate}-{total_before_cell.coordinate})/{total_before_cell.coordinate}"
    format_cell(ws, row, 5, dilution_shares, fill=SAFE_FILL, number_format='0.00%')
    row += 1

    # 우리 지분율
    format_cell(ws, row, 1, "우리 지분율")
    ownership_before = f"={shares_cell.coordinate}/{total_before_cell.coordinate}"
    format_cell(ws, row, 2, ownership_before, number_format='0.00%')
    format_cell(ws, row, 3, diluted_ownership_formula, number_format='0.00%')
    ownership_change = f"={diluted_ownership_formula}-{ownership_before}"
    format_cell(ws, row, 4, ownership_change, number_format='0.00%p')
    ownership_dilution = f"=({ownership_before}-{diluted_ownership_formula})/{ownership_before}"
    format_cell(ws, row, 5, ownership_dilution, fill=SAFE_FILL, number_format='0.00%')
    row += 1

    # PER 15 기준 회수금액 영향
    format_cell(ws, row, 1, "회수금액 (PER 15 기준)")
    if len(per_multiples) >= 2:
        mid_idx = len(per_multiples) // 2
        format_cell(ws, row, 2, f"={s1_data[mid_idx]['rec'].coordinate}", number_format='#,##0"원"')
        format_cell(ws, row, 3, f"={s2_data[mid_idx]['rec'].coordinate}", number_format='#,##0"원"')
        rec_change = f"={s2_data[mid_idx]['rec'].coordinate}-{s1_data[mid_idx]['rec'].coordinate}"
        format_cell(ws, row, 4, rec_change, fill=SAFE_FILL, number_format='#,##0"원"')
        rec_dilution = f"=({s1_data[mid_idx]['rec'].coordinate}-{s2_data[mid_idx]['rec'].coordinate})/{s1_data[mid_idx]['rec'].coordinate}"
        format_cell(ws, row, 5, rec_dilution, fill=SAFE_FILL, number_format='0.00%')
    row += 2

    # === 설명 및 범례 ===
    format_cell(ws, row, 1, "💡 분석 가이드", BOLD_FONT)
    row += 1
    ws.cell(row=row, column=1, value="【시나리오 1】 SAFE 전환 전 기본 Exit - SAFE가 전환되지 않은 상태")
    row += 1
    ws.cell(row=row, column=1, value="【시나리오 2】 SAFE 전환 후 Exit - 밸류에이션 캡 50억으로 SAFE 전환, 지분 희석 반영")
    row += 1
    ws.cell(row=row, column=1, value="【시나리오 3】 콜옵션 행사 - 회사가 투자단가 × 1.5배로 주식 매입")
    row += 1
    ws.cell(row=row, column=1, value="【시나리오 4】 부분 매각 - 2029년 50% 매각 + 2030년 50% 매각 (SAFE 전환 후)")
    row += 1
    ws.cell(row=row, column=1, value="【시나리오 5】 NPV 분석 - 할인율 10% 적용한 현재가치 기준 분석")
    row += 2

    format_cell(ws, row, 1, "🎨 범례", BOLD_FONT)
    row += 1
    legend_cell = ws.cell(row=row, column=1, value="파란색 텍스트")
    legend_cell.font = BLUE_FONT
    ws.cell(row=row, column=2, value="= 입력값 (수정 가능)")
    row += 1
    ws.cell(row=row, column=1, value="노란색 배경").fill = INPUT_FILL
    ws.cell(row=row, column=2, value="= 핵심 가정")
    row += 1
    ws.cell(row=row, column=1, value="녹색 배경").fill = RESULT_FILL
    ws.cell(row=row, column=2, value="= 주요 결과 (멀티플, IRR)")
    row += 1
    ws.cell(row=row, column=1, value="주황색 배경").fill = SAFE_FILL
    ws.cell(row=row, column=2, value="= SAFE 관련")
    row += 1
    ws.cell(row=row, column=1, value="회색 배경").fill = CALL_FILL
    ws.cell(row=row, column=2, value="= 콜옵션 관련")

    # 저장
    if output_path is None:
        output_path = f"{company_name}_Complete_Exit_프로젝션.xlsx"

    wb.save(output_path)
    print(f"✅ Complete Exit 프로젝션 생성 완료: {output_path}")
    print(f"   - 시나리오 1: 2029년 전체 매각 (SAFE 전환 전)")
    print(f"   - 시나리오 2: SAFE 전환 후 매각 (밸류캡 {safe_valuation_cap:,}원)")
    print(f"   - 시나리오 3: 콜옵션 행사 ({call_option_price_multiplier}x)")
    print(f"   - 시나리오 4: 부분 매각 (2029: {partial_exit_ratio*100}% / 2030: {(1-partial_exit_ratio)*100}%)")
    print(f"   - 시나리오 5: 할인율 {discount_rate*100}% NPV 분석")
    print(f"   - SAFE 희석 효과 분석 포함")
    return output_path

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Complete Exit 프로젝션 (SAFE + 콜옵션)')
    parser.add_argument('--investment_amount', type=float, required=True)
    parser.add_argument('--price_per_share', type=float, required=True)
    parser.add_argument('--shares', type=int, required=True)
    parser.add_argument('--total_shares_before_safe', type=int, required=True, help='SAFE 전환 전 총 발행주식수')
    parser.add_argument('--net_income_2029', type=float, required=True)
    parser.add_argument('--net_income_2030', type=float, required=True)
    parser.add_argument('--company_name', type=str, required=True)
    parser.add_argument('--per_multiples', type=str, default="10,15,20")
    parser.add_argument('--safe_amount', type=float, default=100000000, help='SAFE 투자금액 (기본: 1억)')
    parser.add_argument('--safe_valuation_cap', type=float, default=5000000000, help='밸류에이션 캡 (기본: 50억)')
    parser.add_argument('--call_option_price_multiplier', type=float, default=1.5, help='콜옵션 행사가 배수 (기본: 1.5x)')
    parser.add_argument('--partial_exit_ratio', type=float, default=0.5)
    parser.add_argument('--discount_rate', type=float, default=0.10)
    parser.add_argument('--investment_year', type=int, default=2025)
    parser.add_argument('--output', '-o', type=str, default=None)

    args = parser.parse_args()
    per_list = [int(x.strip()) for x in args.per_multiples.split(',')]

    generate_complete_exit_projection(
        investment_amount=args.investment_amount,
        price_per_share=args.price_per_share,
        shares=args.shares,
        total_shares_before_safe=args.total_shares_before_safe,
        net_income_2029=args.net_income_2029,
        net_income_2030=args.net_income_2030,
        company_name=args.company_name,
        per_multiples=per_list,
        safe_amount=args.safe_amount,
        safe_valuation_cap=args.safe_valuation_cap,
        call_option_price_multiplier=args.call_option_price_multiplier,
        partial_exit_ratio=args.partial_exit_ratio,
        discount_rate=args.discount_rate,
        investment_year=args.investment_year,
        output_path=args.output
    )
