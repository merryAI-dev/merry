"""
SQS-driven analysis worker.

Responsibilities:
- Poll SQS for job messages (legacy) and fan-out task messages (v2)
- Legacy path: {teamId, jobId} → download all files → process entire job
- Fan-out path: {version:2, teamId, jobId, taskId, fileId} → process single file
- Concurrent processing via ThreadPoolExecutor with per-thread boto3 sessions
- Atomic progress tracking via DynamoDB counters
- Automatic CSV assembly when all tasks complete
"""

from __future__ import annotations

import csv
import gc
import io
import json
import logging
import math
import os
import random
import shutil
import threading
import time
import traceback
from concurrent.futures import Future, ThreadPoolExecutor
from decimal import Decimal
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import boto3
import boto3.session
from dotenv import load_dotenv


PROJECT_ROOT = Path(__file__).resolve().parent.parent
TEMP_ROOT = PROJECT_ROOT / "temp"

# 로컬 개발: web/.env.local을 단일 소스로 사용
load_dotenv(PROJECT_ROOT / "web" / ".env.local", override=False)

# ── Structured JSON Logging ──
_log_context = threading.local()


class _JsonFormatter(logging.Formatter):
    """Structured JSON log formatter for CloudWatch Logs Insights."""

    def format(self, record: logging.LogRecord) -> str:
        entry: Dict[str, Any] = {
            "timestamp": datetime.fromtimestamp(record.created, tz=timezone.utc)
                .isoformat().replace("+00:00", "Z"),
            "level": record.levelname,
            "logger": record.name,
            "message": record.getMessage(),
            "thread": record.threadName,
        }
        # Attach correlation context (job_id, task_id, team_id).
        ctx = getattr(_log_context, "ctx", None)
        if ctx:
            entry.update(ctx)
        if record.exc_info and record.exc_info[1]:
            entry["exception"] = self.formatException(record.exc_info)
        return json.dumps(entry, ensure_ascii=False)


def _set_log_context(**kwargs: Any) -> None:
    """Set per-thread logging context (job_id, task_id, team_id, correlation_id)."""
    _log_context.ctx = {k: v for k, v in kwargs.items() if v}


def _clear_log_context() -> None:
    """Clear per-thread logging context."""
    _log_context.ctx = None


# Use JSON formatter in production (MERRY_LOG_JSON=true).
# Default to human-readable format for local dev.
_USE_JSON_LOG = os.getenv("MERRY_LOG_JSON", "false").lower() == "true"

if _USE_JSON_LOG:
    _handler = logging.StreamHandler()
    _handler.setFormatter(_JsonFormatter())
    logging.root.handlers.clear()
    logging.root.addHandler(_handler)
    logging.root.setLevel(logging.INFO)
else:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(threadName)s] %(levelname)s %(message)s",
        datefmt="%Y-%m-%dT%H:%M:%S",
    )
log = logging.getLogger("worker")

# ── Concurrency settings ──
WORKER_CONCURRENCY = int(os.getenv("WORKER_CONCURRENCY", "5"))
# Per-file task should complete well within 5 minutes.
FANOUT_VISIBILITY_TIMEOUT = 300
# Legacy (whole-job) messages keep 15 min timeout.
LEGACY_VISIBILITY_TIMEOUT = 900
# Stale task claim threshold: re-claimable after 10 minutes.
STALE_CLAIM_SECONDS = 600
# When True, DDB Streams + Lambda handles assembly. Worker skips it.
# Set to "true" in production (after deploy_lambdas.sh). Defaults to "false"
# so local development still works with the worker-side assembly fallback.
LAMBDA_ASSEMBLY_ENABLED = os.getenv("MERRY_LAMBDA_ASSEMBLY", "false").lower() == "true"
# Circuit breaker: trip when error rate exceeds threshold after min samples.
CB_ERROR_RATE_THRESHOLD = float(os.getenv("MERRY_CB_ERROR_RATE", "0.5"))  # 50% default
CB_MIN_SAMPLES = int(os.getenv("MERRY_CB_MIN_SAMPLES", "5"))  # need at least 5 tasks done
# Job-level timeout in minutes: running fan-out jobs older than this are force-failed.
JOB_TIMEOUT_MINUTES = int(os.getenv("MERRY_JOB_TIMEOUT_MINUTES", "60"))
# Webhook URL for job completion/failure notifications (Slack compatible).
WEBHOOK_URL = os.getenv("MERRY_WEBHOOK_URL", "")
WEBHOOK_MAX_RETRIES = int(os.getenv("MERRY_WEBHOOK_MAX_RETRIES", "3"))
# CloudWatch EMF namespace (set to empty to disable metrics).
CW_NAMESPACE = os.getenv("MERRY_CW_NAMESPACE", "Merry/Worker")
# Bedrock/DDB retry parameters for _call_with_backoff.
BEDROCK_MAX_RETRIES = int(os.getenv("MERRY_BEDROCK_MAX_RETRIES", "3"))
BEDROCK_RETRY_DELAY = float(os.getenv("MERRY_BEDROCK_RETRY_DELAY", "1.5"))


# ── CloudWatch Embedded Metric Format (EMF) ──
class _JobTypeStats:
    """Per-job-type counters."""
    __slots__ = ("succeeded", "failed", "processing_ms", "input_tokens", "output_tokens")

    def __init__(self) -> None:
        self.succeeded = 0
        self.failed = 0
        self.processing_ms = 0.0
        self.input_tokens = 0
        self.output_tokens = 0


class _MetricsAccumulator:
    """Thread-safe accumulator for CloudWatch EMF metrics.

    Emits a JSON log line in EMF format that CloudWatch automatically
    extracts as custom metrics — zero API calls needed.
    """

    def __init__(self) -> None:
        self._lock = threading.Lock()
        self._tasks_succeeded = 0
        self._tasks_failed = 0
        self._total_processing_ms = 0.0
        self._empty_polls = 0
        self._nonempty_polls = 0
        self._input_tokens = 0
        self._output_tokens = 0
        self._retries = 0
        self._by_type: Dict[str, _JobTypeStats] = {}

    def record_task(
        self, succeeded: bool, elapsed_ms: float,
        input_tokens: int = 0, output_tokens: int = 0,
        job_type: str = "",
    ) -> None:
        with self._lock:
            if succeeded:
                self._tasks_succeeded += 1
            else:
                self._tasks_failed += 1
            self._total_processing_ms += elapsed_ms
            self._input_tokens += input_tokens
            self._output_tokens += output_tokens
            # Per-type breakdown.
            if job_type:
                s = self._by_type.get(job_type)
                if s is None:
                    s = _JobTypeStats()
                    self._by_type[job_type] = s
                if succeeded:
                    s.succeeded += 1
                else:
                    s.failed += 1
                s.processing_ms += elapsed_ms
                s.input_tokens += input_tokens
                s.output_tokens += output_tokens

    def record_retry(self) -> None:
        with self._lock:
            self._retries += 1

    def record_poll(self, empty: bool) -> None:
        with self._lock:
            if empty:
                self._empty_polls += 1
            else:
                self._nonempty_polls += 1

    def flush(self, in_flight: int) -> None:
        """Emit EMF log and reset counters."""
        if not CW_NAMESPACE:
            return
        with self._lock:
            succeeded = self._tasks_succeeded
            failed = self._tasks_failed
            avg_ms = (self._total_processing_ms / max(succeeded + failed, 1))
            empty = self._empty_polls
            nonempty = self._nonempty_polls
            input_tok = self._input_tokens
            output_tok = self._output_tokens
            retries = self._retries
            by_type = dict(self._by_type)
            self._tasks_succeeded = 0
            self._tasks_failed = 0
            self._total_processing_ms = 0.0
            self._empty_polls = 0
            self._nonempty_polls = 0
            self._input_tokens = 0
            self._output_tokens = 0
            self._retries = 0
            self._by_type.clear()

        # Aggregate EMF.
        emf = {
            "_aws": {
                "Timestamp": int(time.time() * 1000),
                "CloudWatchMetrics": [{
                    "Namespace": CW_NAMESPACE,
                    "Dimensions": [["Service"]],
                    "Metrics": [
                        {"Name": "TasksSucceeded", "Unit": "Count"},
                        {"Name": "TasksFailed", "Unit": "Count"},
                        {"Name": "AvgProcessingTime", "Unit": "Milliseconds"},
                        {"Name": "InFlightTasks", "Unit": "Count"},
                        {"Name": "EmptyPolls", "Unit": "Count"},
                        {"Name": "NonEmptyPolls", "Unit": "Count"},
                        {"Name": "InputTokens", "Unit": "Count"},
                        {"Name": "OutputTokens", "Unit": "Count"},
                        {"Name": "Retries", "Unit": "Count"},
                    ],
                }],
            },
            "Service": "merry-worker",
            "TasksSucceeded": succeeded,
            "TasksFailed": failed,
            "AvgProcessingTime": round(avg_ms, 1),
            "InFlightTasks": in_flight,
            "EmptyPolls": empty,
            "NonEmptyPolls": nonempty,
            "InputTokens": input_tok,
            "OutputTokens": output_tok,
            "Retries": retries,
        }
        # EMF log must be a single JSON line on stdout.
        print(json.dumps(emf), flush=True)

        # Per-job-type EMF (separate log lines for distinct dimensions).
        for jt, s in by_type.items():
            total = s.succeeded + s.failed
            if total == 0:
                continue
            jt_emf = {
                "_aws": {
                    "Timestamp": int(time.time() * 1000),
                    "CloudWatchMetrics": [{
                        "Namespace": CW_NAMESPACE,
                        "Dimensions": [["Service", "JobType"]],
                        "Metrics": [
                            {"Name": "TasksSucceeded", "Unit": "Count"},
                            {"Name": "TasksFailed", "Unit": "Count"},
                            {"Name": "AvgProcessingTime", "Unit": "Milliseconds"},
                            {"Name": "InputTokens", "Unit": "Count"},
                            {"Name": "OutputTokens", "Unit": "Count"},
                        ],
                    }],
                },
                "Service": "merry-worker",
                "JobType": jt,
                "TasksSucceeded": s.succeeded,
                "TasksFailed": s.failed,
                "AvgProcessingTime": round(s.processing_ms / total, 1),
                "InputTokens": s.input_tokens,
                "OutputTokens": s.output_tokens,
            }
            print(json.dumps(jt_emf), flush=True)


_metrics = _MetricsAccumulator()


def _env(name: str, default: Optional[str] = None, required: bool = False) -> str:
    val = os.environ.get(name) or default or ""
    if required and not val:
        raise RuntimeError(f"Missing env {name}")
    return val


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _send_webhook(
    job_id: str,
    title: str,
    status: str,
    *,
    total: int = 0,
    success: int = 0,
    failed: int = 0,
    error: str = "",
) -> None:
    """Send a Slack-compatible webhook notification with exponential backoff retry."""
    if os.getenv("PYTEST_CURRENT_TEST") or not WEBHOOK_URL:
        return
    import urllib.request
    import urllib.error

    emoji = "\u2705" if status == "succeeded" else "\u274c"
    lines = [f"{emoji} *Job {status}*: {title} (`{job_id}`)"]
    if total > 0:
        lines.append(f"  Total: {total} | Success: {success} | Failed: {failed}")
    if error:
        lines.append(f"  Error: {error[:200]}")
    payload_bytes = json.dumps({"text": "\n".join(lines)}).encode()

    for attempt in range(WEBHOOK_MAX_RETRIES):
        try:
            req = urllib.request.Request(
                WEBHOOK_URL,
                data=payload_bytes,
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            urllib.request.urlopen(req, timeout=5)
            return  # Success
        except urllib.error.HTTPError as e:
            if e.code < 500 and e.code != 429:
                log.warning("Webhook non-retryable HTTP %d for job=%s", e.code, job_id)
                return
            log.warning(
                "Webhook attempt %d/%d failed (HTTP %d) for job=%s",
                attempt + 1, WEBHOOK_MAX_RETRIES, e.code, job_id,
            )
        except Exception as e:
            log.warning(
                "Webhook attempt %d/%d failed for job=%s: %s",
                attempt + 1, WEBHOOK_MAX_RETRIES, job_id, e,
            )
        if attempt < WEBHOOK_MAX_RETRIES - 1:
            delay = min(2 ** attempt + random.uniform(0, 1), 10)
            time.sleep(delay)

    log.error("Webhook delivery failed after %d attempts for job=%s", WEBHOOK_MAX_RETRIES, job_id)


def _pk_team(team_id: str) -> str:
    return f"TEAM#{team_id}"


def _pk_team_jobs(team_id: str) -> str:
    return f"TEAM#{team_id}#JOBS"


def _sk_job(job_id: str) -> str:
    return f"JOB#{job_id}"


def _sk_created_job(created_at: str, job_id: str) -> str:
    return f"CREATED#{created_at}#JOB#{job_id}"


def _sk_file(file_id: str) -> str:
    return f"FILE#{file_id}"


def _safe_filename(name: str) -> str:
    # Keep it simple and filesystem-safe.
    cleaned = "".join(c if c.isalnum() or c in ("-", "_", ".", " ") else "_" for c in (name or "file"))
    cleaned = cleaned.strip().strip(".")
    return cleaned[:160] or "file"


def _ddb_sanitize(value: Any) -> Any:
    """
    DynamoDB (boto3) does not accept Python float types.
    Recursively convert floats to Decimal and ensure maps/lists are serializable.
    """

    if value is None:
        return None
    if isinstance(value, (str, bool, int, Decimal)):
        return value
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        # Use string to avoid binary float issues (e.g., 0.1).
        return Decimal(str(value))
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, dict):
        out: Dict[str, Any] = {}
        for k, v in value.items():
            out[str(k)] = _ddb_sanitize(v)
        return out
    if isinstance(value, (list, tuple)):
        return [_ddb_sanitize(v) for v in value]
    return str(value)


class AwsCtx:
    """Thread-safe AWS context.

    The SQS client is shared (used only by the main polling thread).
    DynamoDB and S3 clients are created per-thread via threading.local()
    because boto3 clients are NOT thread-safe.
    """

    def __init__(self) -> None:
        self.region = _env("AWS_REGION", os.environ.get("AWS_DEFAULT_REGION"), required=True)
        self.ddb_table = _env("MERRY_DDB_TABLE", required=True)
        self.bucket = _env("MERRY_S3_BUCKET", required=True)
        self.queue_url = _env("MERRY_SQS_QUEUE_URL", required=True)
        self.delete_inputs = (_env("MERRY_DELETE_INPUTS", "true").lower() != "false")

        # Tune connection pool size for concurrent workers.
        from botocore.config import Config as BotoConfig
        self._boto_config = BotoConfig(
            max_pool_connections=max(WORKER_CONCURRENCY * 2, 20),
            retries={"max_attempts": 2, "mode": "adaptive"},
        )

        # SQS client: main thread only (polling loop).
        self._sqs = boto3.client("sqs", region_name=self.region, config=self._boto_config)
        # Per-thread storage for DDB and S3 clients.
        self._local = threading.local()

    def _ensure_thread_clients(self) -> None:
        """Lazily create boto3 clients for the current thread."""
        if not hasattr(self._local, "ddb"):
            session = boto3.session.Session()
            self._local.ddb = session.resource(
                "dynamodb", region_name=self.region,
                config=self._boto_config,
            ).Table(self.ddb_table)
            self._local.s3 = session.client(
                "s3", region_name=self.region,
                config=self._boto_config,
            )
            # Low-level DDB client for TransactWriteItems.
            self._local.ddb_client = session.client(
                "dynamodb", region_name=self.region,
                config=self._boto_config,
            )

    @property
    def ddb(self):
        self._ensure_thread_clients()
        return self._local.ddb

    @property
    def ddb_client(self):
        self._ensure_thread_clients()
        return self._local.ddb_client

    @property
    def s3(self):
        self._ensure_thread_clients()
        return self._local.s3

    @property
    def sqs(self):
        return self._sqs

    def warmup(self) -> None:
        """Pre-warm connections to AWS services (TLS handshakes, DNS resolution).

        Called once at worker startup on the main thread to populate
        connection pools and fail fast if credentials/configs are invalid.
        """
        t0 = time.time()
        errors: list = []

        # DDB: describe table (fast, validates credentials + table name).
        try:
            self._ensure_thread_clients()
            self._local.ddb.meta.client.describe_table(TableName=self.ddb_table)
        except Exception as e:
            errors.append(f"DDB: {e}")

        # S3: head bucket (validates credentials + bucket name).
        try:
            self._ensure_thread_clients()
            self._local.s3.head_bucket(Bucket=self.bucket)
        except Exception as e:
            errors.append(f"S3: {e}")

        # SQS: get queue attributes (validates queue URL).
        try:
            self._sqs.get_queue_attributes(
                QueueUrl=self.queue_url,
                AttributeNames=["QueueArn"],
            )
        except Exception as e:
            errors.append(f"SQS: {e}")

        elapsed = round((time.time() - t0) * 1000)
        if errors:
            for err in errors:
                log.warning("Warmup failed: %s", err)
            log.info("Warmup completed with %d errors in %dms", len(errors), elapsed)
        else:
            log.info("Warmup OK: DDB + S3 + SQS connections verified in %dms", elapsed)


def ddb_get_item(ctx: AwsCtx, pk: str, sk: str) -> Optional[Dict[str, Any]]:
    def _get():
        return ctx.ddb.get_item(Key={"pk": pk, "sk": sk}).get("Item")
    return _call_with_backoff(_get, max_retries=2, base_delay=1.0)


def ddb_update_job(
    ctx: AwsCtx,
    team_id: str,
    job_id: str,
    *,
    status: Optional[str] = None,
    error: Optional[str] = None,
    artifacts: Optional[List[Dict[str, Any]]] = None,
    metrics: Optional[Dict[str, Any]] = None,
    usage: Optional[Dict[str, Any]] = None,
) -> None:
    pk = _pk_team(team_id)
    sk = _sk_job(job_id)

    exprs = []
    names: Dict[str, str] = {"#updated_at": "updated_at"}
    values: Dict[str, Any] = {":updated_at": _now_iso()}

    exprs.append("#updated_at = :updated_at")
    if status is not None:
        names["#status"] = "status"
        values[":status"] = status
        exprs.append("#status = :status")
    if error is not None:
        names["#error"] = "error"
        values[":error"] = error
        exprs.append("#error = :error")
    if artifacts is not None:
        names["#artifacts"] = "artifacts"
        values[":artifacts"] = _ddb_sanitize(artifacts)
        exprs.append("#artifacts = :artifacts")
    if metrics is not None:
        names["#metrics"] = "metrics"
        values[":metrics"] = _ddb_sanitize(metrics)
        exprs.append("#metrics = :metrics")
    if usage is not None:
        names["#usage"] = "usage"
        values[":usage"] = _ddb_sanitize(usage)
        exprs.append("#usage = :usage")

    # Final guard: ensure *all* ExpressionAttributeValues contain no Python floats.
    # This protects against any future fields added without explicit sanitization.
    values = _ddb_sanitize(values)

    _call_with_backoff(
        ctx.ddb.update_item,
        Key={"pk": pk, "sk": sk},
        UpdateExpression="SET " + ", ".join(exprs),
        ExpressionAttributeNames=names,
        ExpressionAttributeValues=values,
        max_retries=2, base_delay=1.0,
    )
    if status is not None:
        ddb_update_job_index_state(ctx, team_id, job_id, status=status)


def _ddb_get_job_created_at(ctx: AwsCtx, team_id: str, job_id: str) -> str:
    item = ddb_get_item(ctx, _pk_team(team_id), _sk_job(job_id)) or {}
    return str(item.get("created_at") or "")


def ddb_update_job_index_state(
    ctx: AwsCtx,
    team_id: str,
    job_id: str,
    *,
    status: Optional[str] = None,
    fanout: Optional[bool] = None,
    fanout_status: Optional[str] = None,
) -> None:
    if status is None and fanout is None and fanout_status is None:
        return

    created_at = _ddb_get_job_created_at(ctx, team_id, job_id)
    if not created_at:
        return

    exprs = ["updated_at = :updated_at"]
    names: Dict[str, str] = {}
    values: Dict[str, Any] = {":updated_at": _now_iso()}

    if status is not None:
        names["#status"] = "status"
        values[":status"] = status
        exprs.append("#status = :status")
    if fanout is not None:
        names["#fanout"] = "fanout"
        values[":fanout"] = fanout
        exprs.append("#fanout = :fanout")
    if fanout_status is not None:
        names["#fanout_status"] = "fanout_status"
        values[":fanout_status"] = fanout_status
        exprs.append("#fanout_status = :fanout_status")

    try:
        _call_with_backoff(
            ctx.ddb.update_item,
            Key={"pk": _pk_team_jobs(team_id), "sk": _sk_created_job(created_at, job_id)},
            UpdateExpression="SET " + ", ".join(exprs),
            ExpressionAttributeNames=names,
            ExpressionAttributeValues=_ddb_sanitize(values),
            ConditionExpression="attribute_exists(pk) AND attribute_exists(sk)",
            max_retries=2, base_delay=1.0,
        )
    except ctx.ddb.meta.client.exceptions.ConditionalCheckFailedException:
        return


def ddb_mark_file_deleted(ctx: AwsCtx, team_id: str, file_id: str) -> None:
    pk = _pk_team(team_id)
    sk = _sk_file(file_id)
    values = _ddb_sanitize({":deleted": "deleted", ":deleted_at": _now_iso()})
    _call_with_backoff(
        ctx.ddb.update_item,
        Key={"pk": pk, "sk": sk},
        UpdateExpression="SET #status = :deleted, deleted_at = :deleted_at",
        ExpressionAttributeNames={"#status": "status"},
        ExpressionAttributeValues=values,
        max_retries=2, base_delay=1.0,
    )


def s3_download(ctx: AwsCtx, bucket: str, key: str, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    _call_with_backoff(ctx.s3.download_file, bucket, key, str(dest), max_retries=2, base_delay=1.0)


def s3_upload(ctx: AwsCtx, bucket: str, key: str, src: Path, content_type: str) -> int:
    _call_with_backoff(
        ctx.s3.upload_file, str(src), bucket, key,
        ExtraArgs={"ContentType": content_type},
        max_retries=2, base_delay=1.0,
    )
    return int(src.stat().st_size)


def s3_delete(ctx: AwsCtx, bucket: str, key: str) -> None:
    _call_with_backoff(ctx.s3.delete_object, Bucket=bucket, Key=key, max_retries=2, base_delay=1.0)


MAX_PDF_SIZE = 100 * 1024 * 1024  # 100MB


def _validate_pdf(path: Path, original_name: str) -> None:
    """Quick pre-validation of a downloaded PDF file.

    Raises RuntimeError with a user-friendly message if the file is invalid.
    """
    size = path.stat().st_size
    if size == 0:
        raise RuntimeError(f"빈 파일입니다: {original_name}")
    if size > MAX_PDF_SIZE:
        raise RuntimeError(
            f"파일 크기 초과: {original_name} ({size / 1024 / 1024:.1f}MB > {MAX_PDF_SIZE // 1024 // 1024}MB)"
        )
    # Check PDF magic header.
    with open(path, "rb") as f:
        header = f.read(5)
    if header != b"%PDF-":
        raise RuntimeError(f"유효한 PDF 파일이 아닙니다: {original_name}")


def _job_temp_dir(team_id: str, job_id: str) -> Path:
    # Keep under repo temp/ to satisfy existing security validators.
    p = TEMP_ROOT / team_id / "jobs" / job_id
    p.mkdir(parents=True, exist_ok=True)
    return p


def _write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _sk_task(job_id: str, task_id: str) -> str:
    return f"TASK#{job_id}#{task_id}"


def _pk_team_tasks(team_id: str, job_id: str) -> str:
    return f"TEAM#{team_id}#TASKS#{job_id}"


def _is_retryable(exc: Exception) -> Tuple[bool, str]:
    """Determine if an exception is retryable and return (retryable, category).

    Categories:
    - "throttle": Bedrock/AWS rate limiting (429, 503 capacity)
    - "transient": Network timeouts, connection resets, temporary service errors (5xx)
    - "": Not retryable (4xx client errors, validation, auth)
    """
    exc_type = type(exc).__name__
    exc_str = str(exc).lower()

    # 1. Throttling (Bedrock, DDB provisioned throughput, S3 SlowDown).
    throttle_keywords = (
        "throttl", "too many requests", "rate exceeded",
        "limit exceeded", "slowdown", "provisioned throughput exceeded",
        "request rate too large", "model is overloaded",
    )
    if any(k in exc_str for k in throttle_keywords):
        return True, "throttle"

    # 2. Transient network / connection errors.
    transient_type_names = (
        "ConnectionError", "ConnectionResetError", "TimeoutError",
        "ConnectTimeoutError", "ReadTimeoutError", "EndpointConnectionError",
        "ConnectionClosedError", "BrokenPipeError",
    )
    if exc_type in transient_type_names:
        return True, "transient"

    # Bedrock streaming/timeout exceptions.
    bedrock_retryable_types = (
        "ModelStreamErrorException", "ModelTimeoutException",
        "ModelErrorException", "ServiceQuotaExceededException",
    )
    if exc_type in bedrock_retryable_types:
        return True, "throttle" if "Quota" in exc_type else "transient"

    transient_keywords = (
        "connection reset", "connection aborted", "timed out",
        "timeout", "temporary failure", "name resolution",
        "broken pipe", "connection refused", "network is unreachable",
        "internal server error", "service unavailable", "bad gateway",
        "502", "503", "504",
    )
    if any(k in exc_str for k in transient_keywords):
        return True, "transient"

    # 3. botocore specific retryable codes.
    if hasattr(exc, "response"):
        resp = getattr(exc, "response", {}) or {}
        code = resp.get("Error", {}).get("Code", "")
        retryable_codes = {
            "InternalServerError", "ServiceUnavailable", "RequestTimeout",
            "RequestLimitExceeded", "ProvisionedThroughputExceededException",
            "TransactionConflictException", "ItemCollectionSizeLimitExceededException",
            "ThrottlingException", "TooManyRequestsException",
            "ModelStreamErrorException", "ModelTimeoutException",
            "ServiceQuotaExceededException", "LimitExceededException",
        }
        if code in retryable_codes:
            is_throttle = any(k in code for k in ("Throughput", "Limit", "Throttl", "Quota", "TooMany"))
            return True, "throttle" if is_throttle else "transient"

        # HTTP status code based classification.
        http_status = resp.get("ResponseMetadata", {}).get("HTTPStatusCode", 0)
        if http_status == 429:
            return True, "throttle"
        if http_status in (500, 502, 503, 504):
            return True, "transient"

    return False, ""


def _call_with_backoff(
    fn, *args, max_retries: int = 3, base_delay: float = 2.0, **kwargs
) -> Any:
    """Call fn with exponential backoff + jitter for retryable errors.

    Retries on:
    - Bedrock/DDB/S3 throttling (rate limiting)
    - Transient network errors (timeouts, connection resets, DNS failures)
    - AWS service errors (5xx, InternalServerError, ServiceUnavailable)
    """
    for attempt in range(max_retries + 1):
        try:
            return fn(*args, **kwargs)
        except Exception as exc:
            retryable, category = _is_retryable(exc)
            if not retryable or attempt >= max_retries:
                raise
            delay = base_delay * (2 ** attempt) + random.uniform(0, 1)
            log.warning(
                "Retryable error [%s] (attempt %d/%d), retrying in %.1fs: %s",
                category, attempt + 1, max_retries, delay, exc,
            )
            _metrics.record_retry()
            time.sleep(delay)


def ddb_claim_task(
    ctx: AwsCtx, team_id: str, job_id: str, task_id: str, worker_id: str,
) -> bool:
    """Conditionally claim a task: pending → processing.

    Also reclaims stale tasks stuck in 'processing' for > STALE_CLAIM_SECONDS.
    Returns True if claimed successfully, False if already claimed by another worker.
    """
    now = _now_iso()
    pk = _pk_team(team_id)
    sk = _sk_task(job_id, task_id)
    try:
        _call_with_backoff(
            ctx.ddb.update_item,
            Key={"pk": pk, "sk": sk},
            UpdateExpression="SET #status = :processing, #started_at = :now, #worker_id = :wid, #updated_at = :now",
            ConditionExpression="#status = :pending",
            ExpressionAttributeNames={
                "#status": "status",
                "#started_at": "started_at",
                "#worker_id": "worker_id",
                "#updated_at": "updated_at",
            },
            ExpressionAttributeValues={
                ":pending": "pending",
                ":processing": "processing",
                ":now": now,
                ":wid": worker_id,
            },
            max_retries=2, base_delay=1.0,
        )
        return True
    except ctx.ddb.meta.client.exceptions.ConditionalCheckFailedException:
        pass

    # Try to reclaim stale task.
    stale_cutoff = datetime.now(timezone.utc).timestamp() - STALE_CLAIM_SECONDS
    stale_iso = datetime.fromtimestamp(stale_cutoff, tz=timezone.utc).isoformat().replace("+00:00", "Z")
    try:
        _call_with_backoff(
            ctx.ddb.update_item,
            Key={"pk": pk, "sk": sk},
            UpdateExpression="SET #status = :processing, #started_at = :now, #worker_id = :wid, #updated_at = :now",
            ConditionExpression="#status = :processing AND #started_at < :stale",
            ExpressionAttributeNames={
                "#status": "status",
                "#started_at": "started_at",
                "#worker_id": "worker_id",
                "#updated_at": "updated_at",
            },
            ExpressionAttributeValues={
                ":processing": "processing",
                ":now": now,
                ":wid": worker_id,
                ":stale": stale_iso,
            },
            max_retries=2, base_delay=1.0,
        )
        log.info("Reclaimed stale task %s/%s", job_id, task_id)
        return True
    except ctx.ddb.meta.client.exceptions.ConditionalCheckFailedException:
        return False


def ddb_complete_task(
    ctx: AwsCtx,
    team_id: str,
    job_id: str,
    task_id: str,
    *,
    succeeded: bool,
    result: Optional[Dict[str, Any]] = None,
    error: Optional[str] = None,
) -> None:
    """Atomically mark task complete and increment job counter.

    Uses TransactWriteItems for consistency:
    1. Update TASK status + result
    2. Increment JOB processed_count (and failed_count if failed)
    """
    now = _now_iso()
    status = "succeeded" if succeeded else "failed"
    pk = _pk_team(team_id)

    task_update: Dict[str, Any] = {
        "Update": {
            "TableName": ctx.ddb_table,
            "Key": {"pk": {"S": pk}, "sk": {"S": _sk_task(job_id, task_id)}},
            "UpdateExpression": "SET #status = :status, #ended_at = :now, #updated_at = :now",
            "ExpressionAttributeNames": {
                "#status": "status",
                "#ended_at": "ended_at",
                "#updated_at": "updated_at",
            },
            "ExpressionAttributeValues": {
                ":status": {"S": status},
                ":now": {"S": now},
            },
        }
    }
    if result:
        task_update["Update"]["UpdateExpression"] += ", #result = :result"
        task_update["Update"]["ExpressionAttributeNames"]["#result"] = "result"
        # Serialize result as JSON string to avoid DDB type issues.
        # Cap at 350KB to stay well under DDB 400KB item limit.
        result_json = json.dumps(result, ensure_ascii=False, default=str)
        if len(result_json) > 350_000:
            # Truncate large results — keep essential fields, drop evidence/text.
            truncated = {k: v for k, v in result.items() if k not in ("extracted", "readable_text")}
            truncated["_truncated"] = True
            result_json = json.dumps(truncated, ensure_ascii=False, default=str)
            if len(result_json) > 350_000:
                result_json = json.dumps({"_truncated": True, "error": "Result too large"})
        task_update["Update"]["ExpressionAttributeValues"][":result"] = {
            "S": result_json
        }
    if error:
        task_update["Update"]["UpdateExpression"] += ", #error = :error"
        task_update["Update"]["ExpressionAttributeNames"]["#error"] = "error"
        task_update["Update"]["ExpressionAttributeValues"][":error"] = {"S": error[:2000]}

    # Also update the TASK index record.
    task_index_update: Dict[str, Any] = {
        "Update": {
            "TableName": ctx.ddb_table,
            "Key": {
                "pk": {"S": _pk_team_tasks(team_id, job_id)},
                "sk": {"S": f"TASK#{task_id}"},
            },
            "UpdateExpression": "SET #status = :status, #updated_at = :now",
            "ExpressionAttributeNames": {
                "#status": "status",
                "#updated_at": "updated_at",
            },
            "ExpressionAttributeValues": {
                ":status": {"S": status},
                ":now": {"S": now},
            },
        }
    }

    # Job counter increment.
    job_update_expr = "ADD #processed_count :one"
    job_update_names: Dict[str, str] = {"#processed_count": "processed_count"}
    job_update_values: Dict[str, Any] = {":one": {"N": "1"}}
    if not succeeded:
        job_update_expr += ", #failed_count :one"
        job_update_names["#failed_count"] = "failed_count"
    job_update_expr += " SET #updated_at = :now"
    job_update_names["#updated_at"] = "updated_at"
    job_update_values[":now"] = {"S": now}

    job_counter: Dict[str, Any] = {
        "Update": {
            "TableName": ctx.ddb_table,
            "Key": {"pk": {"S": pk}, "sk": {"S": _sk_job(job_id)}},
            "UpdateExpression": job_update_expr,
            "ExpressionAttributeNames": job_update_names,
            "ExpressionAttributeValues": job_update_values,
        }
    }

    _call_with_backoff(
        ctx.ddb_client.transact_write_items,
        TransactItems=[task_update, task_index_update, job_counter],
        max_retries=BEDROCK_MAX_RETRIES, base_delay=BEDROCK_RETRY_DELAY,
    )


def ddb_try_claim_assembly(ctx: AwsCtx, team_id: str, job_id: str) -> bool:
    """Try to transition fanout_status: running → assembling. Returns True if claimed."""
    try:
        _call_with_backoff(
            ctx.ddb.update_item,
            Key={"pk": _pk_team(team_id), "sk": _sk_job(job_id)},
            UpdateExpression="SET #fs = :assembling, #updated_at = :now",
            ConditionExpression="#fs = :running",
            ExpressionAttributeNames={
                "#fs": "fanout_status",
                "#updated_at": "updated_at",
            },
            ExpressionAttributeValues={
                ":running": "running",
                ":assembling": "assembling",
                ":now": _now_iso(),
            },
            max_retries=2, base_delay=1.0,
        )
        ddb_update_job_index_state(ctx, team_id, job_id, fanout=True, fanout_status="assembling")
        return True
    except ctx.ddb.meta.client.exceptions.ConditionalCheckFailedException:
        return False


def _check_circuit_breaker(ctx: AwsCtx, team_id: str, job_id: str) -> bool:
    """Check if the error rate exceeds the threshold and trip the circuit breaker.

    Returns True if the breaker tripped (job cancelled).
    """
    if CB_MIN_SAMPLES <= 0:
        return False  # disabled

    job = ddb_get_item(ctx, _pk_team(team_id), _sk_job(job_id))
    if not job:
        return False

    total = int(job.get("total_tasks") or 0)
    processed = int(job.get("processed_count") or 0)
    failed = int(job.get("failed_count") or 0)
    fanout_status = str(job.get("fanout_status") or "")

    # Only check while the job is running.
    if fanout_status != "running":
        return False

    # Need minimum sample size before checking.
    if processed < CB_MIN_SAMPLES:
        return False

    error_rate = failed / processed if processed > 0 else 0
    if error_rate < CB_ERROR_RATE_THRESHOLD:
        return False

    # Trip the breaker!
    log.warning(
        "Circuit breaker tripped: job=%s failed=%d/%d (%.0f%%) threshold=%.0f%%",
        job_id, failed, processed, error_rate * 100, CB_ERROR_RATE_THRESHOLD * 100,
    )

    # Try to claim cancellation (running → failed).
    try:
        _call_with_backoff(
            ctx.ddb.update_item,
            Key={"pk": _pk_team(team_id), "sk": _sk_job(job_id)},
            UpdateExpression=(
                "SET #status = :failed, #fs = :failed, "
                "#error = :error, #updated_at = :now"
            ),
            ConditionExpression="#fs = :running",
            ExpressionAttributeNames={
                "#status": "status",
                "#fs": "fanout_status",
                "#error": "error",
                "#updated_at": "updated_at",
            },
            ExpressionAttributeValues={
                ":running": "running",
                ":failed": "failed",
                ":error": f"Circuit breaker: {failed}/{processed} tasks failed ({error_rate:.0%})",
                ":now": _now_iso(),
            },
            max_retries=2, base_delay=1.0,
        )
        ddb_update_job_index_state(ctx, team_id, job_id, status="failed", fanout=True, fanout_status="failed")
        log.info("Job %s cancelled by circuit breaker", job_id)
        _send_webhook(
            job_id, f"Job {job_id}", "failed",
            total=total, success=processed - failed, failed=failed,
            error=f"Circuit breaker: {failed}/{processed} tasks failed ({error_rate:.0%})",
        )
        return True
    except ctx.ddb.meta.client.exceptions.ConditionalCheckFailedException:
        # Another worker already changed the status.
        return False


def _job_timeout_watchdog(ctx: AwsCtx) -> None:
    """Scan for fan-out jobs stuck in 'running' beyond JOB_TIMEOUT_MINUTES and force-fail them.

    This runs periodically on the main thread (not in the worker pool).
    It does a Scan with a filter for running fan-out jobs — admin-only cost,
    bounded to a few pages.
    """
    if JOB_TIMEOUT_MINUTES <= 0:
        return  # Disabled.

    timeout_seconds = JOB_TIMEOUT_MINUTES * 60
    cutoff = datetime.now(timezone.utc).timestamp() - timeout_seconds
    cutoff_iso = datetime.fromtimestamp(cutoff, tz=timezone.utc).isoformat().replace("+00:00", "Z")

    try:
        # Scan for running fan-out jobs (limited to 2 pages to cap cost).
        timed_out_jobs: List[Dict[str, Any]] = []
        last_key: Optional[Dict[str, Any]] = None
        max_pages = 2

        for _ in range(max_pages):
            scan_kwargs: Dict[str, Any] = {
                "FilterExpression": (
                    "#entity = :job AND #fanout = :true "
                    "AND #fanout_status = :running "
                    "AND #created_at < :cutoff"
                ),
                "ExpressionAttributeNames": {
                    "#entity": "entity",
                    "#fanout": "fanout",
                    "#fanout_status": "fanout_status",
                    "#created_at": "created_at",
                },
                "ExpressionAttributeValues": {
                    ":job": "job",
                    ":true": True,
                    ":running": "running",
                    ":cutoff": cutoff_iso,
                },
                "ProjectionExpression": "pk, job_id, title, total_tasks, processed_count, failed_count, created_at",
            }
            if last_key:
                scan_kwargs["ExclusiveStartKey"] = last_key

            resp = _call_with_backoff(
                ctx.ddb.scan, **scan_kwargs,
                max_retries=1, base_delay=1.0,
            )
            for item in resp.get("Items") or []:
                timed_out_jobs.append(item)
            last_key = resp.get("LastEvaluatedKey")
            if not last_key:
                break

        # Force-fail each timed-out job.
        for item in timed_out_jobs:
            team_id = str(item.get("pk", "")).replace("TEAM#", "")
            job_id = str(item.get("job_id", ""))
            title = str(item.get("title", job_id))
            total = int(item.get("total_tasks") or 0)
            processed = int(item.get("processed_count") or 0)
            failed = int(item.get("failed_count") or 0)
            created = str(item.get("created_at", ""))

            log.warning(
                "Job timeout watchdog: job=%s team=%s created=%s processed=%d/%d (timeout=%dm)",
                job_id, team_id, created, processed, total, JOB_TIMEOUT_MINUTES,
            )

            try:
                _call_with_backoff(
                    ctx.ddb.update_item,
                    Key={"pk": _pk_team(team_id), "sk": _sk_job(job_id)},
                    UpdateExpression=(
                        "SET #status = :failed, #fs = :failed, "
                        "#error = :error, #updated_at = :now"
                    ),
                    ConditionExpression="#fs = :running",
                    ExpressionAttributeNames={
                        "#status": "status",
                        "#fs": "fanout_status",
                        "#error": "error",
                        "#updated_at": "updated_at",
                    },
                    ExpressionAttributeValues={
                        ":running": "running",
                        ":failed": "failed",
                        ":error": (
                            f"Job timeout: {JOB_TIMEOUT_MINUTES}분 초과 "
                            f"(processed {processed}/{total}, failed {failed})"
                        ),
                        ":now": _now_iso(),
                    },
                    max_retries=2, base_delay=1.0,
                )
                ddb_update_job_index_state(ctx, team_id, job_id, status="failed", fanout=True, fanout_status="failed")
                log.info("Job %s force-failed by timeout watchdog", job_id)
                _send_webhook(
                    job_id, title, "failed",
                    total=total, success=processed - failed, failed=failed,
                    error=f"Job timeout: {JOB_TIMEOUT_MINUTES}분 초과",
                )
            except Exception:
                # ConditionalCheckFailedException or other — already handled elsewhere.
                pass

        if timed_out_jobs:
            log.info("Watchdog processed %d timed-out jobs", len(timed_out_jobs))

    except Exception as e:
        log.warning("Job timeout watchdog error: %s", e)


def ddb_query_all_tasks(
    ctx: AwsCtx, team_id: str, job_id: str,
) -> List[Dict[str, Any]]:
    """Query all task records for a job using the TASK index."""
    pk = _pk_team(team_id)
    sk_prefix = f"TASK#{job_id}#"
    items: List[Dict[str, Any]] = []
    kwargs: Dict[str, Any] = {
        "KeyConditionExpression": "pk = :pk AND begins_with(sk, :prefix)",
        "ExpressionAttributeValues": {":pk": pk, ":prefix": sk_prefix},
    }
    while True:
        resp = _call_with_backoff(ctx.ddb.query, max_retries=2, base_delay=1.0, **kwargs)
        items.extend(resp.get("Items", []))
        if "LastEvaluatedKey" not in resp:
            break
        kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]
    return items


def _handle_exit_projection(input_path: Path, params: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    from agent.tools.extraction_tools import execute_analyze_and_generate_projection

    def _to_int(v: Any) -> Optional[int]:
        if v is None:
            return None
        try:
            if isinstance(v, Decimal):
                return int(v)
            if isinstance(v, bool):
                return None
            if isinstance(v, (int,)):
                return int(v)
            if isinstance(v, float):
                if not math.isfinite(v):
                    return None
                return int(v)
            if isinstance(v, str) and v.strip():
                return int(float(v.strip()))
        except Exception:
            return None
        return None

    target_year = _to_int(params.get("targetYear") or params.get("target_year")) or 2030
    per_multiples = params.get("perMultiples") or params.get("per_multiples") or params.get("perMultiples")
    if not isinstance(per_multiples, list) or not per_multiples:
        per_multiples = [10, 20, 30]
    per = [float(x) for x in per_multiples]

    result = execute_analyze_and_generate_projection(
        excel_path=str(input_path),
        target_year=target_year,
        per_multiples=per,
        company_name=str(params.get("companyName") or params.get("company_name") or "").strip() or None,
        output_filename=f"exit_projection_{target_year}_{int(time.time())}.xlsx",
        investment_year=_to_int(params.get("investmentYear") or params.get("investment_year")),
        investment_amount=_to_int(params.get("investmentAmount") or params.get("investment_amount")),
        price_per_share=_to_int(params.get("pricePerShare") or params.get("price_per_share")),
        shares=_to_int(params.get("shares")),
        total_shares=_to_int(params.get("totalShares") or params.get("total_shares")),
        net_income=_to_int(params.get("netIncomeTargetYear") or params.get("net_income_target_year") or params.get("net_income")),
    )
    if not result.get("success"):
        raise RuntimeError(result.get("error") or "Exit projection failed")

    output_path = Path(str(result.get("output_file") or "")).resolve()
    if not output_path.exists():
        raise RuntimeError("Exit projection output missing")

    artifacts = [
        {
            "artifactId": "exit_projection_xlsx",
            "label": "Exit 프로젝션 (XLSX)",
            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "localPath": str(output_path),
        }
    ]
    metrics = {"projection_summary": result.get("projection_summary"), "assumptions": result.get("assumptions")}
    return artifacts, metrics


def _handle_diagnosis(input_path: Path) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    from agent.tools.diagnosis_tools import execute_analyze_company_diagnosis_sheet

    result = execute_analyze_company_diagnosis_sheet(str(input_path))
    if not result.get("success"):
        raise RuntimeError(result.get("error") or "Diagnosis analysis failed")

    out_path = input_path.parent / "diagnosis_result.json"
    _write_json(out_path, result)
    artifacts = [
        {
            "artifactId": "diagnosis_json",
            "label": "기업진단 분석 결과 (JSON)",
            "contentType": "application/json",
            "localPath": str(out_path),
        }
    ]
    metrics = {"page": None}
    return artifacts, metrics


def _handle_pdf_evidence(input_path: Path, params: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    from agent.tools.underwriter_tools import execute_extract_pdf_market_evidence

    max_pages = int(params.get("maxPages") or 30)
    max_results = int(params.get("maxResults") or 20)
    result = execute_extract_pdf_market_evidence(
        pdf_path=str(input_path),
        max_pages=max_pages,
        max_results=max_results,
        keywords=None,
    )
    if not result.get("success"):
        raise RuntimeError(result.get("error") or "PDF evidence extraction failed")

    out_path = input_path.parent / "pdf_evidence.json"
    _write_json(out_path, result)
    artifacts = [
        {
            "artifactId": "pdf_evidence_json",
            "label": "PDF 근거 추출 결과 (JSON)",
            "contentType": "application/json",
            "localPath": str(out_path),
        }
    ]
    metrics = {"pages_read": result.get("pages_read"), "evidence_count": result.get("evidence_count")}
    return artifacts, metrics


def _handle_pdf_parse(input_path: Path, params: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    from agent.tools.pdf_tools import execute_read_pdf_as_text

    max_pages = int(params.get("maxPages") or 30)
    output_mode = str(params.get("outputMode") or params.get("output_mode") or "structured").strip()
    if output_mode not in ("text_only", "structured", "tables_only"):
        output_mode = "structured"

    result = execute_read_pdf_as_text(
        pdf_path=str(input_path),
        max_pages=max_pages,
        output_mode=output_mode,
        extract_financial_tables=True,
    )
    if not result.get("success"):
        raise RuntimeError(result.get("error") or "PDF parse failed")

    out_path = input_path.parent / "pdf_parse.json"
    _write_json(out_path, result)
    artifacts = [
        {
            "artifactId": "pdf_parse_json",
            "label": "PDF 파싱 결과 (JSON)",
            "contentType": "application/json",
            "localPath": str(out_path),
        }
    ]

    metrics: Dict[str, Any] = {
        "doc_type": result.get("doc_type"),
        "pages_read": result.get("pages_read"),
        "total_pages": result.get("total_pages"),
        "processing_method": result.get("processing_method"),
        "processing_time_seconds": result.get("processing_time_seconds"),
        "output_mode": output_mode,
    }

    usage = result.get("usage") if isinstance(result.get("usage"), dict) else {}
    model = result.get("model") if isinstance(result.get("model"), str) else ""
    provider = result.get("provider") if isinstance(result.get("provider"), str) else ""
    if usage or model or provider:
        llm = {**usage}
        if model:
            llm["model"] = model
        if provider:
            llm["provider"] = provider
        metrics["llm"] = llm

    return artifacts, metrics


def _handle_condition_check(
    input_paths: List[Path],
    params: Dict[str, Any],
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    조건 충족 여부 검사 (1-N 문서, 동일 조건 세트).

    params:
        conditions: list[str]  — 검사할 조건 목록
        model_id:   str        — (선택) Nova 모델 ID
        region:     str        — (선택) AWS 리전
    """
    import time as _time
    from ralph.playground_parser import (
        extract_text, assess_text_quality,
        render_first_page, render_pages, analyze_pages,
        call_nova_visual, build_presentation_prompt, _PROMPT_OCR,
    )
    from ralph.condition_checker import check_conditions_nova

    conditions: List[str] = [str(c) for c in (params.get("conditions") or []) if c]
    if not conditions:
        raise RuntimeError("conditions 파라미터가 비어 있습니다")

    model_id = str(params.get("model_id") or os.getenv("RALPH_VLM_NOVA_MODEL_ID", "us.amazon.nova-pro-v1:0"))
    model_lite = os.getenv("RALPH_VLM_NOVA_LITE_MODEL_ID", "us.amazon.nova-lite-v1:0")
    region = str(params.get("region") or os.getenv("RALPH_VLM_NOVA_REGION", "us-east-1"))
    default_use_vlm = "false" if os.getenv("PYTEST_CURRENT_TEST") else "true"
    use_vlm = os.getenv("RALPH_USE_VLM", default_use_vlm).lower() != "false"

    rows: List[Dict[str, Any]] = []
    for pdf_path in input_paths:
        t0 = _time.time()
        entry: Dict[str, Any] = {"filename": pdf_path.name}
        try:
            text, pages, blocks = extract_text(str(pdf_path))
            _, is_poor, is_fragmented = assess_text_quality(text, blocks, page_count=pages)

            extracted = ""
            method = "pymupdf"
            if use_vlm and is_poor:
                img = render_first_page(str(pdf_path))
                vd = call_nova_visual(img, model_lite, region, _PROMPT_OCR)
                extracted = vd.get("readable_text") or ""
                method = "nova_hybrid"
            elif use_vlm and is_fragmented:
                imgs = render_pages(str(pdf_path), max_pages=10, dpi=100)
                info = analyze_pages(str(pdf_path), max_pages=10)
                prompt = build_presentation_prompt(info)
                vd = call_nova_visual(imgs, model_id, region, prompt, max_tokens=5000)
                extracted = vd.get("readable_text") or ""
                method = "nova_presentation"

            full_text = "\n\n".join(filter(None, [extracted, text]))
            check = check_conditions_nova(full_text, conditions, model_id, region)

            entry.update({
                "method": method,
                "pages": pages,
                "company_name": check.get("company_name"),
                "conditions": check.get("conditions", []),
                "parse_warning": check.get("parse_warning"),
                "raw_response": check.get("raw_response"),
                "elapsed_s": round(_time.time() - t0, 1),
            })
        except Exception as exc:
            entry.update({"error": str(exc), "elapsed_s": round(_time.time() - t0, 1)})
        rows.append(entry)

    csv_path, json_path = _build_condition_check_csv(input_paths[0].parent, rows, conditions)

    artifacts: List[Dict[str, Any]] = []

    # Generate XLSX (preferred download format).
    try:
        xlsx_path = _build_condition_check_xlsx(input_paths[0].parent, rows, conditions)
        artifacts.append({
            "artifactId": "condition_check_xlsx",
            "label": "조건 검사 결과 (Excel)",
            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "localPath": str(xlsx_path),
        })
    except Exception as e:
        log.warning("Legacy XLSX generation failed: %s", e)

    artifacts.extend([
        {
            "artifactId": "condition_check_csv",
            "label": "조건 검사 결과 (CSV)",
            "contentType": "text/csv",
            "localPath": str(csv_path),
        },
        {
            "artifactId": "condition_check_json",
            "label": "조건 검사 결과 (JSON)",
            "contentType": "application/json",
            "localPath": str(json_path),
        },
    ])
    success = sum(1 for r in rows if "error" not in r)
    warning_count = sum(1 for r in rows if r.get("parse_warning"))
    metrics: Dict[str, Any] = {
        "total": len(rows),
        "success_count": success,
        "failed_count": len(rows) - success,
        "warning_count": warning_count,
        "conditions": conditions,
    }
    return artifacts, metrics


def _handle_document_extraction(
    input_paths: List[Path],
    file_rows: List[Dict[str, Any]],
    params: Dict[str, Any],
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """다건 문서 일괄 추출 (HITL 파이프라인)."""
    from ralph.batch_pipeline import process_batch, BatchDocInput

    # params.typeMap: { file_id: doc_type } — 사용자가 HITL에서 확정한 문서 타입
    type_map = params.get("typeMap") or {}

    docs: List[BatchDocInput] = []
    for path, row in zip(input_paths, file_rows):
        file_id = str(row.get("file_id") or path.stem)
        filename = str(row.get("original_name") or path.name)
        doc_type = str(type_map.get(file_id, "unknown"))
        if doc_type == "unknown":
            continue
        docs.append(BatchDocInput(
            file_id=file_id,
            filename=filename,
            pdf_path=str(path),
            doc_type=doc_type,
        ))

    if not docs:
        raise RuntimeError("No documents with valid type mapping")

    output_dir = str(input_paths[0].parent / "extraction_results")
    batch_result = process_batch(docs, output_dir=output_dir)

    artifacts: List[Dict[str, Any]] = []

    # 개별 JSON 결과
    for r in batch_result.results:
        json_file = Path(output_dir) / f"{r.file_id}.json"
        if json_file.exists():
            artifacts.append({
                "artifactId": f"extraction_{r.file_id}",
                "label": f"{r.filename} 추출 결과",
                "contentType": "application/json",
                "localPath": str(json_file),
            })

    # ZIP 번들
    if batch_result.zip_path and Path(batch_result.zip_path).exists():
        artifacts.append({
            "artifactId": "all_results_zip",
            "label": "전체 추출 결과 (ZIP)",
            "contentType": "application/zip",
            "localPath": batch_result.zip_path,
        })

    metrics: Dict[str, Any] = {
        "total": batch_result.total,
        "success_count": batch_result.success_count,
        "failed_count": batch_result.failed_count,
        "total_elapsed_seconds": batch_result.total_elapsed_seconds,
        "results_summary": [
            {
                "file_id": r.file_id,
                "filename": r.filename,
                "doc_type": r.doc_type,
                "success": r.success,
                "confidence": float(r.confidence),
                "method": r.method,
            }
            for r in batch_result.results
        ],
    }
    return artifacts, metrics


def _handle_contract_review_single(input_path: Path) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    raise NotImplementedError("Use _handle_contract_review instead")


def _handle_contract_review(
    input_paths: List[Path], params: Dict[str, Any]
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Contract review job.

    - Supports 1 or 2 documents:
      - [0] Term Sheet
      - [1] Investment agreement (optional)
    - Keeps the existing extraction/compare logic in shared.contract_review.
    - OCR is OFF by default (can be enabled later via params).
    """

    from shared.contract_review import (
        build_review_opinion,
        compare_fields,
        detect_clauses,
        extract_fields,
        load_document,
    )

    raw_mode = params.get("ocrMode") or params.get("ocr_mode") or "off"
    ocr_mode = str(raw_mode).strip().lower() if raw_mode is not None else "off"
    if ocr_mode not in ("off", "auto", "force"):
        ocr_mode = "off"

    def _parse(path: Path, doc_type: str) -> Tuple[Dict[str, Any], Dict[str, Any]]:
        loaded = load_document(Path(path), ocr_mode=ocr_mode, api_key="")
        segments = loaded.get("segments", [])
        fields = extract_fields(segments)
        clauses = detect_clauses(segments, doc_type, loaded.get("segment_scores", {}))
        doc: Dict[str, Any] = {
            "name": path.name,
            "doc_type": doc_type,
            "fields": fields,
            "clauses": clauses,
            "page_count": loaded.get("page_count", 0),
            "ocr_used": loaded.get("ocr_used", False),
            "ocr_error": loaded.get("ocr_error", ""),
            "ocr_engine": loaded.get("ocr_engine", ""),
            "priority_segments": loaded.get("priority_segments", []),
        }
        return doc, loaded

    term_doc: Optional[Dict[str, Any]] = None
    invest_doc: Optional[Dict[str, Any]] = None
    term_loaded: Dict[str, Any] = {}
    invest_loaded: Dict[str, Any] = {}

    if input_paths:
        term_doc, term_loaded = _parse(input_paths[0], "term_sheet")
    if len(input_paths) >= 2:
        invest_doc, invest_loaded = _parse(input_paths[1], "investment_agreement")

    comparisons: List[Dict[str, Any]] = []
    if term_doc and invest_doc:
        comparisons = compare_fields(term_doc.get("fields", {}), invest_doc.get("fields", {}))

    opinion = build_review_opinion(term_doc, invest_doc, comparisons)
    out: Dict[str, Any] = {
        "term_sheet": term_doc,
        "investment_agreement": invest_doc,
        "comparisons": comparisons,
        "opinion": opinion,
        "ocr_mode": ocr_mode,
    }

    out_path = input_paths[0].parent / "contract_review.json"
    _write_json(out_path, out)

    artifacts = [
        {
            "artifactId": "contract_review_json",
            "label": "계약서 검토 결과 (JSON)",
            "contentType": "application/json",
            "localPath": str(out_path),
        }
    ]
    metrics = {
        "term_sheet_pages": term_loaded.get("page_count", 0) if term_loaded else 0,
        "investment_pages": invest_loaded.get("page_count", 0) if invest_loaded else 0,
        "comparison_count": len(comparisons),
        "ocr_mode": ocr_mode,
    }
    return artifacts, metrics


def _generate_worker_id() -> str:
    """Generate a unique worker ID for task claim tracking."""
    import socket
    host = socket.gethostname()[:16]
    tid = threading.current_thread().name
    return f"{host}-{tid}-{int(time.time())}"


def process_fanout_task(
    ctx: AwsCtx,
    team_id: str,
    job_id: str,
    task_id: str,
    file_id: str,
    correlation_id: str = "",
) -> None:
    """Process a single file as part of a fan-out job.

    Steps:
    1. Claim task (conditional write: pending → processing)
    2. Load job params (conditions list)
    3. Download single PDF from S3
    4. Extract text → assess quality → VLM fallback → check conditions
    5. TransactWriteItems: TASK result + JOB counter increment
    6. Check completion → maybe trigger assembly
    """
    worker_id = _generate_worker_id()
    job_type = ""
    if not correlation_id:
        correlation_id = f"{job_id}-{task_id}"
    _set_log_context(
        team_id=team_id, job_id=job_id, task_id=task_id,
        file_id=file_id, correlation_id=correlation_id, worker_id=worker_id,
    )
    log.info("Fan-out task: job=%s task=%s file=%s worker=%s", job_id, task_id, file_id, worker_id)

    # 1. Claim task.
    if not ddb_claim_task(ctx, team_id, job_id, task_id, worker_id):
        log.info("Task already claimed, skipping: %s/%s", job_id, task_id)
        return

    t0 = time.time()
    try:
        # 2. Load job params — also check if job is already cancelled/failed.
        job = ddb_get_item(ctx, _pk_team(team_id), _sk_job(job_id)) or {}
        job_status = str(job.get("status") or "")
        if job_status == "failed":
            log.info("Job already failed/cancelled, skipping task: %s/%s", job_id, task_id)
            return
        job_type = str(job.get("type") or "")
        params = job.get("params") if isinstance(job.get("params"), dict) else {}

        # 3. Load file metadata and download.
        file_row = ddb_get_item(ctx, _pk_team(team_id), _sk_file(file_id))
        if not file_row:
            raise RuntimeError(f"Missing file metadata: {file_id}")

        task_dir = _job_temp_dir(team_id, job_id) / f"task_{task_id}"
        task_dir.mkdir(parents=True, exist_ok=True)

        s3_key = str(file_row.get("s3_key") or "")
        s3_bucket = str(file_row.get("s3_bucket") or ctx.bucket)
        original_name = str(file_row.get("original_name") or file_id)
        ext = Path(s3_key).suffix or Path(original_name).suffix
        local_path = task_dir / f"{file_id}{ext}"
        s3_download(ctx, s3_bucket, s3_key, local_path)

        # 3.5. Validate downloaded file.
        _validate_pdf(local_path, original_name)

        # 4. Process based on job type.
        result: Dict[str, Any] = {"filename": original_name}

        if job_type == "condition_check":
            result = _process_single_condition_check(ctx, local_path, original_name, params)
        elif job_type == "document_extraction":
            result = _process_single_document_extraction(
                ctx, local_path, file_id, original_name, params,
            )
        elif job_type == "financial_extraction":
            result = _process_single_financial_extraction(
                ctx, local_path, file_id, original_name,
            )
        else:
            raise RuntimeError(f"Unsupported fan-out job type: {job_type}")

        # 5. Mark task succeeded + increment counter.
        ddb_complete_task(
            ctx, team_id, job_id, task_id,
            succeeded=True, result=result,
        )
        elapsed = time.time() - t0
        tu = result.get("token_usage") if isinstance(result.get("token_usage"), dict) else {}
        log.info(
            "Task succeeded: job=%s task=%s elapsed=%.1fs tokens=%d",
            job_id, task_id, elapsed, tu.get("total_tokens", 0),
        )
        _metrics.record_task(
            True, elapsed * 1000,
            input_tokens=tu.get("input_tokens", 0),
            output_tokens=tu.get("output_tokens", 0),
            job_type=job_type,
        )

    except Exception as exc:
        elapsed = time.time() - t0
        err_text = f"{type(exc).__name__}: {exc}"
        log.error("Task failed: job=%s task=%s error=%s", job_id, task_id, err_text)
        _metrics.record_task(False, elapsed * 1000, job_type=job_type)
        try:
            ddb_complete_task(
                ctx, team_id, job_id, task_id,
                succeeded=False,
                result={"filename": file_id, "error": err_text},
                error=err_text,
            )
        except Exception as inner:
            log.error("Failed to record task failure: %s", inner)
    finally:
        # Clean up task temp directory.
        task_dir = _job_temp_dir(team_id, job_id) / f"task_{task_id}"
        if task_dir.exists():
            shutil.rmtree(task_dir, ignore_errors=True)

    # 6. Circuit breaker check (after failures).
    if _check_circuit_breaker(ctx, team_id, job_id):
        _clear_log_context()
        return  # Job cancelled by circuit breaker; skip assembly.

    # 7. Check completion.
    # In production, DDB Streams + Lambda handles assembly automatically.
    # In local/dev mode, worker does it as a fallback.
    if not LAMBDA_ASSEMBLY_ENABLED:
        _check_and_maybe_assemble(ctx, team_id, job_id)
    else:
        log.info("Assembly delegated to Lambda (MERRY_LAMBDA_ASSEMBLY=true)")

    _clear_log_context()


CACHE_TTL_DAYS = int(os.getenv("MERRY_CACHE_TTL_DAYS", "7"))
CACHE_ENABLED = os.getenv("MERRY_RESULT_CACHE", "true").lower() != "false"


def _cache_key(pdf_path: Path, conditions: List[str]) -> str:
    """Generate a deterministic cache key: SHA256(file_content + sorted_conditions)."""
    import hashlib
    h = hashlib.sha256()
    with open(pdf_path, "rb") as f:
        while True:
            chunk = f.read(65536)
            if not chunk:
                break
            h.update(chunk)
    # Sort conditions for deterministic key regardless of input order.
    for c in sorted(conditions):
        h.update(c.encode("utf-8"))
    return h.hexdigest()[:32]


def _cache_get(ctx: AwsCtx, team_id: str, cache_key: str) -> Optional[Dict[str, Any]]:
    """Lookup result cache. Returns cached result or None."""
    if not CACHE_ENABLED:
        return None
    try:
        item = ddb_get_item(ctx, _pk_team(team_id), f"CACHE#{cache_key}")
        if not item:
            return None
        # Check TTL.
        expires = str(item.get("expires_at") or "")
        if expires and expires < _now_iso():
            return None  # Expired.
        result_json = str(item.get("result") or "")
        if result_json:
            return json.loads(result_json)
    except Exception:
        pass  # Cache miss on any error.
    return None


def _cache_put(ctx: AwsCtx, team_id: str, cache_key: str, result: Dict[str, Any]) -> None:
    """Store result in cache with TTL."""
    if not CACHE_ENABLED:
        return
    try:
        from datetime import timedelta
        now = datetime.now(timezone.utc)
        expires = (now + timedelta(days=CACHE_TTL_DAYS)).isoformat().replace("+00:00", "Z")
        result_json = json.dumps(result, ensure_ascii=False, default=str)
        # Skip caching very large results.
        if len(result_json) > 200_000:
            return
        _call_with_backoff(
            ctx.ddb.put_item,
            Item=_ddb_sanitize({
                "pk": _pk_team(team_id),
                "sk": f"CACHE#{cache_key}",
                "entity": "cache",
                "result": result_json,
                "created_at": _now_iso(),
                "expires_at": expires,
                "ttl": int((now + timedelta(days=CACHE_TTL_DAYS)).timestamp()),
            }),
            max_retries=1, base_delay=0.5,
        )
    except Exception as e:
        log.debug("Cache put failed: %s", e)  # Best-effort.


def _process_single_condition_check(
    ctx: AwsCtx,
    pdf_path: Path,
    filename: str,
    params: Dict[str, Any],
) -> Dict[str, Any]:
    """Process a single PDF for condition checking with backoff."""
    from ralph.playground_parser import (
        extract_text, assess_text_quality,
        render_first_page, render_pages, analyze_pages,
        call_nova_visual, build_presentation_prompt, _PROMPT_OCR,
    )
    from ralph.condition_checker import check_conditions_nova

    conditions: List[str] = [str(c) for c in (params.get("conditions") or []) if c]
    if not conditions:
        raise RuntimeError("conditions 파라미터가 비어 있습니다")

    # Cache lookup.
    team_id = getattr(_log_context, "ctx", {}).get("team_id", "") if hasattr(_log_context, "ctx") and _log_context.ctx else ""
    ck = _cache_key(pdf_path, conditions)
    cached = _cache_get(ctx, team_id, ck) if team_id else None
    if cached:
        cached["filename"] = filename
        cached["_cached"] = True
        log.info("Cache hit for %s (key=%s)", filename, ck[:8])
        return cached

    model_id = str(params.get("model_id") or os.getenv("RALPH_VLM_NOVA_MODEL_ID", "us.amazon.nova-pro-v1:0"))
    model_lite = os.getenv("RALPH_VLM_NOVA_LITE_MODEL_ID", "us.amazon.nova-lite-v1:0")
    region = str(params.get("region") or os.getenv("RALPH_VLM_NOVA_REGION", "us-east-1"))
    default_use_vlm = "false" if os.getenv("PYTEST_CURRENT_TEST") else "true"
    use_vlm = os.getenv("RALPH_USE_VLM", default_use_vlm).lower() != "false"

    t0 = time.time()
    entry: Dict[str, Any] = {"filename": filename}

    text, pages, blocks = _call_with_backoff(extract_text, str(pdf_path))
    _, is_poor, is_fragmented = assess_text_quality(text, blocks, page_count=pages)

    extracted = ""
    method = "pymupdf"
    total_input_tokens = 0
    total_output_tokens = 0

    if use_vlm and is_poor:
        img = render_first_page(str(pdf_path))
        vd = _call_with_backoff(call_nova_visual, img, model_lite, region, _PROMPT_OCR)
        extracted = vd.get("readable_text") or ""
        method = "nova_hybrid"
        vlm_usage = vd.pop("_usage", {})
        total_input_tokens += vlm_usage.get("input_tokens", 0)
        total_output_tokens += vlm_usage.get("output_tokens", 0)
    elif use_vlm and is_fragmented:
        imgs = render_pages(str(pdf_path), max_pages=10, dpi=100)
        info = analyze_pages(str(pdf_path), max_pages=10)
        prompt = build_presentation_prompt(info)
        vd = _call_with_backoff(call_nova_visual, imgs, model_id, region, prompt, max_tokens=5000)
        extracted = vd.get("readable_text") or ""
        method = "nova_presentation"
        vlm_usage = vd.pop("_usage", {})
        total_input_tokens += vlm_usage.get("input_tokens", 0)
        total_output_tokens += vlm_usage.get("output_tokens", 0)

    full_text = "\n\n".join(filter(None, [extracted, text]))
    check = _call_with_backoff(check_conditions_nova, full_text, conditions, model_id, region)

    # Aggregate token usage from condition check call.
    check_usage = check.pop("_usage", {})
    total_input_tokens += check_usage.get("input_tokens", 0)
    total_output_tokens += check_usage.get("output_tokens", 0)

    entry.update({
        "method": method,
        "pages": pages,
        "company_name": check.get("company_name"),
        "conditions": check.get("conditions", []),
        "parse_warning": check.get("parse_warning"),
        "raw_response": check.get("raw_response"),
        "elapsed_s": round(time.time() - t0, 1),
        "token_usage": {
            "input_tokens": total_input_tokens,
            "output_tokens": total_output_tokens,
            "total_tokens": total_input_tokens + total_output_tokens,
        },
    })

    # Store in cache for future identical requests.
    if team_id:
        _cache_put(ctx, team_id, ck, entry)

    return entry


def _process_single_document_extraction(
    ctx: AwsCtx,
    pdf_path: Path,
    file_id: str,
    filename: str,
    params: Dict[str, Any],
) -> Dict[str, Any]:
    """Process a single document for extraction (fan-out version)."""
    from ralph.batch_pipeline import process_batch, BatchDocInput

    type_map = params.get("typeMap") or {}
    doc_type = str(type_map.get(file_id, "unknown"))
    if doc_type == "unknown":
        raise RuntimeError(f"No type mapping for file {file_id}")

    t0 = time.time()
    docs = [BatchDocInput(
        file_id=file_id,
        filename=filename,
        pdf_path=str(pdf_path),
        doc_type=doc_type,
    )]

    output_dir = str(pdf_path.parent / "extraction_output")
    batch_result = process_batch(docs, output_dir=output_dir)

    if not batch_result.results:
        raise RuntimeError("No extraction results")

    r = batch_result.results[0]

    # Capture token usage from VLM calls (if any).
    vlm_usage = getattr(r, "vlm_usage", {}) or {}
    input_tokens = vlm_usage.get("input_tokens", 0)
    output_tokens = vlm_usage.get("output_tokens", 0)

    result: Dict[str, Any] = {
        "file_id": file_id,
        "filename": filename,
        "doc_type": doc_type,
        "success": r.success,
        "confidence": float(r.confidence),
        "method": r.method,
        "elapsed_s": round(time.time() - t0, 1),
        "token_usage": {
            "input_tokens": input_tokens,
            "output_tokens": output_tokens,
            "total_tokens": input_tokens + output_tokens,
        },
    }

    # Read the JSON result file if it exists.
    json_file = Path(output_dir) / f"{file_id}.json"
    if json_file.exists():
        try:
            extracted = json.loads(json_file.read_text(encoding="utf-8"))
            result["extracted"] = extracted
        except Exception:
            pass

    if not r.success:
        result["error"] = getattr(r, "error", "Extraction failed")

    return result


def _process_single_financial_extraction(
    ctx: AwsCtx,
    pdf_path: Path,
    file_id: str,
    filename: str,
) -> Dict[str, Any]:
    """Process a single financial statement PDF."""
    from ralph.batch_pipeline import BatchDocInput, process_batch

    t0 = time.time()
    docs = [BatchDocInput(
        file_id=file_id,
        filename=filename,
        pdf_path=str(pdf_path),
        doc_type="financial_stmt",
    )]

    output_dir = str(pdf_path.parent / "financial_output")
    batch_result = process_batch(docs, output_dir=output_dir)
    if not batch_result.results:
        raise RuntimeError("No financial extraction results")

    r = batch_result.results[0]
    vlm_usage = getattr(r, "vlm_usage", {}) or {}
    input_tokens = vlm_usage.get("input_tokens", 0)
    output_tokens = vlm_usage.get("output_tokens", 0)

    result: Dict[str, Any] = {
        "file_id": file_id,
        "filename": filename,
        "doc_type": "financial_stmt",
        "success": r.success,
        "confidence": float(r.confidence),
        "method": r.method,
        "elapsed_s": round(time.time() - t0, 1),
        "token_usage": {
            "input_tokens": input_tokens,
            "output_tokens": output_tokens,
            "total_tokens": input_tokens + output_tokens,
        },
    }

    json_file = Path(output_dir) / f"{file_id}.json"
    extracted: Dict[str, Any] = {}
    if json_file.exists():
        try:
            extracted = json.loads(json_file.read_text(encoding="utf-8"))
            result["extracted"] = extracted
        except Exception:
            extracted = {}

    data = extracted.get("data") if isinstance(extracted.get("data"), dict) else {}
    corp_name = str(data.get("corp_name") or Path(filename).stem)
    statements = data.get("statements") if isinstance(data.get("statements"), list) else []
    result["company_name"] = corp_name
    result["statement_count"] = len(statements)

    if not r.success:
        result["error"] = "; ".join(getattr(r, "errors", []) or []) or "Financial extraction failed"

    return result


def _check_and_maybe_assemble(
    ctx: AwsCtx, team_id: str, job_id: str,
) -> None:
    """Check if all tasks are done. If so, assemble CSV and finalize job."""
    job = ddb_get_item(ctx, _pk_team(team_id), _sk_job(job_id)) or {}
    total = int(job.get("total_tasks") or 0)
    processed = int(job.get("processed_count") or 0)
    failed = int(job.get("failed_count") or 0)

    if total <= 0 or processed < total:
        return  # Not done yet.

    log.info(
        "All tasks done: job=%s total=%d processed=%d failed=%d",
        job_id, total, processed, failed,
    )

    # Try to claim assembly (only one worker succeeds).
    if not ddb_try_claim_assembly(ctx, team_id, job_id):
        log.info("Assembly already claimed by another worker: %s", job_id)
        return

    log.info("Assembling results for job=%s", job_id)
    try:
        _assemble_fanout_results(ctx, team_id, job_id, job)
    except Exception as exc:
        log.error("Assembly failed for job=%s: %s", job_id, exc)
        # Mark job failed so it doesn't hang in "assembling" forever.
        ddb_update_job(
            ctx, team_id, job_id,
            status="failed",
            error=f"Assembly error: {exc}",
        )
        # Also update fanout_status.
        ctx.ddb.update_item(
            Key={"pk": _pk_team(team_id), "sk": _sk_job(job_id)},
            UpdateExpression="SET #fs = :failed",
            ExpressionAttributeNames={"#fs": "fanout_status"},
            ExpressionAttributeValues={":failed": "failed"},
        )
        ddb_update_job_index_state(ctx, team_id, job_id, fanout=True, fanout_status="failed")


def _assemble_fanout_results(
    ctx: AwsCtx, team_id: str, job_id: str, job: Dict[str, Any],
) -> None:
    """Query all task results, build CSV + JSON, upload to S3, mark job succeeded."""
    job_type = str(job.get("type") or "")
    params = job.get("params") if isinstance(job.get("params"), dict) else {}
    conditions: List[str] = [str(c) for c in (params.get("conditions") or []) if c]

    # Query all tasks.
    tasks = ddb_query_all_tasks(ctx, team_id, job_id)
    tasks.sort(key=lambda t: str(t.get("sk", "")))

    # Parse task results.
    rows: List[Dict[str, Any]] = []
    for task in tasks:
        result_raw = task.get("result")
        if isinstance(result_raw, str):
            try:
                result_data = json.loads(result_raw)
            except json.JSONDecodeError:
                result_data = {"filename": str(task.get("file_id", "")), "error": result_raw}
        elif isinstance(result_raw, dict):
            result_data = result_raw
        else:
            result_data = {
                "filename": str(task.get("file_id", "")),
                "error": str(task.get("error", "unknown")),
            }
        rows.append(result_data)

    # Build CSV (same format as legacy _handle_condition_check).
    job_dir = _job_temp_dir(team_id, job_id)
    try:
        _assemble_fanout_inner(ctx, team_id, job_id, job, job_type, conditions, rows, job_dir)
    finally:
        if job_dir.exists():
            shutil.rmtree(job_dir, ignore_errors=True)


def _assemble_fanout_inner(
    ctx: AwsCtx,
    team_id: str,
    job_id: str,
    job: Dict[str, Any],
    job_type: str,
    conditions: List[str],
    rows: List[Dict[str, Any]],
    job_dir: Path,
) -> None:
    assembly_dir = job_dir / "assembly"
    assembly_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path: Optional[Path] = None
    companies: List[Dict[str, Any]] = []
    if job_type == "condition_check" and conditions:
        csv_path, json_path = _build_condition_check_csv(assembly_dir, rows, conditions)
        try:
            xlsx_path = _build_condition_check_xlsx(assembly_dir, rows, conditions)
        except Exception as e:
            log.warning("XLSX generation failed (CSV still available): %s", e)
    elif job_type == "financial_extraction":
        csv_path, json_path, xlsx_path, companies = _build_financial_extraction_artifacts(
            assembly_dir, rows,
        )
    else:
        # Generic JSON-only output for other fan-out types.
        json_path = assembly_dir / "results.json"
        _write_json(json_path, {"total": len(rows), "results": rows})
        csv_path = None
        companies = []

    # Artifact labels by job type.
    _labels = {
        "condition_check": ("조건 검사 결과", "condition_check"),
        "document_extraction": ("문서 추출 결과", "document_extraction"),
        "financial_extraction": ("재무 데이터 추출 결과", "financial_extraction"),
    }
    label_prefix, artifact_prefix = _labels.get(job_type, ("결과", "results"))

    # Upload artifacts.
    uploaded: List[Dict[str, Any]] = []
    total_bytes = 0

    if xlsx_path and xlsx_path.exists():
        dest_key = f"artifacts/{team_id}/{job_id}/{artifact_prefix}_xlsx{xlsx_path.suffix}"
        xlsx_ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        size = s3_upload(ctx, ctx.bucket, dest_key, xlsx_path, xlsx_ct)
        total_bytes += size
        uploaded.append({
            "artifactId": f"{artifact_prefix}_xlsx",
            "label": f"{label_prefix} (Excel)",
            "contentType": xlsx_ct,
            "s3Bucket": ctx.bucket,
            "s3Key": dest_key,
            "sizeBytes": size,
        })

    if csv_path and csv_path.exists():
        dest_key = f"artifacts/{team_id}/{job_id}/{artifact_prefix}_csv{csv_path.suffix}"
        size = s3_upload(ctx, ctx.bucket, dest_key, csv_path, "text/csv")
        total_bytes += size
        uploaded.append({
            "artifactId": f"{artifact_prefix}_csv",
            "label": f"{label_prefix} (CSV)",
            "contentType": "text/csv",
            "s3Bucket": ctx.bucket,
            "s3Key": dest_key,
            "sizeBytes": size,
        })

    if json_path and json_path.exists():
        dest_key = f"artifacts/{team_id}/{job_id}/{artifact_prefix}_json{json_path.suffix}"
        size = s3_upload(ctx, ctx.bucket, dest_key, json_path, "application/json")
        total_bytes += size
        uploaded.append({
            "artifactId": f"{artifact_prefix}_json",
            "label": f"{label_prefix} (JSON)",
            "contentType": "application/json",
            "s3Bucket": ctx.bucket,
            "s3Key": dest_key,
            "sizeBytes": size,
        })

    # Delete input files from S3 (security).
    deleted_inputs: List[str] = []
    if ctx.delete_inputs:
        file_ids = job.get("input_file_ids") if isinstance(job.get("input_file_ids"), list) else []
        for fid in file_ids:
            fid = str(fid)
            if not fid:
                continue
            try:
                row = ddb_get_item(ctx, _pk_team(team_id), _sk_file(fid))
                if row:
                    bucket = str(row.get("s3_bucket") or ctx.bucket)
                    key = str(row.get("s3_key") or "")
                    if key:
                        s3_delete(ctx, bucket, key)
                        ddb_mark_file_deleted(ctx, team_id, fid)
                        deleted_inputs.append(fid)
            except Exception:
                pass  # Best-effort; lifecycle rule handles leftovers.

    success_count = sum(1 for r in rows if "error" not in r)
    warning_count = sum(1 for r in rows if r.get("parse_warning"))

    # Aggregate token usage across all tasks.
    total_input_tokens = 0
    total_output_tokens = 0
    for r in rows:
        tu = r.get("token_usage") if isinstance(r.get("token_usage"), dict) else {}
        total_input_tokens += int(tu.get("input_tokens", 0))
        total_output_tokens += int(tu.get("output_tokens", 0))

    metrics: Dict[str, Any] = {
        "total": len(rows),
        "success_count": success_count,
        "failed_count": len(rows) - success_count,
        "warning_count": warning_count,
        "conditions": conditions,
        "companies": companies,
        "artifacts_bytes": total_bytes,
        "deleted_inputs": deleted_inputs,
        "ended_at": _now_iso(),
        "token_usage": {
            "input_tokens": total_input_tokens,
            "output_tokens": total_output_tokens,
            "total_tokens": total_input_tokens + total_output_tokens,
        },
    }

    # Finalize job.
    final_status = "succeeded"  # Even if some tasks failed, the job "succeeded" overall.
    ddb_update_job(
        ctx, team_id, job_id,
        status=final_status,
        error="",
        artifacts=uploaded,
        metrics=metrics,
    )
    # Update fanout_status to match.
    ctx.ddb.update_item(
        Key={"pk": _pk_team(team_id), "sk": _sk_job(job_id)},
        UpdateExpression="SET #fs = :status",
        ExpressionAttributeNames={"#fs": "fanout_status"},
        ExpressionAttributeValues={":status": "succeeded"},
    )
    ddb_update_job_index_state(ctx, team_id, job_id, fanout=True, fanout_status="succeeded")

    log.info(
        "Assembly complete: job=%s artifacts=%d success=%d failed=%d",
        job_id, len(uploaded), success_count, len(rows) - success_count,
    )

    # Webhook notification.
    job_title = str(job.get("title") or job_type)
    _send_webhook(
        job_id, job_title, "succeeded",
        total=len(rows), success=success_count, failed=len(rows) - success_count,
    )

    # Release memory after assembly — rows can be large for 700+ files.
    del rows
    gc.collect()


def _build_condition_check_csv(
    output_dir: Path,
    rows: List[Dict[str, Any]],
    conditions: List[str],
) -> Tuple[Path, Path]:
    """Build CSV and JSON files from condition check results (same format as legacy)."""
    buf = io.StringIO()
    fieldnames = ["filename", "company_name", "method", "pages", "elapsed_s", "warning", "error"]
    for c in conditions:
        short = c[:30].replace(" ", "_")
        fieldnames += [f"{short}_result", f"{short}_evidence"]
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        row: Dict[str, Any] = {
            "filename": r.get("filename", ""),
            "company_name": r.get("company_name", ""),
            "method": r.get("method", ""),
            "pages": r.get("pages", ""),
            "elapsed_s": r.get("elapsed_s", ""),
            "warning": r.get("parse_warning", ""),
            "error": r.get("error", ""),
        }
        cond_results = r.get("conditions") or []
        for j, c in enumerate(conditions):
            short = c[:30].replace(" ", "_")
            if j < len(cond_results):
                cr = cond_results[j]
                row[f"{short}_result"] = "✓" if cr.get("result") else "✗"
                row[f"{short}_evidence"] = cr.get("evidence", "")
            else:
                row[f"{short}_result"] = ""
                row[f"{short}_evidence"] = ""
        writer.writerow(row)

    csv_path = output_dir / "condition_check_results.csv"
    csv_path.write_text(buf.getvalue(), encoding="utf-8-sig")

    json_path = output_dir / "condition_check_results.json"
    warning_count = sum(1 for r in rows if r.get("parse_warning"))
    _write_json(json_path, {
        "conditions": conditions,
        "total": len(rows),
        "warning_count": warning_count,
        "results": rows,
    })
    return csv_path, json_path


def _build_condition_check_xlsx(
    output_dir: Path,
    rows: List[Dict[str, Any]],
    conditions: List[str],
) -> Path:
    """Build a styled XLSX workbook from condition check results.

    - Header row: bold white text on blue background, frozen.
    - Result cells: green fill for ✓, red fill for ✗.
    - Auto-adjusted column widths.
    - Evidence columns word-wrapped.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "조건 검사 결과"

    # ── Styles ──
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="3182F6", end_color="3182F6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D6DB"),
        right=Side(style="thin", color="D1D6DB"),
        top=Side(style="thin", color="D1D6DB"),
        bottom=Side(style="thin", color="D1D6DB"),
    )
    pass_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    pass_font = Font(name="맑은 고딕", color="065F46", bold=True, size=10)
    fail_fill = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
    fail_font = Font(name="맑은 고딕", color="9F1239", bold=True, size=10)
    body_font = Font(name="맑은 고딕", size=9)
    body_align = Alignment(vertical="top", wrap_text=False)
    evidence_align = Alignment(vertical="top", wrap_text=True)
    warning_font = Font(name="맑은 고딕", color="B45309", size=9)
    error_font = Font(name="맑은 고딕", color="DC2626", size=9)

    # ── Headers ──
    base_headers = ["파일명", "회사명", "추출 방식", "페이지 수", "처리 시간(초)", "경고", "에러"]
    headers = list(base_headers)
    for c in conditions:
        short = c[:40]
        headers.append(f"[결과] {short}")
        headers.append(f"[근거] {short}")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.freeze_panes = "A2"

    # ── Data rows ──
    for row_idx, r in enumerate(rows, 2):
        base_values = [
            r.get("filename", ""),
            r.get("company_name", ""),
            r.get("method", ""),
            r.get("pages", ""),
            r.get("elapsed_s", ""),
            r.get("parse_warning", ""),
            r.get("error", ""),
        ]
        cond_results = r.get("conditions") or []

        for col_idx, val in enumerate(base_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")
            if col_idx == 6 and val:
                cell.font = warning_font
            elif col_idx == 7 and val:
                cell.font = error_font
            else:
                cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border

        for j, _c in enumerate(conditions):
            result_col = len(base_values) + (j * 2) + 1
            evidence_col = result_col + 1

            if j < len(cond_results):
                cr = cond_results[j]
                is_pass = bool(cr.get("result"))
                result_text = "✓ 충족" if is_pass else "✗ 미충족"
                evidence_text = str(cr.get("evidence", ""))

                rc = ws.cell(row=row_idx, column=result_col, value=result_text)
                rc.fill = pass_fill if is_pass else fail_fill
                rc.font = pass_font if is_pass else fail_font
                rc.alignment = Alignment(horizontal="center", vertical="center")
                rc.border = thin_border

                ec = ws.cell(row=row_idx, column=evidence_col, value=evidence_text)
                ec.font = body_font
                ec.alignment = evidence_align
                ec.border = thin_border
            else:
                for col in (result_col, evidence_col):
                    cell = ws.cell(row=row_idx, column=col, value="")
                    cell.border = thin_border

    # ── Auto column widths ──
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 50))
        col_letter = get_column_letter(col_idx)
        # Evidence columns get wider.
        is_evidence = col_idx > len(base_headers) and (col_idx - len(base_headers)) % 2 == 0
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60) if is_evidence else min(max_len + 3, 30)

    # ── Summary sheet ──
    ws2 = wb.create_sheet("요약")
    total = len(rows)
    warnings = sum(1 for r in rows if r.get("parse_warning"))
    errors = sum(1 for r in rows if r.get("error"))
    ws2.cell(row=1, column=1, value="총 파일 수").font = Font(bold=True)
    ws2.cell(row=1, column=2, value=total)
    ws2.cell(row=2, column=1, value="경고 파일 수").font = Font(bold=True)
    ws2.cell(row=2, column=2, value=warnings)
    ws2.cell(row=3, column=1, value="에러 파일 수").font = Font(bold=True)
    ws2.cell(row=3, column=2, value=errors)
    ws2.cell(row=4, column=1, value="검사 조건 수").font = Font(bold=True)
    ws2.cell(row=4, column=2, value=len(conditions))

    for i, c in enumerate(conditions):
        pass_count = 0
        fail_count = 0
        for r in rows:
            cond_results = r.get("conditions") or []
            if i < len(cond_results):
                if cond_results[i].get("result"):
                    pass_count += 1
                else:
                    fail_count += 1
        ws2.cell(row=6 + i, column=1, value=c[:60]).font = body_font
        ws2.cell(row=6 + i, column=2, value=f"✓ {pass_count}").font = pass_font
        ws2.cell(row=6 + i, column=3, value=f"✗ {fail_count}").font = fail_font

    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    xlsx_path = output_dir / "condition_check_results.xlsx"
    wb.save(str(xlsx_path))
    return xlsx_path


def _build_financial_extraction_artifacts(
    output_dir: Path,
    rows: List[Dict[str, Any]],
) -> Tuple[Path, Path, Path, List[Dict[str, Any]]]:
    """Build summary artifacts for financial extraction fan-out jobs."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    financial_rows: List[Dict[str, Any]] = []
    inventory_rows: List[Dict[str, Any]] = []
    companies: Dict[str, Dict[str, Any]] = {}

    for r in rows:
        company_name = str(r.get("company_name") or Path(str(r.get("filename") or "unknown")).stem)
        filename = str(r.get("filename") or "")
        extracted = r.get("extracted") if isinstance(r.get("extracted"), dict) else {}
        data = extracted.get("data") if isinstance(extracted.get("data"), dict) else {}
        statements = data.get("statements") if isinstance(data.get("statements"), list) else []

        inventory_rows.append({
            "company_name": company_name,
            "filename": filename,
            "success": "error" not in r,
            "confidence": r.get("confidence"),
            "method": r.get("method"),
            "statement_count": len(statements),
            "error": r.get("error", ""),
        })

        company = companies.setdefault(company_name, {
            "company_name": company_name,
            "file_count": 0,
            "has_financials": False,
            "has_cap_table": False,
        })
        company["file_count"] += 1
        company["has_financials"] = company["has_financials"] or bool(statements)

        for stmt in statements:
            if not isinstance(stmt, dict):
                continue
            financial_rows.append({
                "company_name": company_name,
                "source_file": filename,
                "year": stmt.get("year"),
                "revenue": stmt.get("revenue"),
                "operating_income": stmt.get("operating_income"),
                "net_income": stmt.get("net_income"),
                "total_assets": stmt.get("total_assets"),
                "total_liabilities": stmt.get("total_liabilities"),
                "equity": stmt.get("equity"),
            })

    financial_rows.sort(key=lambda row: (str(row.get("company_name") or ""), str(row.get("year") or "")))
    company_rows = sorted(companies.values(), key=lambda row: row["company_name"])

    csv_path = output_dir / "financial_summary.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "company_name",
                "year",
                "source_file",
                "revenue",
                "operating_income",
                "net_income",
                "total_assets",
                "total_liabilities",
                "equity",
            ],
        )
        writer.writeheader()
        writer.writerows(financial_rows)

    json_path = output_dir / "financial_summary.json"
    _write_json(json_path, {
        "total": len(rows),
        "companies": company_rows,
        "financials": financial_rows,
        "results": rows,
    })

    wb = Workbook()
    wb.active.title = "Financial Summary"
    header_fill = PatternFill(fill_type="solid", fgColor="3182F6")
    header_font = Font(bold=True, color="FFFFFF")
    body_align = Alignment(vertical="center")

    def write_sheet(name: str, data_rows: List[Dict[str, Any]], columns: List[str]) -> None:
        sheet = wb[name] if name in wb.sheetnames else wb.create_sheet(title=name)
        for idx, col in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=idx, value=col)
            cell.fill = header_fill
            cell.font = header_font
        for row_idx, data in enumerate(data_rows, 2):
            for col_idx, col in enumerate(columns, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=data.get(col))
                cell.alignment = body_align
        for col_idx, col in enumerate(columns, 1):
            max_len = len(col)
            for row_idx in range(2, len(data_rows) + 2):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, min(len(str(value)), 40))
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 42)
        sheet.freeze_panes = "A2"

    write_sheet("Financial Summary", financial_rows, [
        "company_name",
        "year",
        "source_file",
        "revenue",
        "operating_income",
        "net_income",
        "total_assets",
        "total_liabilities",
        "equity",
    ])
    write_sheet("Companies", company_rows, [
        "company_name",
        "file_count",
        "has_financials",
        "has_cap_table",
    ])
    write_sheet("Document Inventory", inventory_rows, [
        "company_name",
        "filename",
        "success",
        "confidence",
        "method",
        "statement_count",
        "error",
    ])

    xlsx_path = output_dir / "financial_report.xlsx"
    wb.save(str(xlsx_path))
    return csv_path, json_path, xlsx_path, company_rows


def process_job(ctx: AwsCtx, team_id: str, job_id: str) -> None:
    started = time.time()
    started_at = _now_iso()
    ddb_update_job(ctx, team_id, job_id, status="running", metrics={"started_at": started_at})

    job = ddb_get_item(ctx, _pk_team(team_id), _sk_job(job_id)) or {}
    job_type = str(job.get("type") or "")
    params = job.get("params") if isinstance(job.get("params"), dict) else {}
    file_ids = job.get("input_file_ids") if isinstance(job.get("input_file_ids"), list) else []
    file_ids = [str(x) for x in file_ids if x]
    if not file_ids:
        raise RuntimeError("Job has no inputs")

    job_dir = _job_temp_dir(team_id, job_id)
    try:
        _process_job_inner(ctx, team_id, job_id, job_type, params, file_ids, job_dir, started, started_at)
    finally:
        # Always clean up temp directory regardless of success or failure.
        if job_dir.exists():
            shutil.rmtree(job_dir, ignore_errors=True)


def _process_job_inner(
    ctx: AwsCtx,
    team_id: str,
    job_id: str,
    job_type: str,
    params: Dict[str, Any],
    file_ids: List[str],
    job_dir: Path,
    started: float,
    started_at: str,
) -> None:
    input_dir = job_dir / "inputs"
    input_dir.mkdir(parents=True, exist_ok=True)

    # Download all inputs.
    file_rows: List[Dict[str, Any]] = []
    local_inputs: List[Path] = []
    for file_id in file_ids:
        row = ddb_get_item(ctx, _pk_team(team_id), _sk_file(file_id))
        if not row:
            raise RuntimeError(f"Missing file metadata: {file_id}")
        file_rows.append(row)

        key = str(row.get("s3_key") or "")
        bucket = str(row.get("s3_bucket") or ctx.bucket)
        original = str(row.get("original_name") or file_id)
        ext = Path(key).suffix or Path(original).suffix
        dest = input_dir / f"{file_id}{ext}"
        s3_download(ctx, bucket, key, dest)
        local_inputs.append(dest)

    artifacts: List[Dict[str, Any]] = []
    metrics: Dict[str, Any] = {"started_at": started_at}

    # Dispatch.
    if job_type == "exit_projection":
        artifacts, extra = _handle_exit_projection(local_inputs[0], params)
        metrics.update(extra)
    elif job_type == "diagnosis_analysis":
        artifacts, extra = _handle_diagnosis(local_inputs[0])
        metrics.update(extra)
    elif job_type == "pdf_evidence":
        artifacts, extra = _handle_pdf_evidence(local_inputs[0], params)
        metrics.update(extra)
    elif job_type == "pdf_parse":
        artifacts, extra = _handle_pdf_parse(local_inputs[0], params)
        metrics.update(extra)
    elif job_type == "contract_review":
        artifacts, extra = _handle_contract_review(local_inputs[:2], params)
        metrics.update(extra)
    elif job_type == "document_extraction":
        artifacts, extra = _handle_document_extraction(local_inputs, file_rows, params)
        metrics.update(extra)
    elif job_type == "condition_check":
        artifacts, extra = _handle_condition_check(local_inputs, params)
        metrics.update(extra)
    else:
        raise RuntimeError(f"Unsupported job type: {job_type}")

    # Upload artifacts.
    uploaded: List[Dict[str, Any]] = []
    total_out = 0
    for a in artifacts:
        local_path = Path(str(a.get("localPath") or ""))
        if not local_path.exists():
            continue
        artifact_id = str(a.get("artifactId") or _safe_filename(local_path.name))
        label = str(a.get("label") or local_path.name)
        content_type = str(a.get("contentType") or "application/octet-stream")
        dest_key = f"artifacts/{team_id}/{job_id}/{artifact_id}{local_path.suffix}"
        size = s3_upload(ctx, ctx.bucket, dest_key, local_path, content_type)
        total_out += size
        uploaded.append(
            {
                "artifactId": artifact_id,
                "label": label,
                "contentType": content_type,
                "s3Bucket": ctx.bucket,
                "s3Key": dest_key,
                "sizeBytes": size,
            }
        )

    # Delete originals after processing (security).
    deleted_inputs = []
    if ctx.delete_inputs:
        for row in file_rows:
            bucket = str(row.get("s3_bucket") or ctx.bucket)
            key = str(row.get("s3_key") or "")
            file_id = str(row.get("file_id") or "")
            if not key or not file_id:
                continue
            try:
                s3_delete(ctx, bucket, key)
                ddb_mark_file_deleted(ctx, team_id, file_id)
                deleted_inputs.append(file_id)
            except Exception:
                # Best-effort. Lifecycle rule can still clean up.
                pass

    ended_at = _now_iso()
    duration_ms = int((time.time() - started) * 1000)
    metrics.update(
        {
            "ended_at": ended_at,
            "duration_ms": duration_ms,
            "artifacts_bytes": total_out,
            "deleted_inputs": deleted_inputs,
        }
    )

    usage = metrics.get("llm") if isinstance(metrics.get("llm"), dict) else None
    ddb_update_job(
        ctx,
        team_id,
        job_id,
        status="succeeded",
        error="",
        artifacts=uploaded,
        metrics=metrics,
        usage=usage,
    )


def _process_message(ctx: AwsCtx, payload: Dict[str, Any]) -> None:
    """Route a single SQS message to the appropriate handler."""
    version = payload.get("version")
    team_id = str(payload.get("teamId") or "")
    job_id = str(payload.get("jobId") or "")
    correlation_id = str(payload.get("correlationId") or job_id)

    if not team_id or not job_id:
        raise ValueError("Missing teamId/jobId in message")

    if version == 2:
        # Fan-out v2: per-file task.
        task_id = str(payload.get("taskId") or "")
        file_id = str(payload.get("fileId") or "")
        if not task_id or not file_id:
            raise ValueError("Missing taskId/fileId in v2 message")
        process_fanout_task(ctx, team_id, job_id, task_id, file_id, correlation_id=correlation_id)
    else:
        # Legacy: entire job in one message.
        _set_log_context(team_id=team_id, job_id=job_id, correlation_id=correlation_id)
        try:
            process_job(ctx, team_id, job_id)
        except Exception as exc:
            ended_at = _now_iso()
            err_text = f"{type(exc).__name__}: {exc}"
            ddb_update_job(
                ctx,
                team_id,
                job_id,
                status="failed",
                error=err_text,
                metrics={
                    "ended_at": ended_at,
                    "traceback": traceback.format_exc(limit=20),
                },
            )
        finally:
            _clear_log_context()


def worker_loop() -> None:
    ctx = AwsCtx()
    log.info(
        "Worker starting: region=%s table=%s bucket=%s concurrency=%d",
        ctx.region, ctx.ddb_table, ctx.bucket, WORKER_CONCURRENCY,
    )
    ctx.warmup()

    executor = ThreadPoolExecutor(
        max_workers=WORKER_CONCURRENCY,
        thread_name_prefix="task",
    )
    # Track in-flight futures keyed by SQS receipt handle.
    in_flight: Dict[str, Future] = {}

    # ── Graceful shutdown via SIGTERM/SIGINT ──
    import signal

    _shutdown_requested = threading.Event()

    def _signal_handler(signum: int, _frame: Any) -> None:
        sig_name = signal.Signals(signum).name
        log.info("Received %s — initiating graceful shutdown (in_flight=%d)", sig_name, len(in_flight))
        _shutdown_requested.set()

    signal.signal(signal.SIGTERM, _signal_handler)
    signal.signal(signal.SIGINT, _signal_handler)

    # ── Health check HTTP server (for ECS) ──
    _start_health_server(in_flight, _shutdown_requested)

    # ── Periodic task timers ──
    _last_metrics_flush = [time.time()]
    _last_temp_cleanup = [time.time()]
    _last_watchdog_run = [time.time()]
    TEMP_CLEANUP_INTERVAL = 300  # Purge stale temp dirs every 5 minutes.
    TEMP_MAX_AGE = 1800  # Temp dirs older than 30 minutes are stale.
    WATCHDOG_INTERVAL = 120  # Check for timed-out jobs every 2 minutes.

    # ── Main loop ──
    while not _shutdown_requested.is_set():
        # Drain completed futures and delete their SQS messages.
        _drain_completed(ctx, in_flight)

        # Flush metrics every ~60s (each SQS long-poll is ~20s, so every ~3 loops).
        if time.time() - _last_metrics_flush[0] >= 60:
            _metrics.flush(len(in_flight))
            _last_metrics_flush[0] = time.time()

        # Periodic temp directory cleanup to prevent disk/memory leak.
        if time.time() - _last_temp_cleanup[0] >= TEMP_CLEANUP_INTERVAL:
            _cleanup_stale_temp_dirs(TEMP_MAX_AGE)
            _last_temp_cleanup[0] = time.time()

        # Job timeout watchdog: force-fail jobs running beyond JOB_TIMEOUT_MINUTES.
        if time.time() - _last_watchdog_run[0] >= WATCHDOG_INTERVAL:
            _job_timeout_watchdog(ctx)
            _last_watchdog_run[0] = time.time()

        available = WORKER_CONCURRENCY - len(in_flight)
        if available <= 0:
            # All slots full — wait a bit, then drain again.
            time.sleep(1)
            continue

        # Determine visibility timeout: use shorter timeout if we expect fan-out tasks.
        # SQS ReceiveMessage can fetch up to 10, but we limit to available slots.
        fetch_count = min(available, 10)
        try:
            resp = ctx.sqs.receive_message(
                QueueUrl=ctx.queue_url,
                MaxNumberOfMessages=fetch_count,
                WaitTimeSeconds=20,
                VisibilityTimeout=FANOUT_VISIBILITY_TIMEOUT,
            )
        except Exception:
            if _shutdown_requested.is_set():
                break
            raise
        msgs = resp.get("Messages") or []
        _metrics.record_poll(empty=len(msgs) == 0)
        if not msgs:
            continue

        for msg in msgs:
            if _shutdown_requested.is_set():
                break
            receipt = msg.get("ReceiptHandle")
            body_raw = msg.get("Body") or ""
            try:
                payload = json.loads(body_raw)
            except json.JSONDecodeError:
                log.warning("Bad JSON in SQS message, deleting")
                if receipt:
                    ctx.sqs.delete_message(QueueUrl=ctx.queue_url, ReceiptHandle=receipt)
                continue

            # For legacy messages, extend visibility timeout since they take longer.
            if payload.get("version") != 2 and receipt:
                try:
                    ctx.sqs.change_message_visibility(
                        QueueUrl=ctx.queue_url,
                        ReceiptHandle=receipt,
                        VisibilityTimeout=LEGACY_VISIBILITY_TIMEOUT,
                    )
                except Exception:
                    pass  # Best-effort.

            future = executor.submit(_process_message, ctx, payload)
            if receipt:
                in_flight[receipt] = future
            else:
                log.warning("Message without ReceiptHandle")

    # ── Graceful drain: wait for in-flight tasks to finish ──
    DRAIN_TIMEOUT = int(os.getenv("MERRY_DRAIN_TIMEOUT", "120"))
    # How often to extend visibility of in-flight SQS messages during drain.
    VISIBILITY_EXTEND_INTERVAL = 30  # seconds
    # Extension amount (must be > DRAIN_TIMEOUT to prevent re-delivery during drain).
    VISIBILITY_EXTEND_SECONDS = max(DRAIN_TIMEOUT + 60, FANOUT_VISIBILITY_TIMEOUT)

    if in_flight:
        log.info("Draining %d in-flight tasks (timeout=%ds)...", len(in_flight), DRAIN_TIMEOUT)
        deadline = time.time() + DRAIN_TIMEOUT
        last_visibility_extend = time.time()

        while in_flight and time.time() < deadline:
            _drain_completed(ctx, in_flight)
            if not in_flight:
                break

            # Periodically extend SQS visibility for still-running tasks
            # to prevent re-delivery while we wait for them to finish.
            if time.time() - last_visibility_extend >= VISIBILITY_EXTEND_INTERVAL:
                for receipt in list(in_flight.keys()):
                    try:
                        ctx.sqs.change_message_visibility(
                            QueueUrl=ctx.queue_url,
                            ReceiptHandle=receipt,
                            VisibilityTimeout=VISIBILITY_EXTEND_SECONDS,
                        )
                    except Exception:
                        pass  # Best-effort; receipt may be stale.
                log.info(
                    "Extended visibility for %d in-flight messages (+%ds)",
                    len(in_flight), VISIBILITY_EXTEND_SECONDS,
                )
                last_visibility_extend = time.time()

            time.sleep(1)

        if in_flight:
            log.warning("Drain timeout: %d tasks still running, forcing shutdown", len(in_flight))
        else:
            log.info("All in-flight tasks drained successfully")

    executor.shutdown(wait=False)
    _metrics.flush(0)
    log.info("Worker shut down")


def _start_health_server(
    in_flight: Dict[str, Future],
    shutdown_event: threading.Event,
) -> None:
    """Start a lightweight HTTP health check server on a background thread.

    GET /health → 200 {"status":"ok"} (liveness — always 200 unless draining)
    GET /ready  → 200 if worker can accept work, 503 if draining or overloaded
    GET /metrics → 200 Prometheus-style text metrics for observability
    Used by ECS health checks (port configurable via MERRY_HEALTH_PORT, default 8080).
    """
    from http.server import HTTPServer, BaseHTTPRequestHandler

    port = int(os.getenv("MERRY_HEALTH_PORT", "8080"))
    _state: Dict[str, Any] = {
        "start": time.time(),
        "total_processed": 0,
        "total_failed": 0,
    }

    class Handler(BaseHTTPRequestHandler):
        def do_GET(self) -> None:
            if self.path == "/health":
                self._handle_health()
            elif self.path == "/ready":
                self._handle_ready()
            elif self.path == "/metrics":
                self._handle_metrics()
            else:
                self.send_response(404)
                self.end_headers()

        def _handle_health(self) -> None:
            is_shutting_down = shutdown_event.is_set()
            body = json.dumps({
                "status": "draining" if is_shutting_down else "ok",
                "in_flight": len(in_flight),
                "shutdown": is_shutting_down,
                "concurrency": WORKER_CONCURRENCY,
                "uptime_s": round(time.time() - _state["start"]),
            })
            # Return 503 during drain so ECS stops routing traffic.
            code = 503 if is_shutting_down else 200
            self.send_response(code)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(body.encode())

        def _handle_ready(self) -> None:
            is_shutting_down = shutdown_event.is_set()
            has_capacity = len(in_flight) < WORKER_CONCURRENCY
            ready = not is_shutting_down and has_capacity
            body = json.dumps({
                "ready": ready,
                "in_flight": len(in_flight),
                "capacity": WORKER_CONCURRENCY - len(in_flight),
                "shutdown": is_shutting_down,
            })
            code = 200 if ready else 503
            self.send_response(code)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(body.encode())

        def _handle_metrics(self) -> None:
            uptime = time.time() - _state["start"]
            cur_in_flight = len(in_flight)
            # Read current accumulator snapshot (non-destructive).
            with _metrics._lock:
                succeeded = _metrics._tasks_succeeded + _state["total_processed"]
                failed = _metrics._tasks_failed + _state["total_failed"]
                input_tok = _metrics._input_tokens
                output_tok = _metrics._output_tokens
                retries = _metrics._retries

            lines = [
                "# HELP merry_worker_uptime_seconds Worker uptime in seconds",
                "# TYPE merry_worker_uptime_seconds gauge",
                f"merry_worker_uptime_seconds {uptime:.1f}",
                "# HELP merry_worker_in_flight Current in-flight tasks",
                "# TYPE merry_worker_in_flight gauge",
                f"merry_worker_in_flight {cur_in_flight}",
                "# HELP merry_worker_concurrency Max concurrent tasks",
                "# TYPE merry_worker_concurrency gauge",
                f"merry_worker_concurrency {WORKER_CONCURRENCY}",
                "# HELP merry_worker_tasks_total Total tasks processed",
                "# TYPE merry_worker_tasks_total counter",
                f'merry_worker_tasks_total{{status="succeeded"}} {succeeded}',
                f'merry_worker_tasks_total{{status="failed"}} {failed}',
                "# HELP merry_worker_tokens_total Total LLM tokens used",
                "# TYPE merry_worker_tokens_total counter",
                f'merry_worker_tokens_total{{direction="input"}} {input_tok}',
                f'merry_worker_tokens_total{{direction="output"}} {output_tok}',
                "# HELP merry_worker_retries_total Total retries",
                "# TYPE merry_worker_retries_total counter",
                f"merry_worker_retries_total {retries}",
                "",
            ]
            body = "\n".join(lines)
            self.send_response(200)
            self.send_header("Content-Type", "text/plain; version=0.0.4; charset=utf-8")
            self.end_headers()
            self.wfile.write(body.encode())

        def log_message(self, format: str, *args: Any) -> None:
            pass  # Suppress noisy HTTP logs.

    def _serve() -> None:
        try:
            server = HTTPServer(("0.0.0.0", port), Handler)
            server.timeout = 1
            log.info("Health check listening on :%d [/health, /ready, /metrics]", port)
            while not shutdown_event.is_set():
                server.handle_request()
            server.server_close()
        except Exception as e:
            log.warning("Health server error: %s", e)

    t = threading.Thread(target=_serve, name="health", daemon=True)
    t.start()


def _cleanup_stale_temp_dirs(max_age_seconds: int) -> None:
    """Remove temp directories older than max_age_seconds to prevent disk bloat."""
    if not TEMP_ROOT.exists():
        return
    now = time.time()
    cleaned = 0
    try:
        for team_dir in TEMP_ROOT.iterdir():
            if not team_dir.is_dir():
                continue
            jobs_dir = team_dir / "jobs"
            if not jobs_dir.exists():
                continue
            for job_dir in jobs_dir.iterdir():
                if not job_dir.is_dir():
                    continue
                try:
                    age = now - job_dir.stat().st_mtime
                    if age > max_age_seconds:
                        shutil.rmtree(job_dir, ignore_errors=True)
                        cleaned += 1
                except OSError:
                    pass
    except OSError:
        pass
    if cleaned > 0:
        log.info("Cleaned up %d stale temp directories", cleaned)


def _drain_completed(ctx: AwsCtx, in_flight: Dict[str, Future]) -> None:
    """Remove completed futures from in_flight and delete their SQS messages."""
    done_receipts = []
    for receipt, future in in_flight.items():
        if future.done():
            done_receipts.append(receipt)
            # Log any exceptions from the future.
            exc = future.exception()
            if exc:
                log.error("Message processing error: %s", exc)

    for receipt in done_receipts:
        del in_flight[receipt]
        try:
            ctx.sqs.delete_message(QueueUrl=ctx.queue_url, ReceiptHandle=receipt)
        except Exception as e:
            log.warning("Failed to delete SQS message: %s", e)


if __name__ == "__main__":
    worker_loop()
