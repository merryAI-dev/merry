"""Tool definitions for VC Investment Agent"""

import os
import sys
import json
import re
import math
import subprocess
import shlex
import time
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path
from typing import Any, Dict, List, Callable, Optional

from shared.logging_config import get_logger
from shared.cache_utils import compute_file_hash, compute_payload_hash, get_cache_dir, load_json, save_json
from shared.airtable_portfolio import (
    search_portfolio_records,
    summarize_portfolio_records,
)

logger = get_logger("tools")

# 프로젝트 루트를 Python 경로에 추가
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

CACHE_VERSION = 1
CACHE_TTL_SECONDS = 60 * 60 * 6

DEFAULT_UNDERWRITER_DATA_PATH = (
    PROJECT_ROOT
    / "temp"
    / "dart_underwriter_opinion_2025_v5_window"
    / "underwriter_opinion.jsonl"
)

UNDERWRITER_CATEGORY_KEYWORDS = {
    "market_size": {
        "any": ["시장 규모", "시장규모", "TAM", "SAM", "SOM", "CAGR", "연평균", "성장률", "시장 전망", "시장전망", "시장성장"],
        "all": ["시장", "규모"],
    },
    "valuation": {
        "any": ["공모가격", "희망공모가액", "PER", "EV/EBITDA", "PBR", "PSR", "DCF", "할인율"],
        "all": [],
    },
    "comparables": {
        "any": ["비교회사", "비교기업", "유사기업", "선정"],
        "all": [],
    },
    "demand_forecast": {
        "any": ["수요예측", "수요 예측"],
        "all": [],
    },
    "risk": {
        "any": ["위험", "리스크", "유의", "불확실", "변동", "미래예측"],
        "all": [],
    },
}

_UNDERWRITER_TFIDF_CACHE = {
    "path": None,
    "mtime": None,
    "size": None,
    "min_n": None,
    "max_n": None,
    "max_text_chars": None,
    "idf": None,
    "vectors": None,
    "entries": None,
    "texts": None,
}


def _resolve_underwriter_data_path(jsonl_path: str = None) -> tuple:
    candidates = []
    if jsonl_path:
        candidates.append(Path(jsonl_path))

    env_path = os.getenv("UNDERWRITER_DATA_PATH")
    if env_path:
        candidates.append(Path(env_path))

    candidates.append(DEFAULT_UNDERWRITER_DATA_PATH)

    if not jsonl_path:
        temp_root = PROJECT_ROOT / "temp"
        if temp_root.exists():
            patterns = [
                "dart_underwriter_opinion*/underwriter_opinion.jsonl",
                "underwriter_opinion*/underwriter_opinion.jsonl",
                "**/underwriter_opinion.jsonl",
            ]
            seen = set()
            for pattern in patterns:
                for p in temp_root.glob(pattern):
                    if p in seen:
                        continue
                    seen.add(p)
                    candidates.append(p)

    for candidate in candidates:
        if not candidate:
            continue
        try:
            resolved = candidate.resolve()
        except OSError:
            continue

        if resolved.exists():
            return str(resolved), None

    message = (
        "인수인의견 JSONL 파일을 찾을 수 없습니다. "
        "temp/ 하위에 underwriter_opinion.jsonl을 두거나 "
        "UNDERWRITER_DATA_PATH 환경변수를 설정하세요."
    )
    return None, message


# ========================================
# 재시도 데코레이터 (외부 API 호출용)
# ========================================

def retry_with_backoff(
    max_retries: int = 3,
    base_delay: float = 1.0,
    max_delay: float = 30.0,
    exceptions: tuple = (Exception,)
) -> Callable:
    """
    지수 백오프 재시도 데코레이터

    Args:
        max_retries: 최대 재시도 횟수
        base_delay: 기본 대기 시간 (초)
        max_delay: 최대 대기 시간 (초)
        exceptions: 재시도할 예외 튜플
    """
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except exceptions as e:
                    last_exception = e
                    if attempt == max_retries - 1:
                        logger.error(f"{func.__name__} failed after {max_retries} attempts: {e}")
                        raise
                    delay = min(base_delay * (2 ** attempt), max_delay)
                    logger.warning(f"{func.__name__} retry {attempt + 1}/{max_retries} after {delay:.1f}s: {e}")
                    time.sleep(delay)
            raise last_exception
        return wrapper
    return decorator


# ========================================
# 입력 검증 헬퍼 함수
# ========================================

def _sanitize_filename(filename: str) -> str:
    """파일명에서 위험한 문자 제거"""
    if not filename:
        return "unnamed"
    # 허용: 알파벳, 숫자, 한글, 언더스코어, 하이픈, 점, 공백
    sanitized = re.sub(r'[^\w\s가-힣.\-]', '_', filename, flags=re.UNICODE)
    # 연속된 언더스코어 정리
    sanitized = re.sub(r'_+', '_', sanitized)
    return sanitized.strip('_') or "unnamed"


def _validate_file_path(file_path: str, allowed_extensions: List[str] = None, require_temp_dir: bool = True) -> tuple:
    """
    파일 경로 검증 (보안 강화)

    Args:
        file_path: 검증할 파일 경로
        allowed_extensions: 허용된 확장자 리스트
        require_temp_dir: temp 디렉토리 내부 경로만 허용 (기본: True)

    Returns: (is_valid: bool, error_message: str or None)
    """
    if not file_path:
        return False, "파일 경로가 비어있습니다"

    # Path traversal 공격 방어
    try:
        path = Path(file_path).resolve()
        # 상대 경로로 상위 디렉토리 접근 시도 감지
        if ".." in file_path:
            logger.warning(f"Path traversal attempt detected: {file_path}")
            return False, "잘못된 파일 경로입니다"
    except Exception:
        return False, "파일 경로를 해석할 수 없습니다"

    # temp 디렉토리 내부 경로만 허용 (업로드된 파일만 접근 가능)
    if require_temp_dir:
        # 허용된 디렉토리: temp/<user_id>/ 패턴
        temp_dir = (PROJECT_ROOT / "temp").resolve()
        try:
            # path가 temp_dir의 하위 경로인지 확인
            path.relative_to(temp_dir)
        except ValueError:
            logger.warning(f"Access to file outside temp directory blocked: {file_path}")
            return False, "허용되지 않은 경로입니다. 업로드된 파일만 접근할 수 있습니다."

    # 확장자 검증
    if allowed_extensions:
        ext = path.suffix.lower()
        if ext not in [e.lower() if e.startswith('.') else f'.{e.lower()}' for e in allowed_extensions]:
            return False, f"허용되지 않은 파일 형식입니다. 허용: {', '.join(allowed_extensions)}"

    return True, None


def _validate_numeric_param(value: Any, param_name: str, min_val: float = None, max_val: float = None) -> tuple:
    """
    숫자 파라미터 검증
    Returns: (is_valid: bool, validated_value: float or None, error_message: str or None)
    """
    try:
        num_value = float(value)
        if min_val is not None and num_value < min_val:
            return False, None, f"{param_name}은(는) {min_val} 이상이어야 합니다"
        if max_val is not None and num_value > max_val:
            return False, None, f"{param_name}은(는) {max_val} 이하여야 합니다"
        return True, num_value, None
    except (TypeError, ValueError):
        return False, None, f"{param_name}은(는) 유효한 숫자여야 합니다"


def _normalize_text(text: str) -> str:
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()


def _match_underwriter_category(text: str, category: str) -> tuple:
    if not category:
        return True, []
    keywords = UNDERWRITER_CATEGORY_KEYWORDS.get(category)
    if not keywords:
        return False, []

    text_lower = text.lower()
    any_terms = [t.lower() for t in keywords.get("any", []) if t]
    all_terms = [t.lower() for t in keywords.get("all", []) if t]

    if all_terms and not all(term in text_lower for term in all_terms):
        return False, []
    if any_terms and not any(term in text_lower for term in any_terms):
        return False, []

    matched = [term for term in any_terms if term in text_lower]
    matched += [term for term in all_terms if term in text_lower and term not in matched]
    return True, matched


def _extract_snippet(text: str, terms: List[str], max_chars: int) -> str:
    if not text:
        return ""

    text_clean = _normalize_text(text)
    if not text_clean:
        return ""

    if terms:
        text_lower = text_clean.lower()
        for term in terms:
            term_lower = term.lower()
            idx = text_lower.find(term_lower)
            if idx >= 0:
                start = max(0, idx - 120)
                end = min(len(text_clean), idx + max_chars)
                return text_clean[start:end].strip()

    return text_clean[:max_chars].strip()


def _split_sentences(text: str) -> List[str]:
    if not text:
        return []
    normalized = text.replace("\n", " ")
    parts = re.split(r"(?<=[.!?])\s+|(?<=\.)\s+|(?<=다\.)\s+", normalized)
    return [_normalize_text(p) for p in parts if _normalize_text(p)]


def _extract_market_size_sentences(text: str, max_sentences: int = 2) -> str:
    if not text:
        return ""
    candidates = []
    for sentence in _split_sentences(text):
        if len(candidates) >= max_sentences:
            break
        if any(keyword in sentence for keyword in ["시장", "TAM", "SAM", "SOM", "CAGR", "연평균", "성장률", "규모"]):
            if re.search(r"\d", sentence):
                candidates.append(sentence)
    if not candidates:
        return ""
    return " ".join(candidates)


def _extract_numeric_phrases(text: str) -> List[str]:
    if not text:
        return []
    pattern = re.compile(
        r"\d+(?:,\d{3})*(?:\.\d+)?\s*(?:조|억|억원|백만|백만원|천만|만원|원|달러|usd|백만달러|억달러|톤|t|kg|%)",
        re.IGNORECASE,
    )
    return pattern.findall(text)


def _generalize_underwriter_text(text: str, corp_name: str = None) -> str:
    if not text:
        return ""
    generalized = text
    if corp_name:
        generalized = re.sub(re.escape(corp_name), "{회사}", generalized)
    generalized = re.sub(r"\b20\d{2}년\b", "{연도}", generalized)
    generalized = re.sub(r"\b20\d{2}\b", "{연도}", generalized)
    generalized = re.sub(
        r"\d+(?:,\d{3})*(?:\.\d+)?\s*(?:조|억|억원|백만|백만원|천만|만원|원|달러|USD|백만달러|억달러)\b",
        "{금액}",
        generalized,
    )
    generalized = re.sub(r"\d+(?:\.\d+)?\s*%", "{비율}", generalized)
    generalized = re.sub(r"\d{1,3}(?:,\d{3})+", "{숫자}", generalized)
    generalized = re.sub(r"\d+", "{숫자}", generalized)
    generalized = re.sub(r"\s+", " ", generalized).strip()
    return generalized


def _parse_underwriter_jsonl(jsonl_path: str) -> List[Dict[str, Any]]:
    entries = []
    with open(jsonl_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                entries.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return entries


def _normalize_similarity_text(text: str, max_chars: int = 2000) -> str:
    if not text:
        return ""
    normalized = text.lower()
    normalized = re.sub(r"\s+", " ", normalized)
    normalized = re.sub(r"\d", "0", normalized)
    normalized = normalized.strip()
    if max_chars and len(normalized) > max_chars:
        normalized = normalized[:max_chars]
    return normalized


def _char_ngram_counts(text: str, min_n: int = 3, max_n: int = 5) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    length = len(text)
    for n in range(min_n, max_n + 1):
        if length < n:
            continue
        for i in range(length - n + 1):
            token = text[i:i + n]
            counts[token] = counts.get(token, 0) + 1
    return counts


def _build_tfidf_index(texts: List[str], min_n: int = 3, max_n: int = 5) -> tuple:
    counts_list = [_char_ngram_counts(text, min_n=min_n, max_n=max_n) for text in texts]
    df: Dict[str, int] = {}
    for counts in counts_list:
        for term in counts.keys():
            df[term] = df.get(term, 0) + 1

    doc_count = len(counts_list)
    idf = {term: math.log((1 + doc_count) / (1 + freq)) + 1 for term, freq in df.items()}

    vectors = []
    for counts in counts_list:
        vec: Dict[str, float] = {}
        norm = 0.0
        for term, tf in counts.items():
            weight = (1 + math.log(tf)) * idf[term]
            vec[term] = weight
            norm += weight * weight
        vectors.append((vec, math.sqrt(norm)))
    return idf, vectors


def _vectorize_query(text: str, idf: Dict[str, float], min_n: int = 3, max_n: int = 5) -> tuple:
    counts = _char_ngram_counts(text, min_n=min_n, max_n=max_n)
    vec: Dict[str, float] = {}
    norm = 0.0
    for term, tf in counts.items():
        if term not in idf:
            continue
        weight = (1 + math.log(tf)) * idf[term]
        vec[term] = weight
        norm += weight * weight
    return vec, math.sqrt(norm)


def _cosine_similarity(vec_a: Dict[str, float], norm_a: float, vec_b: Dict[str, float], norm_b: float) -> float:
    if norm_a == 0 or norm_b == 0:
        return 0.0
    if len(vec_a) > len(vec_b):
        vec_a, vec_b = vec_b, vec_a
        norm_a, norm_b = norm_b, norm_a
    dot = 0.0
    for term, weight in vec_a.items():
        dot += weight * vec_b.get(term, 0.0)
    return dot / (norm_a * norm_b)


def _get_underwriter_tfidf_index(jsonl_path: str, min_n: int = 3, max_n: int = 5, max_text_chars: int = 2000):
    cache = _UNDERWRITER_TFIDF_CACHE
    try:
        stat = os.stat(jsonl_path)
    except OSError:
        return None, "JSONL 파일 정보를 읽을 수 없습니다"

    if (
        cache["path"] == jsonl_path
        and cache["mtime"] == stat.st_mtime
        and cache["size"] == stat.st_size
        and cache["min_n"] == min_n
        and cache["max_n"] == max_n
        and cache["max_text_chars"] == max_text_chars
        and cache["idf"] is not None
        and cache["vectors"] is not None
        and cache["entries"] is not None
        and cache["texts"] is not None
    ):
        return cache, None

    entries = _parse_underwriter_jsonl(jsonl_path)
    texts = []
    for entry in entries:
        combined = f"{entry.get('section_title', '')}\n{entry.get('section_text', '')}"
        texts.append(_normalize_similarity_text(combined, max_chars=max_text_chars))

    idf, vectors = _build_tfidf_index(texts, min_n=min_n, max_n=max_n)

    cache.update({
        "path": jsonl_path,
        "mtime": stat.st_mtime,
        "size": stat.st_size,
        "min_n": min_n,
        "max_n": max_n,
        "max_text_chars": max_text_chars,
        "idf": idf,
        "vectors": vectors,
        "entries": entries,
        "texts": texts,
    })
    return cache, None


def register_tools() -> List[Dict[str, Any]]:
    """에이전트가 사용할 도구 등록"""

    return [
        {
            "name": "read_excel_as_text",
            "description": "엑셀 파일을 텍스트로 변환하여 읽습니다. 모든 시트의 내용을 텍스트 형식으로 반환하므로, 엑셀 구조가 다양해도 유연하게 대응할 수 있습니다. 이 도구로 먼저 엑셀 내용을 읽은 후, 필요한 정보를 파악하세요.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "excel_path": {
                        "type": "string",
                        "description": "읽을 엑셀 파일 경로"
                    },
                    "sheet_names": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "읽을 시트 이름 리스트 (선택사항, 없으면 모든 시트)"
                    },
                    "max_rows": {
                        "type": "integer",
                        "description": "각 시트에서 읽을 최대 행 수 (기본값: 50)"
                    }
                },
                "required": ["excel_path"]
            }
        },
        {
            "name": "analyze_excel",
            "description": "투자 검토 엑셀 파일을 자동으로 분석하여 투자조건, IS요약(연도별 당기순이익), Cap Table(총발행주식수)을 추출합니다. 일반적인 엑셀 구조에서 작동하지만, 구조가 특이하면 read_excel_as_text를 사용하세요.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "excel_path": {
                        "type": "string",
                        "description": "분석할 엑셀 파일 경로"
                    }
                },
                "required": ["excel_path"]
            }
        },
        {
            "name": "analyze_and_generate_projection",
            "description": "엑셀 파일을 분석하고 즉시 Exit 프로젝션을 생성합니다. 파일에서 투자 조건과 재무 데이터를 자동으로 추출한 후, 지정된 연도와 PER 배수로 Exit 시나리오를 계산하여 새로운 엑셀 파일을 생성합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "excel_path": {
                        "type": "string",
                        "description": "분석할 투자검토 엑셀 파일 경로"
                    },
                    "target_year": {
                        "type": "integer",
                        "description": "Exit 목표 연도 (예: 2028, 2030)"
                    },
                    "per_multiples": {
                        "type": "array",
                        "items": {"type": "number"},
                        "description": "PER 배수 리스트 (예: [10, 20, 30])"
                    },
                    "company_name": {
                        "type": "string",
                        "description": "회사명 (선택사항)"
                    },
                    "output_filename": {
                        "type": "string",
                        "description": "출력 파일명 (선택사항, 기본값: exit_projection_YYYYMMDD_HHMMSS.xlsx)"
                    }
                },
                "required": ["excel_path", "target_year", "per_multiples"]
            }
        },
        {
            "name": "analyze_company_diagnosis_sheet",
            "description": "MYSC 기업현황 진단시트(xlsx)를 분석하여 회사 기본정보, 체크리스트 응답(예/아니오), KPI/마일스톤, 항목별 점수(자가진단 기반)를 구조화된 형태로 반환합니다. 컨설턴트용 분석보고서 작성 전에 반드시 이 도구로 진단시트를 파악하세요.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "excel_path": {
                        "type": "string",
                        "description": "분석할 기업현황 진단시트 엑셀 파일 경로"
                    }
                },
                "required": ["excel_path"]
            }
        },
        {
            "name": "create_company_diagnosis_draft",
            "description": "템플릿 업로드 없이도 대화를 통해 기업현황 진단시트를 작성할 수 있도록, 사용자별 임시 드래프트(JSON)를 생성합니다. 이후 update_company_diagnosis_draft로 내용을 채우고 generate_company_diagnosis_sheet_from_draft로 엑셀을 생성하세요.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "user_id": {
                        "type": "string",
                        "description": "사용자 고유 ID (temp/<user_id>/에 드래프트 생성)"
                    },
                    "template_version": {
                        "type": "string",
                        "description": "템플릿 버전 (기본: 2025)",
                        "default": "2025"
                    }
                },
                "required": ["user_id"]
            }
        },
        {
            "name": "update_company_diagnosis_draft",
            "description": "기업현황 진단시트 드래프트(JSON)에 사용자 응답을 반영하고, 다음으로 질문해야 할 항목(필드/체크리스트 배치)과 진행률을 반환합니다. PII를 포함할 수 있으므로 필요한 값만 최소한으로 전달하세요.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "draft_path": {
                        "type": "string",
                        "description": "create_company_diagnosis_draft로 생성된 드래프트 경로 (temp 내부 .json)"
                    },
                    "company_info": {
                        "type": "object",
                        "description": "기업 기본정보 (선택)",
                        "properties": {
                            "company_name": {"type": "string"},
                            "representative_name": {"type": "string"},
                            "email": {"type": "string"},
                            "phone": {"type": "string"},
                            "incorporation_date": {"type": ["string", "number"]},
                            "business_registration_number": {"type": "string"},
                            "business_type": {"type": "string"},
                            "hq_address": {"type": "string"},
                            "branch_address": {"type": "string"}
                        }
                    },
                    "employees_financials": {
                        "type": "object",
                        "description": "인력/재무 정보 (선택)",
                        "properties": {
                            "employees_full_time": {"type": ["string", "number"]},
                            "employees_contract": {"type": ["string", "number"]},
                            "revenue_2024": {"type": ["string", "number"]},
                            "revenue_2023": {"type": ["string", "number"]},
                            "equity_total": {"type": ["string", "number"]},
                            "certification": {"type": "string"}
                        }
                    },
                    "investment": {
                        "type": "object",
                        "description": "투자 정보 (선택)",
                        "properties": {
                            "investment_history": {"type": "string"},
                            "desired_investment_amount": {"type": "string"},
                            "pre_money_valuation": {"type": "string"}
                        }
                    },
                    "kpi": {
                        "type": "object",
                        "description": "KPI/마일스톤 (선택)",
                        "properties": {
                            "service_intro": {"type": "string"},
                            "revenue_model": {"type": "string"},
                            "core_customer": {"type": "string"},
                            "kpis": {
                                "type": "array",
                                "description": "정량 KPI (최대 5개 권장)",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "name": {"type": "string"},
                                        "current": {"type": "string"},
                                        "target": {"type": "string"}
                                    },
                                    "required": ["name"]
                                }
                            },
                            "milestone": {
                                "type": "object",
                                "properties": {
                                    "domestic_plan_2025": {"type": "string"},
                                    "global_plan_2025": {"type": "string"},
                                    "long_term_goal_3y": {"type": "string"},
                                    "program_expectation": {"type": "string"},
                                    "growth_goal": {"type": "string"},
                                    "concerns": {"type": "string"}
                                }
                            }
                        }
                    },
                    "checklist_answers": {
                        "type": "array",
                        "description": "체크리스트 응답 (선택) - 예/아니오 + 근거/요청(선택)",
                        "items": {
                            "type": "object",
                            "properties": {
                                "id": {"type": "string"},
                                "status": {"type": "string", "enum": ["예", "아니오"]},
                                "detail": {"type": "string"}
                            },
                            "required": ["id", "status"]
                        }
                    }
                },
                "required": ["draft_path"]
            }
        },
        {
            "name": "generate_company_diagnosis_sheet_from_draft",
            "description": "대화로 수집한 드래프트(JSON)를 기반으로 기업현황 진단시트 엑셀(xlsx)을 생성합니다. 생성된 엑셀은 temp 디렉토리에 저장되며, 이후 analyze_company_diagnosis_sheet / write_company_diagnosis_report로 확장할 수 있습니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "draft_path": {
                        "type": "string",
                        "description": "드래프트 JSON 경로 (temp 내부)"
                    },
                    "output_filename": {
                        "type": "string",
                        "description": "출력 파일명 (선택, 기본: diagnosis_sheet_YYYYMMDD_HHMMSS.xlsx)"
                    }
                },
                "required": ["draft_path"]
            }
        },
        {
            "name": "write_company_diagnosis_report",
            "description": "기업현황 진단시트의 '(컨설턴트용) 분석보고서' 시트에 컨설턴트 요약/개선사항/점수를 반영한 새 엑셀 파일을 생성합니다. 원본 파일은 수정하지 않고 temp 디렉토리에 결과 파일을 저장합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "excel_path": {
                        "type": "string",
                        "description": "원본 기업현황 진단시트 엑셀 파일 경로 (temp 내부)"
                    },
                    "company_name": {
                        "type": "string",
                        "description": "기업명 (선택사항, 비어있으면 원본에서 추출)"
                    },
                    "report_datetime": {
                        "type": "string",
                        "description": "작성일시 문자열 (선택사항, 비어있으면 현재 시각)"
                    },
                    "scores": {
                        "type": "object",
                        "description": "항목별 점수 (문제/솔루션/사업화/자금조달/팀/조직/임팩트)",
                        "properties": {
                            "문제": {"type": "number"},
                            "솔루션": {"type": "number"},
                            "사업화": {"type": "number"},
                            "자금조달": {"type": "number"},
                            "팀/조직": {"type": "number"},
                            "임팩트": {"type": "number"}
                        }
                    },
                    "summary_text": {
                        "type": "string",
                        "description": "기업 상황 요약/기업진단 내용 (분석보고서 본문)"
                    },
                    "improvement_text": {
                        "type": "string",
                        "description": "개선 필요사항 (분석보고서 본문)"
                    },
                    "output_filename": {
                        "type": "string",
                        "description": "출력 파일명 (선택사항, 기본값: diagnosis_report_YYYYMMDD_HHMMSS.xlsx)"
                    }
                },
                "required": ["excel_path", "scores", "summary_text", "improvement_text"]
            }
        },
        {
            "name": "calculate_valuation",
            "description": "다양한 방법론으로 기업가치를 계산합니다 (PER, EV/Revenue, EV/EBITDA 등)",
            "input_schema": {
                "type": "object",
                "properties": {
                    "method": {
                        "type": "string",
                        "enum": ["per", "ev_revenue", "ev_ebitda"],
                        "description": "밸류에이션 방법론"
                    },
                    "base_value": {
                        "type": "number",
                        "description": "기준 값 (순이익, 매출, EBITDA 등)"
                    },
                    "multiple": {
                        "type": "number",
                        "description": "적용할 배수"
                    }
                },
                "required": ["method", "base_value", "multiple"]
            }
        },
        {
            "name": "calculate_dilution",
            "description": "SAFE 전환, 신규 투자 라운드 등으로 인한 지분 희석 효과를 계산합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "event_type": {
                        "type": "string",
                        "enum": ["safe", "new_round", "call_option"],
                        "description": "희석 이벤트 종류"
                    },
                    "current_shares": {
                        "type": "number",
                        "description": "현재 총 발행주식수"
                    },
                    "event_details": {
                        "type": "object",
                        "description": "이벤트 상세 정보 (investment_amount, valuation_cap 등)"
                    }
                },
                "required": ["event_type", "current_shares", "event_details"]
            }
        },
        {
            "name": "calculate_irr",
            "description": "현금흐름 기반으로 IRR(내부수익률)과 멀티플을 계산합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "cash_flows": {
                        "type": "array",
                        "description": "현금흐름 리스트 [{year: 2025, amount: -300000000}, {year: 2029, amount: 3150000000}]",
                        "items": {
                            "type": "object",
                            "properties": {
                                "year": {"type": "number"},
                                "amount": {"type": "number"}
                            }
                        }
                    }
                },
                "required": ["cash_flows"]
            }
        },
        {
            "name": "generate_exit_projection",
            "description": "Exit 프로젝션 엑셀 파일을 생성합니다 (basic/advanced/complete 중 선택)",
            "input_schema": {
                "type": "object",
                "properties": {
                    "projection_type": {
                        "type": "string",
                        "enum": ["basic", "advanced", "complete"],
                        "description": "프로젝션 타입 (basic: 기본, advanced: 부분매각+NPV, complete: SAFE+콜옵션)"
                    },
                    "parameters": {
                        "type": "object",
                        "description": "생성에 필요한 파라미터 (investment_amount, company_name, per_multiples 등)"
                    }
                },
                "required": ["projection_type", "parameters"]
            }
        },
        # ========================================
        # Underwriter opinion 데이터 도구
        # ========================================
        {
            "name": "search_underwriter_opinion",
            "description": "인수인의견(분석기관의 평가의견) JSONL 데이터에서 특정 카테고리/기업/키워드에 맞는 문장과 패턴을 추출합니다. 시장규모 근거, 비교기업 선정, 공모가 산정 논리 등 반복 패턴을 일반화할 때 사용하세요.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "jsonl_path": {
                        "type": "string",
                        "description": "underwriter_opinion.jsonl 경로 (선택, 기본: 프로젝트 temp 경로)"
                    },
                    "category": {
                        "type": "string",
                        "description": "카테고리 (market_size, valuation, comparables, demand_forecast, risk 중 선택)"
                    },
                    "corp_name": {
                        "type": "string",
                        "description": "기업명 필터 (선택)"
                    },
                    "query": {
                        "type": "string",
                        "description": "추가 검색어 (선택)"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "최대 결과 수 (기본값: 5)"
                    },
                    "max_chars": {
                        "type": "integer",
                        "description": "문장/본문 최대 길이 (기본값: 800)"
                    },
                    "min_section_length": {
                        "type": "integer",
                        "description": "섹션 길이 최소값 (기본값: 0)"
                    },
                    "return_patterns": {
                        "type": "boolean",
                        "description": "일반화된 패턴 문장 반환 여부 (기본값: true)"
                    },
                    "output_filename": {
                        "type": "string",
                        "description": "결과 저장 파일명 (선택, temp/에 JSON 저장)"
                    }
                }
            }
        },
        {
            "name": "search_underwriter_opinion_similar",
            "description": "인수인의견 JSONL 데이터에서 입력 질의와 유사한 문장을 유사도 기반으로 검색합니다. 키워드 매칭이 어려운 경우 보완용으로 사용하세요.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "유사도 검색 질의 (시장/기술/산업 관련 문장 권장)"
                    },
                    "jsonl_path": {
                        "type": "string",
                        "description": "underwriter_opinion.jsonl 경로 (선택, 기본: 프로젝트 temp 경로)"
                    },
                    "category": {
                        "type": "string",
                        "description": "카테고리 필터 (market_size, valuation, comparables, demand_forecast, risk 중 선택, 선택사항)"
                    },
                    "corp_name": {
                        "type": "string",
                        "description": "기업명 필터 (선택)"
                    },
                    "top_k": {
                        "type": "integer",
                        "description": "상위 결과 수 (기본값: 5)"
                    },
                    "min_score": {
                        "type": "number",
                        "description": "유사도 최소 점수 (기본값: 0.05)"
                    },
                    "max_chars": {
                        "type": "integer",
                        "description": "문장/본문 최대 길이 (기본값: 800)"
                    },
                    "min_section_length": {
                        "type": "integer",
                        "description": "섹션 길이 최소값 (기본값: 0)"
                    },
                    "max_text_chars": {
                        "type": "integer",
                        "description": "유사도 계산용 텍스트 최대 길이 (기본값: 2000)"
                    },
                    "return_patterns": {
                        "type": "boolean",
                        "description": "일반화된 패턴 문장 반환 여부 (기본값: true)"
                    },
                    "output_filename": {
                        "type": "string",
                        "description": "결과 저장 파일명 (선택, temp/에 JSON 저장)"
                    }
                },
                "required": ["query"]
            }
        },
        {
            "name": "extract_pdf_market_evidence",
            "description": "PDF에서 시장규모 근거가 될 수 있는 문장을 페이지 번호와 함께 추출합니다. 숫자/단위가 포함된 문장을 우선 수집합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "pdf_path": {
                        "type": "string",
                        "description": "읽을 PDF 파일 경로"
                    },
                    "max_pages": {
                        "type": "integer",
                        "description": "읽을 최대 페이지 수 (기본값: 30)"
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "최대 결과 수 (기본값: 20)"
                    },
                    "keywords": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "추가 키워드 (선택)"
                    }
                },
                "required": ["pdf_path"]
            }
        },
        {
            "name": "fetch_underwriter_opinion_data",
            "description": "인수인의견 원천 데이터를 API에서 수집하여 JSONL로 생성합니다. API 키가 필요하며, 결과는 temp/ 하위에 저장됩니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "start_date": {
                        "type": "string",
                        "description": "시작일 (YYYYMMDD, 선택)"
                    },
                    "end_date": {
                        "type": "string",
                        "description": "종료일 (YYYYMMDD, 선택)"
                    },
                    "lookback_days": {
                        "type": "integer",
                        "description": "종료일 기준 조회 기간 (일, 기본값: 365)"
                    },
                    "out_dir": {
                        "type": "string",
                        "description": "출력 디렉토리 (선택, temp/ 하위로 제한)"
                    },
                    "max_items": {
                        "type": "integer",
                        "description": "최대 처리 건수 (기본값: 200)"
                    },
                    "last_only": {
                        "type": "boolean",
                        "description": "최종 보고서만 수집 (기본값: true)"
                    },
                    "include_corrections": {
                        "type": "boolean",
                        "description": "정정/확정 보고서 포함 여부 (기본값: false)"
                    },
                    "min_length": {
                        "type": "integer",
                        "description": "섹션 최소 길이 (기본값: 600)"
                    },
                    "min_score": {
                        "type": "number",
                        "description": "섹션 최소 점수 (기본값: 3.0)"
                    },
                    "api_key": {
                        "type": "string",
                        "description": "API 키 (선택, 없으면 환경변수 사용)"
                    }
                }
            }
        },
        # ========================================
        # Peer PER 분석 도구
        # ========================================
        {
            "name": "read_pdf_as_text",
            "description": "PDF 파일(기업 소개서, IR 자료, 사업계획서 등)을 Claude Vision으로 분석합니다. 테이블 구조를 보존하고 재무제표를 자동으로 추출합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "pdf_path": {
                        "type": "string",
                        "description": "읽을 PDF 파일 경로"
                    },
                    "max_pages": {
                        "type": "integer",
                        "description": "읽을 최대 페이지 수 (기본값: 30)"
                    },
                    "output_mode": {
                        "type": "string",
                        "enum": ["text_only", "structured", "tables_only"],
                        "description": "출력 모드 (text_only: 텍스트만, structured: 전체 구조+재무제표, tables_only: 테이블만)"
                    },
                    "extract_financial_tables": {
                        "type": "boolean",
                        "description": "재무제표 테이블 자동 추출 여부 (IS/BS/CF)"
                    }
                },
                "required": ["pdf_path"]
            }
        },
        {
            "name": "parse_pdf_dolphin",
            "description": "Claude Vision을 사용하여 PDF를 구조화된 형태로 파싱합니다. 테이블, 재무제표, Cap Table을 자동 인식합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "pdf_path": {
                        "type": "string",
                        "description": "분석할 PDF 파일 경로"
                    },
                    "max_pages": {
                        "type": "integer",
                        "description": "분석할 최대 페이지 수 (기본값: 30)"
                    },
                    "output_mode": {
                        "type": "string",
                        "enum": ["text_only", "structured", "tables_only"],
                        "description": "출력 모드"
                    },
                    "extract_financial_tables": {
                        "type": "boolean",
                        "description": "재무제표 테이블 자동 추출 여부 (기본값: true)"
                    }
                },
                "required": ["pdf_path"]
            }
        },
        {
            "name": "extract_pdf_tables",
            "description": "PDF에서 테이블만 추출합니다. 재무제표(IS/BS/CF), Cap Table 등을 구조화된 데이터로 반환합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "pdf_path": {
                        "type": "string",
                        "description": "PDF 파일 경로"
                    },
                    "max_pages": {
                        "type": "integer",
                        "description": "처리할 최대 페이지 수 (기본값: 50)"
                    }
                },
                "required": ["pdf_path"]
            }
        },
        # 대화형 투자 분석 도구
        {
            "name": "start_analysis_session",
            "description": "대화형 투자 분석 세션을 시작합니다. 여러 파일과 텍스트 입력을 받아서 점진적으로 분석을 완성합니다. 세션 ID를 반환하며, 이후 add_supplementary_data와 함께 사용합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "initial_pdf_path": {
                        "type": "string",
                        "description": "초기 분석할 PDF 파일 경로 (선택)"
                    },
                    "max_pages": {
                        "type": "integer",
                        "description": "PDF 분석할 최대 페이지 수 (기본값: 30)"
                    }
                },
                "required": []
            }
        },
        {
            "name": "add_supplementary_data",
            "description": "기존 분석 세션에 추가 데이터를 입력합니다. PDF 파일 또는 텍스트(재무 데이터, Cap Table, 투자 조건 등)를 추가할 수 있습니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "session_id": {
                        "type": "string",
                        "description": "분석 세션 ID (start_analysis_session에서 반환)"
                    },
                    "pdf_path": {
                        "type": "string",
                        "description": "추가할 PDF 파일 경로 (선택)"
                    },
                    "text_input": {
                        "type": "string",
                        "description": "추가할 텍스트 데이터 (선택). 예: '2024년 매출 100억, 순이익 15억'"
                    },
                    "data_type": {
                        "type": "string",
                        "enum": ["financial", "cap_table", "investment_terms", "general"],
                        "description": "텍스트 데이터 유형 (기본값: general)"
                    },
                    "max_pages": {
                        "type": "integer",
                        "description": "PDF 분석할 최대 페이지 수 (기본값: 30)"
                    }
                },
                "required": ["session_id"]
            }
        },
        {
            "name": "get_analysis_status",
            "description": "분석 세션의 현재 상태를 확인합니다. 어떤 데이터가 수집되었고, 어떤 데이터가 부족한지 알려줍니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "session_id": {
                        "type": "string",
                        "description": "분석 세션 ID"
                    }
                },
                "required": ["session_id"]
            }
        },
        {
            "name": "complete_analysis",
            "description": "분석 세션을 완료하고 최종 분석 결과를 반환합니다. 필수 데이터가 부족하면 부족한 항목을 안내합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "session_id": {
                        "type": "string",
                        "description": "분석 세션 ID"
                    }
                },
                "required": ["session_id"]
            }
        },
        {
            "name": "get_stock_financials",
            "description": "yfinance를 사용하여 상장 기업의 재무 지표를 조회합니다. PER, PSR, 매출, 영업이익률, 시가총액 등을 반환합니다. 한국 주식은 티커 뒤에 .KS(KOSPI) 또는 .KQ(KOSDAQ)를 붙입니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "ticker": {
                        "type": "string",
                        "description": "주식 티커 심볼 (예: AAPL, MSFT, 005930.KS, 035720.KQ)"
                    }
                },
                "required": ["ticker"]
            }
        },
        {
            "name": "analyze_peer_per",
            "description": "여러 Peer 기업의 PER을 일괄 조회하고 비교 분석합니다. 평균, 중간값, 범위를 계산하여 적정 PER 배수를 제안합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "tickers": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "비교할 기업 티커 리스트 (예: ['AAPL', 'MSFT', 'GOOGL'])"
                    },
                    "include_forward_per": {
                        "type": "boolean",
                        "description": "Forward PER 포함 여부 (기본값: true)"
                    }
                },
                "required": ["tickers"]
            }
        },
        {
            "name": "query_investment_portfolio",
            "description": "투자기업 CSV(투자기업-Grid view.csv)를 조회합니다. 질문/필터/정렬 조건을 주면 해당하는 기업 리스트와 요약을 반환합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "검색어 (기업명/서비스/키워드 등)"
                    },
                    "filters": {
                        "type": "object",
                        "description": "컬럼명-값 맵 (값은 문자열 또는 문자열 리스트). 예: {\"본점 소재지\": \"강원\", \"카테고리1\": [\"푸드\", \"환경\"]}"
                    },
                    "limit": {
                        "type": "integer",
                        "description": "최대 조회할 결과 개수 (기본: 5)"
                    },
                    "sort_by": {
                        "type": "string",
                        "description": "정렬 기준 컬럼명"
                    },
                    "sort_order": {
                        "type": "string",
                        "enum": ["asc", "desc"],
                        "description": "정렬 방향 (기본: desc)"
                    }
                }
            }
        },
        # ========================================
        # 스타트업 발굴 지원 도구
        # ========================================
        {
            "name": "analyze_government_policy",
            "description": "정부 정책 PDF/아티클을 분석하여 핵심 정책 방향, 예산 배분, 타겟 산업을 추출합니다. 여러 PDF를 한번에 분석할 수 있습니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "pdf_paths": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "분석할 PDF 파일 경로 리스트"
                    },
                    "focus_keywords": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "집중 분석할 키워드 (선택)"
                    },
                    "max_pages_per_pdf": {
                        "type": "integer",
                        "description": "PDF당 최대 페이지 수 (기본: 30)"
                    }
                },
                "required": ["pdf_paths"]
            }
        },
        {
            "name": "search_iris_plus_metrics",
            "description": "IRIS+ 임팩트 메트릭 카탈로그에서 키워드/카테고리로 관련 지표를 검색합니다. SDG(지속가능발전목표) 연계 정보도 반환합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "검색 키워드 (예: clean energy, 탄소중립)"
                    },
                    "category": {
                        "type": "string",
                        "enum": ["environmental", "social", "governance"],
                        "description": "카테고리 필터 (선택)"
                    },
                    "sdg_filter": {
                        "type": "array",
                        "items": {"type": "integer"},
                        "description": "SDG 번호 필터 (예: [7, 13])"
                    },
                    "top_k": {
                        "type": "integer",
                        "description": "상위 결과 수 (기본: 10)"
                    }
                },
                "required": ["query"]
            }
        },
        {
            "name": "map_policy_to_iris",
            "description": "정책 분석 결과를 IRIS+ 메트릭에 자동 매핑합니다. 정책-임팩트 연관도 점수를 계산합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "policy_themes": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "정책 테마 리스트 (예: ['탄소중립', '디지털전환'])"
                    },
                    "target_industries": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "타겟 산업 리스트 (선택)"
                    },
                    "min_relevance_score": {
                        "type": "number",
                        "description": "최소 연관도 점수 (기본: 0.3)"
                    }
                },
                "required": ["policy_themes"]
            }
        },
        {
            "name": "generate_industry_recommendation",
            "description": "정책 분석 + IRIS+ 매핑 결과를 종합하여 유망 산업/스타트업 영역을 추천합니다. 근거와 함께 순위를 제공합니다.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "policy_analysis": {
                        "type": "object",
                        "description": "analyze_government_policy 결과"
                    },
                    "iris_mapping": {
                        "type": "object",
                        "description": "map_policy_to_iris 결과"
                    },
                    "interest_areas": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "사용자 관심 분야 (선택)"
                    },
                    "recommendation_count": {
                        "type": "integer",
                        "description": "추천 개수 (기본: 5)"
                    },
                    "document_weight": {
                        "type": "number",
                        "description": "문서 기반 가중치 (0~1, 기본: 0.7)"
                    }
                },
                "required": ["policy_analysis", "iris_mapping"]
            }
        }
    ]


# === Tool Execution Functions ===

def execute_read_excel_as_text(
    excel_path: str,
    sheet_names: List[str] = None,
    max_rows: int = 50
) -> Dict[str, Any]:
    """엑셀 파일을 텍스트로 변환하여 읽기"""
    # 입력 검증: 파일 경로 (temp 디렉토리 내부만 허용)
    is_valid, error = _validate_file_path(excel_path, allowed_extensions=['.xlsx', '.xls'], require_temp_dir=True)
    if not is_valid:
        return {"success": False, "error": error}

    wb = None
    try:
        from openpyxl import load_workbook

        # 파일 존재 확인
        if not os.path.exists(excel_path):
            return {"success": False, "error": f"파일을 찾을 수 없습니다: {excel_path}"}

        wb = load_workbook(excel_path, data_only=True)

        sheets_data = {}
        target_sheets = sheet_names if sheet_names else wb.sheetnames

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            sheet_text = []

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx > max_rows:
                    break

                # None이 아닌 값들만 필터링
                row_values = [str(cell) if cell is not None else "" for cell in row]

                # 빈 행 스킵
                if not any(val.strip() for val in row_values):
                    continue

                # 행 텍스트 생성
                row_text = " | ".join(row_values[:15])  # 처음 15개 컬럼만
                sheet_text.append(f"Row {row_idx}: {row_text}")

            sheets_data[sheet_name] = "\n".join(sheet_text)

        # 텍스트 결과 생성
        result_text = ""
        for sheet_name, content in sheets_data.items():
            result_text += f"\n{'='*60}\n"
            result_text += f"시트: {sheet_name}\n"
            result_text += f"{'='*60}\n"
            result_text += content + "\n"

        logger.info(f"Excel read successfully: {excel_path} ({len(sheets_data)} sheets)")
        return {
            "success": True,
            "file_path": excel_path,
            "sheets": list(sheets_data.keys()),
            "content": result_text,
            "total_sheets": len(sheets_data)
        }

    except FileNotFoundError:
        return {"success": False, "error": f"파일을 찾을 수 없습니다: {excel_path}"}
    except PermissionError:
        return {"success": False, "error": f"파일 접근 권한이 없습니다: {excel_path}"}
    except Exception as e:
        logger.error(f"Failed to read excel {excel_path}: {e}", exc_info=True)
        return {"success": False, "error": f"엑셀 파일 읽기 실패: {str(e)}"}
    finally:
        if wb is not None:
            wb.close()


def execute_analyze_excel(excel_path: str) -> Dict[str, Any]:
    """엑셀 파일 분석 실행 - openpyxl로 직접 읽기"""
    # 입력 검증: 파일 경로 (temp 디렉토리 내부만 허용)
    is_valid, error = _validate_file_path(excel_path, allowed_extensions=['.xlsx', '.xls'], require_temp_dir=True)
    if not is_valid:
        return {"success": False, "error": error}

    from openpyxl import load_workbook

    wb = None
    try:
        # 파일 존재 확인
        if not os.path.exists(excel_path):
            return {"success": False, "error": f"파일을 찾을 수 없습니다: {excel_path}"}

        # 엑셀 파일 열기
        wb = load_workbook(excel_path, data_only=True)

        result = {
            "success": True,
            "file_path": excel_path,
            "sheets": wb.sheetnames,
            "investment_terms": {},
            "income_statement": {},
            "cap_table": {}
        }

        # IS요약 시트에서 순이익 데이터 추출
        is_sheet = None
        for sheet_name in wb.sheetnames:
            if 'IS' in sheet_name or '손익' in sheet_name:
                is_sheet = wb[sheet_name]
                break

        if is_sheet:
            # 헤더 행 찾기 (구분, 2021년, 2022년... 형태)
            year_row_idx = None
            year_cols = {}

            for row_idx, row in enumerate(is_sheet.iter_rows(min_row=1, max_row=10), start=1):
                for col_idx, cell in enumerate(row):
                    if cell.value and isinstance(cell.value, str) and '년' in cell.value:
                        try:
                            year_val = int(cell.value.replace('년', '').replace(',', ''))
                            if 2020 <= year_val <= 2040:
                                year_row_idx = row_idx
                                year_cols[year_val] = col_idx
                        except ValueError:
                            pass

            # 당기순이익 행 찾기
            net_income_data = {}
            if year_cols:
                for row in is_sheet.iter_rows(min_row=year_row_idx if year_row_idx else 1):
                    first_cell = row[1].value if len(row) > 1 else None  # 2번째 컬럼 확인
                    if first_cell and '당기순이익' in str(first_cell):
                        for year, col_idx in year_cols.items():
                            if col_idx < len(row):
                                value = row[col_idx].value
                                if value and isinstance(value, (int, float)):
                                    net_income_data[year] = int(value)
                        break

            result["income_statement"] = {
                "years": sorted(year_cols.keys()) if year_cols else [],
                "net_income": net_income_data
            }

        # Cap Table에서 총 발행주식수 추출
        cap_sheet = None
        for sheet_name in wb.sheetnames:
            if 'cap' in sheet_name.lower() or '주주' in sheet_name:
                cap_sheet = wb[sheet_name]
                break

        if cap_sheet:
            # "합계" 행에서 주식수 찾기
            for row in cap_sheet.iter_rows():
                first_cell = row[0].value if row else None
                if first_cell and '합계' in str(first_cell):
                    # Incorporation 라운드의 주식수 (4번째 컬럼)
                    if len(row) > 3 and row[3].value and isinstance(row[3].value, (int, float)):
                        incorporation_shares = int(row[3].value)
                        # Seed 라운드 주식수 (7번째 컬럼)
                        seed_shares = 0
                        if len(row) > 6 and row[6].value and isinstance(row[6].value, (int, float)):
                            seed_shares = int(row[6].value)

                        total_shares = incorporation_shares + seed_shares
                        result["cap_table"]["total_shares"] = total_shares
                        result["cap_table"]["incorporation_shares"] = incorporation_shares
                        result["cap_table"]["seed_shares"] = seed_shares
                        break

        # 투자조건 시트에서 투자 정보 추출
        invest_sheet = None
        for sheet_name in wb.sheetnames:
            if '투자조건' in sheet_name:
                invest_sheet = wb[sheet_name]
                break

        if invest_sheet:
            for row_idx, row in enumerate(invest_sheet.iter_rows(min_row=1, max_row=30)):
                # 두 번째 컬럼이 주 정보 컬럼
                if len(row) < 4:
                    continue

                second_cell = row[1].value if row[1] else None
                if not second_cell:
                    continue

                second_val = str(second_cell)

                # 투자금액(원)
                if '투자금액' in second_val and '원' in second_val:
                    # 4번째 컬럼부터 찾기 (투자조건 열)
                    for cell in row[3:]:
                        if cell.value and isinstance(cell.value, (int, float)):
                            result["investment_terms"]["investment_amount"] = int(cell.value)
                            break

                # 투자단가(원)
                if '투자단가' in second_val and '원' in second_val:
                    for cell in row[3:]:
                        if cell.value and isinstance(cell.value, (int, float)):
                            result["investment_terms"]["price_per_share"] = int(cell.value)
                            break

                # 투자주식수
                if '투자주식수' in second_val:
                    for cell in row[3:]:
                        if cell.value and isinstance(cell.value, (int, float)):
                            result["investment_terms"]["shares"] = int(cell.value)
                            break

        logger.info(f"Excel analyzed successfully: {excel_path}")
        return result

    except FileNotFoundError:
        return {"success": False, "error": f"파일을 찾을 수 없습니다: {excel_path}"}
    except PermissionError:
        return {"success": False, "error": f"파일 접근 권한이 없습니다: {excel_path}"}
    except Exception as e:
        logger.error(f"Failed to analyze excel {excel_path}: {e}", exc_info=True)
        return {
            "success": False,
            "error": f"엑셀 파일 분석 실패: {str(e)}"
        }
    finally:
        if wb is not None:
            wb.close()


def execute_calculate_valuation(
    method: str,
    base_value: float,
    multiple: float
) -> Dict[str, Any]:
    """기업가치 계산 실행"""

    enterprise_value = base_value * multiple

    return {
        "success": True,
        "method": method,
        "base_value": base_value,
        "multiple": multiple,
        "enterprise_value": enterprise_value,
        "formatted": f"{enterprise_value:,.0f}원"
    }


def execute_calculate_dilution(
    event_type: str,
    current_shares: float,
    event_details: Dict[str, Any]
) -> Dict[str, Any]:
    """지분 희석 계산 실행"""

    if event_type == "safe":
        safe_amount = event_details.get("safe_amount")
        valuation_cap = event_details.get("valuation_cap")

        new_shares = (safe_amount / valuation_cap) * current_shares

    elif event_type == "new_round":
        investment = event_details.get("investment_amount")
        pre_money = event_details.get("pre_money_valuation")

        new_shares = (investment / pre_money) * current_shares

    elif event_type == "call_option":
        new_shares = 0  # 콜옵션은 희석 없음 (주식 매입)

    else:
        return {
            "success": False,
            "error": f"Unknown event type: {event_type}"
        }

    total_shares = current_shares + new_shares
    dilution_ratio = new_shares / total_shares if total_shares > 0 else 0

    return {
        "success": True,
        "event_type": event_type,
        "current_shares": current_shares,
        "new_shares": new_shares,
        "total_shares": total_shares,
        "dilution_ratio": dilution_ratio,
        "dilution_percentage": f"{dilution_ratio * 100:.2f}%"
    }


def execute_calculate_irr(cash_flows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """IRR 계산 실행"""

    if len(cash_flows) < 2:
        return {
            "success": False,
            "error": "최소 2개의 현금흐름이 필요합니다 (투자 + 회수)"
        }

    # 간단한 IRR 계산 (Newton's method)
    def npv(rate, cfs):
        return sum([
            cf["amount"] / ((1 + rate) ** (cf["year"] - cfs[0]["year"]))
            for cf in cfs
        ])

    # IRR 추정 (초기값 10%)
    rate = 0.1
    for _ in range(100):  # 최대 100번 반복
        npv_value = npv(rate, cash_flows)

        # NPV가 0에 가까우면 종료
        if abs(npv_value) < 1:
            break

        # Newton's method로 업데이트
        delta = 0.0001
        npv_delta = npv(rate + delta, cash_flows)
        derivative = (npv_delta - npv_value) / delta

        if abs(derivative) < 1e-10:
            break

        rate = rate - npv_value / derivative

    # 멀티플 계산
    initial_investment = abs(cash_flows[0]["amount"])
    total_return = sum([cf["amount"] for cf in cash_flows[1:]])
    multiple = total_return / initial_investment if initial_investment > 0 else 0

    # 투자기간
    holding_period = cash_flows[-1]["year"] - cash_flows[0]["year"]

    return {
        "success": True,
        "irr": rate,
        "irr_percentage": f"{rate * 100:.1f}%",
        "multiple": multiple,
        "multiple_formatted": f"{multiple:.2f}x",
        "holding_period": holding_period,
        "cash_flows": cash_flows
    }


def execute_generate_exit_projection(
    projection_type: str,
    parameters: Dict[str, Any]
) -> Dict[str, Any]:
    """Exit 프로젝션 엑셀 생성 실행"""

    # 스크립트 선택 (화이트리스트 방식)
    script_map = {
        "basic": "generate_exit_projection.py",
        "advanced": "generate_advanced_exit_projection.py",
        "complete": "generate_complete_exit_projection.py"
    }

    script_name = script_map.get(projection_type)
    if not script_name:
        return {
            "success": False,
            "error": f"Unknown projection type: {projection_type}. 허용: basic, advanced, complete"
        }

    script_path = PROJECT_ROOT / "scripts" / script_name

    # 허용된 파라미터 키 화이트리스트
    allowed_params = {
        "investment_amount", "price_per_share", "shares", "total_shares",
        "net_income_company", "net_income_reviewer", "target_year",
        "company_name", "per_multiples", "output",
        "net_income_2029", "net_income_2030", "partial_exit_ratio", "discount_rate",
        "total_shares_before_safe", "safe_amount", "safe_valuation_cap",
        "call_option_price_multiplier"
    }

    # 파라미터를 CLI 인자로 변환 (화이트리스트 검증)
    cmd = [sys.executable, str(script_path)]

    for key, value in parameters.items():
        # 허용되지 않은 파라미터 키 거부
        if key not in allowed_params:
            logger.warning(f"Rejected unknown parameter: {key}")
            continue

        # 파라미터 키 검증 (알파벳, 숫자, 언더스코어만 허용)
        if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', key):
            logger.warning(f"Invalid parameter key format: {key}")
            continue

        # 값 sanitize
        if key == "company_name":
            value = _sanitize_filename(str(value))
        elif key == "output":
            value = _sanitize_filename(str(value))
            if not value.endswith('.xlsx'):
                value += '.xlsx'

        cmd.append(f"--{key}")
        cmd.append(str(value))

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            check=True
        )

        # 출력 파일 경로 추출 (stdout에서)
        output_file = None
        for line in result.stdout.split('\n'):
            if '생성 완료' in line or 'xlsx' in line:
                # 파일 경로 추출
                parts = line.split(':')
                if len(parts) > 1:
                    output_file = parts[-1].strip()

        return {
            "success": True,
            "projection_type": projection_type,
            "output_file": output_file,
            "message": result.stdout
        }

    except subprocess.CalledProcessError as e:
        return {
            "success": False,
            "error": e.stderr
        }


def execute_analyze_and_generate_projection(
    excel_path: str,
    target_year: int,
    per_multiples: List[float],
    company_name: str = None,
    output_filename: str = None
) -> Dict[str, Any]:
    """엑셀 파일 분석 후 즉시 Exit 프로젝션 생성"""

    # 입력 검증: 파일 경로
    is_valid, error = _validate_file_path(excel_path, allowed_extensions=['.xlsx', '.xls'])
    if not is_valid:
        return {"success": False, "error": error}

    # 입력 검증: target_year
    is_valid, year_val, error = _validate_numeric_param(target_year, "target_year", min_val=2020, max_val=2050)
    if not is_valid:
        return {"success": False, "error": error}
    target_year = int(year_val)

    # 입력 검증: per_multiples
    if not per_multiples or not isinstance(per_multiples, list):
        return {"success": False, "error": "PER 멀티플 리스트가 필요합니다"}
    validated_multiples = []
    for m in per_multiples:
        is_valid, val, error = _validate_numeric_param(m, "PER multiple", min_val=0.1, max_val=1000)
        if not is_valid:
            return {"success": False, "error": error}
        validated_multiples.append(val)
    per_multiples = validated_multiples

    # 1단계: 엑셀 파일 분석
    analysis = execute_analyze_excel(excel_path)

    if not analysis["success"]:
        return analysis

    data = analysis
    income_statement = data.get("income_statement", {})
    net_income_data = income_statement.get("net_income", {})
    investment_terms = data.get("investment_terms", {})
    cap_table = data.get("cap_table", {})

    # 2단계: 필수 데이터 검증
    if target_year not in net_income_data:
        return {
            "success": False,
            "error": f"{target_year}년 순이익 데이터를 찾을 수 없습니다. 사용 가능한 연도: {list(net_income_data.keys())}"
        }

    net_income = net_income_data[target_year]
    investment_amount = investment_terms.get("investment_amount")
    price_per_share = investment_terms.get("price_per_share")
    shares = investment_terms.get("shares")
    total_shares = cap_table.get("total_shares")

    if not all([investment_amount, price_per_share, shares, total_shares]):
        return {
            "success": False,
            "error": "필수 투자 정보가 부족합니다",
            "found_data": {
                "investment_amount": investment_amount,
                "price_per_share": price_per_share,
                "shares": shares,
                "total_shares": total_shares
            }
        }

    # 3단계: Exit 프로젝션 요약 계산 (UI/리포트용)
    from datetime import datetime

    investment_year = datetime.now().year
    holding_period_years = target_year - investment_year

    projection_summary: List[Dict[str, Any]] = []
    for per in per_multiples:
        enterprise_value = float(net_income) * float(per)
        proceeds = None
        multiple = None
        irr_pct = None

        try:
            per_share_value = enterprise_value / float(total_shares)
            proceeds = per_share_value * float(shares)
            multiple = proceeds / float(investment_amount)
            if holding_period_years > 0 and multiple > 0:
                irr_pct = (multiple ** (1 / holding_period_years) - 1) * 100
        except Exception:
            pass

        projection_summary.append({
            "PER": float(per),
            "IRR": irr_pct,
            "Multiple": multiple
        })

    # 4단계: Exit 프로젝션 생성 (사용자별 temp 디렉토리에 저장)

    # excel_path에서 user_id 추출 (temp/<user_id>/파일명 형식)
    excel_path_obj = Path(excel_path)
    user_id = "cli_user"  # 기본값
    try:
        # temp/<user_id>/파일명 형식인지 확인
        if "temp" in excel_path_obj.parts:
            temp_idx = excel_path_obj.parts.index("temp")
            if len(excel_path_obj.parts) > temp_idx + 1:
                user_id = excel_path_obj.parts[temp_idx + 1]
    except (ValueError, IndexError):
        pass

    # 출력 파일명 생성
    if not output_filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"exit_projection_{timestamp}.xlsx"
    else:
        # 출력 파일명 sanitize
        output_filename = _sanitize_filename(output_filename)
        if not output_filename.endswith('.xlsx'):
            output_filename += '.xlsx'

    # temp/<user_id>/ 디렉토리에 저장
    output_dir = PROJECT_ROOT / "temp" / user_id
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / output_filename

    if not company_name:
        company_name = _sanitize_filename(Path(excel_path).stem)
    else:
        # 회사명 sanitize
        company_name = _sanitize_filename(company_name)

    # generate_exit_projection.py 스크립트 호출
    script_path = PROJECT_ROOT / "scripts" / "generate_exit_projection.py"

    cmd = [
        sys.executable, str(script_path),
        "--investment_amount", str(int(investment_amount)),
        "--price_per_share", str(int(price_per_share)),
        "--shares", str(int(shares)),
        "--total_shares", str(int(total_shares)),
        "--net_income_company", str(int(net_income)),
        "--net_income_reviewer", str(int(net_income)),  # 같은 값 사용
        "--target_year", str(target_year),
        "--company_name", company_name,
        "--per_multiples", ",".join(map(lambda x: str(int(x) if x == int(x) else x), per_multiples)),
        "--output", str(output_path)
    ]

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            check=True,
            cwd=str(PROJECT_ROOT)
        )

        return {
            "success": True,
            "output_file": str(output_path),
            "assumptions": {
                "company_name": company_name,
                "target_year": target_year,
                "investment_year": investment_year,
                "holding_period_years": holding_period_years,
                "investment_amount": investment_amount,
                "shares": shares,
                "total_shares": total_shares,
                "net_income": net_income,
                "per_multiples": per_multiples
            },
            "projection_summary": projection_summary,
            "analysis_data": {
                "target_year": target_year,
                "net_income": net_income,
                "investment_amount": investment_amount,
                "per_multiples": per_multiples,
                "company_name": company_name
            },
            "message": f"Exit 프로젝션 생성 완료: {output_path.name}"
        }

    except subprocess.CalledProcessError as e:
        return {
            "success": False,
            "error": f"Exit 프로젝션 생성 실패: {e.stderr}"
        }


# ========================================
# 기업현황 진단시트 도구 실행 함수
# ========================================

DIAG_SHEET_INFO = "(기업용) 1. 기업정보"
DIAG_SHEET_CHECKLIST = "(기업용) 2. 체크리스트"
DIAG_SHEET_KPI = "(기업용) 3. KPI기대사항"
DIAG_SHEET_EXIT_CHECKLIST = "(기업&컨설턴트용) EXIT 체크리스트"
DIAG_SHEET_REPORT = "(컨설턴트용) 분석보고서"
DIAG_SHEET_EXPORT_SUMMARY = "내보내기 요약"


def _load_company_diagnosis_template(template_version: str = "2025") -> Dict[str, Any]:
    """로컬 리소스에서 기업현황 진단시트 템플릿 정의를 로드"""
    version = (template_version or "2025").strip()
    if version != "2025":
        raise ValueError(f"지원하지 않는 template_version: {version}")

    template_path = PROJECT_ROOT / "shared" / "resources" / "company_diagnosis_template_2025.json"
    with open(template_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _sanitize_user_id(user_id: str, max_length: int = 80) -> str:
    """temp/<user_id>/ 경로용 user_id 정화"""
    if not user_id:
        return "anonymous"
    s = str(user_id).strip()
    s = s.replace("/", "_").replace("\\", "_").replace("..", "_")
    s = re.sub(r"[^\w\-]", "_", s, flags=re.UNICODE)
    s = re.sub(r"_+", "_", s).strip("_")
    if len(s) > max_length:
        s = s[:max_length]
    return s or "anonymous"


def _extract_user_id_from_temp_path(path_str: str, default: str = "cli_user") -> str:
    try:
        p = Path(path_str)
        if "temp" in p.parts:
            temp_idx = p.parts.index("temp")
            if len(p.parts) > temp_idx + 1:
                return p.parts[temp_idx + 1]
    except Exception:
        pass
    return default


def _normalize_optional_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return str(value).strip() or None


def _diagnosis_draft_progress(draft: Dict[str, Any], template: Dict[str, Any]) -> Dict[str, Any]:
    """드래프트 진행률/다음 질문 계산 (PII를 반환하지 않음)"""
    company_fields = template.get("company_info_fields") or []
    employees_fields = template.get("employees_financial_fields") or []
    investment_fields = template.get("investment_fields") or []
    kpi_fields = (template.get("kpi_fields") or {}).get("business") or []
    milestone_fields = (template.get("kpi_fields") or {}).get("milestone") or []
    checklist_items = template.get("checklist_items") or []
    weights = template.get("weights") or {}

    company_info = draft.get("company_info") or {}
    employees_financials = draft.get("employees_financials") or {}
    investment = draft.get("investment") or {}
    kpi = draft.get("kpi") or {}
    milestone = (kpi.get("milestone") or {}) if isinstance(kpi, dict) else {}
    kpis = (kpi.get("kpis") or []) if isinstance(kpi, dict) else []
    checklist_answers = draft.get("checklist_answers") or {}

    missing: List[Dict[str, Any]] = []

    def add_missing(section: str, field: Dict[str, Any]):
        missing.append(
            {
                "type": "field",
                "section": section,
                "key": field.get("key"),
                "label": field.get("label"),
            }
        )

    for f in company_fields:
        key = f.get("key")
        if f.get("optional") is True:
            continue
        if key and not _normalize_optional_text(company_info.get(key)):
            add_missing("company_info", f)

    for f in employees_fields:
        key = f.get("key")
        if f.get("optional") is True:
            continue
        if key and not _normalize_optional_text(employees_financials.get(key)):
            add_missing("employees_financials", f)

    for f in investment_fields:
        key = f.get("key")
        if f.get("optional") is True:
            continue
        if key and not _normalize_optional_text(investment.get(key)):
            add_missing("investment", f)

    for f in kpi_fields:
        key = f.get("key")
        if f.get("optional") is True:
            continue
        if key and not _normalize_optional_text(kpi.get(key)):
            add_missing("kpi", f)

    # KPI 리스트 최소 1개 권장 (없으면 요청)
    if not isinstance(kpis, list) or len([x for x in kpis if (x or {}).get("name")]) == 0:
        missing.append({"type": "kpi_items", "section": "kpi", "key": "kpis", "label": "정량 KPI"})

    for f in milestone_fields:
        key = f.get("key")
        if f.get("optional") is True:
            continue
        if key and not _normalize_optional_text(milestone.get(key)):
            add_missing("kpi.milestone", f)

    # 체크리스트 (배치로)
    answered_ids = set(checklist_answers.keys()) if isinstance(checklist_answers, dict) else set()
    remaining_items = [it for it in checklist_items if it.get("id") and it.get("id") not in answered_ids]

    total_fields = (
        len(company_fields)
        + len(employees_fields)
        + len(investment_fields)
        + len(kpi_fields)
        + 1  # kpi list
        + len(milestone_fields)
        + len(checklist_items)
    )
    # answered counts
    answered_fields = 0
    for f in company_fields:
        k = f.get("key")
        if k and _normalize_optional_text(company_info.get(k)):
            answered_fields += 1
    for f in employees_fields:
        k = f.get("key")
        if k and _normalize_optional_text(employees_financials.get(k)):
            answered_fields += 1
    for f in investment_fields:
        k = f.get("key")
        if k and _normalize_optional_text(investment.get(k)):
            answered_fields += 1
    for f in kpi_fields:
        k = f.get("key")
        if k and _normalize_optional_text(kpi.get(k)):
            answered_fields += 1
    if isinstance(kpis, list) and len([x for x in kpis if (x or {}).get("name")]) > 0:
        answered_fields += 1
    for f in milestone_fields:
        k = f.get("key")
        if k and _normalize_optional_text(milestone.get(k)):
            answered_fields += 1

    answered_checklist = len(checklist_items) - len(remaining_items)
    answered_total = answered_fields + answered_checklist
    completion_pct = round((answered_total / total_fields * 100), 1) if total_fields else 0.0

    # 점수 계산 (현재까지 응답된 체크리스트 기반)
    checklist_by_module: Dict[str, Dict[str, int]] = {}
    for it in checklist_items:
        it_id = it.get("id")
        module = _normalize_diagnosis_category(it.get("module"))
        if not it_id or not module:
            continue
        checklist_by_module.setdefault(module, {"total": 0, "yes": 0, "no": 0})
        checklist_by_module[module]["total"] += 1
        ans = (checklist_answers or {}).get(it_id) if isinstance(checklist_answers, dict) else None
        status = (ans or {}).get("status") if isinstance(ans, dict) else None
        if status == "예":
            checklist_by_module[module]["yes"] += 1
        elif status == "아니오":
            checklist_by_module[module]["no"] += 1

    scores: Dict[str, Any] = {}
    for category, weight in (weights or {}).items():
        stats = checklist_by_module.get(category, {"total": 0, "yes": 0, "no": 0})
        total = stats.get("total", 0) or 0
        yes = stats.get("yes", 0) or 0
        yes_rate = round((yes / total * 100), 1) if total else None
        score = round((yes / total * float(weight)), 1) if total else None
        scores[category] = {
            "weight": float(weight),
            "score": score,
            "yes": yes,
            "no": stats.get("no", 0) or 0,
            "total": total,
            "yes_rate_pct": yes_rate,
        }

    next_step: Dict[str, Any]
    if missing:
        m0 = missing[0]
        if m0["type"] == "field":
            next_step = {
                "type": "field",
                "section": m0.get("section"),
                "key": m0.get("key"),
                "label": m0.get("label"),
                "prompt": f"{m0.get('label')}을(를) 알려주세요.",
            }
        elif m0["type"] == "kpi_items":
            next_step = {
                "type": "kpi_items",
                "section": "kpi",
                "key": "kpis",
                "label": "정량 KPI",
                "prompt": "정량 KPI를 1~5개까지 알려주세요. 예: '매출: 현재 월 6,700만원 → 목표 월 1억; 고용: 현재 7명 → 목표 9명'",
            }
        else:
            next_step = {"type": "field", "prompt": "다음 정보를 알려주세요."}
    elif remaining_items:
        first = remaining_items[0]
        module = _normalize_diagnosis_category(first.get("module"))
        batch_size = 6
        batch: List[Dict[str, Any]] = []
        for it in remaining_items:
            if _normalize_diagnosis_category(it.get("module")) != module:
                break
            batch.append({"id": it.get("id"), "question": it.get("question")})
            if len(batch) >= batch_size:
                break
        next_step = {
            "type": "checklist_batch",
            "module": module,
            "items": batch,
            "prompt": "다음 체크리스트에 대해 각 항목을 '예/아니오'로 답해주세요. 형식 예: '문제_01 예, 문제_02 아니오(사유...)'",
        }
    else:
        next_step = {"type": "complete", "prompt": "모든 항목이 채워졌습니다. 엑셀로 저장할까요?"}

    return {
        "template_version": template.get("version"),
        "completion_pct": completion_pct,
        "answered": {
            "fields": answered_fields,
            "checklist": answered_checklist,
            "total": answered_total,
        },
        "total": {
            "fields": len(company_fields) + len(employees_fields) + len(investment_fields) + len(kpi_fields) + 1 + len(milestone_fields),
            "checklist": len(checklist_items),
            "overall": total_fields,
        },
        "scores": scores,
        "next": next_step,
    }


def _normalize_diagnosis_category(value: Any) -> str:
    """체크리스트 모듈/분석보고서 헤더를 표준 카테고리로 정규화"""
    if value is None:
        return ""
    s = str(value).strip().replace("\n", " ")
    # "(비즈니스)" 같은 보조 설명 제거
    s = s.split("(")[0].strip()
    # 복수 공백 정리
    s = re.sub(r"\s+", " ", s)
    return s


def _extract_diagnosis_company_info(wb) -> Dict[str, Any]:
    """기업정보 시트에서 기본 정보 추출 (가능한 경우)"""
    info: Dict[str, Any] = {}

    # 1) 기업정보 시트 우선
    if DIAG_SHEET_INFO in wb.sheetnames:
        ws = wb[DIAG_SHEET_INFO]
        info["company_name"] = ws["B6"].value
        info["representative_name"] = ws["C6"].value
        info["email"] = ws["D6"].value
        info["phone"] = ws["E6"].value
        info["incorporation_date"] = ws["F6"].value
        info["business_registration_number"] = ws["G6"].value
        info["business_type"] = ws["H6"].value
        info["hq_address"] = ws["B9"].value
        info["branch_address"] = ws["G9"].value

    # 2) EXIT 체크리스트 보조
    if not info.get("company_name") and DIAG_SHEET_EXIT_CHECKLIST in wb.sheetnames:
        ws = wb[DIAG_SHEET_EXIT_CHECKLIST]
        # B2: 기업명, C2: 값 형태가 일반적
        maybe = ws["C2"].value
        if maybe:
            info["company_name"] = maybe

    # 3) 분석보고서 보조
    if not info.get("company_name") and DIAG_SHEET_REPORT in wb.sheetnames:
        ws = wb[DIAG_SHEET_REPORT]
        maybe = ws["D6"].value
        if maybe:
            info["company_name"] = maybe

    # 정리 (문자열 trim)
    for k, v in list(info.items()):
        if isinstance(v, str):
            info[k] = v.strip()

    return info


def execute_create_company_diagnosis_draft(user_id: str, template_version: str = "2025") -> Dict[str, Any]:
    """템플릿 없이 대화로 작성할 수 있는 진단시트 드래프트(JSON) 생성"""
    try:
        template = _load_company_diagnosis_template(template_version)
    except Exception as e:
        return {"success": False, "error": f"템플릿 로드 실패: {str(e)}"}

    safe_user_id = _sanitize_user_id(user_id)
    output_dir = PROJECT_ROOT / "temp" / safe_user_id
    output_dir.mkdir(parents=True, exist_ok=True)

    from datetime import datetime

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    draft_path = output_dir / f"diagnosis_draft_{timestamp}.json"

    draft = {
        "template_version": template.get("version", template_version),
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat(),
        "company_info": {},
        "employees_financials": {},
        "investment": {},
        "kpi": {"kpis": [], "milestone": {}},
        "checklist_answers": {},
    }

    try:
        with open(draft_path, "w", encoding="utf-8") as f:
            json.dump(draft, f, ensure_ascii=False, indent=2)
    except Exception as e:
        return {"success": False, "error": f"드래프트 저장 실패: {str(e)}"}

    progress = _diagnosis_draft_progress(draft, template)
    return {
        "success": True,
        "draft_path": str(draft_path),
        "progress": progress,
        "message": "드래프트를 생성했습니다.",
    }


def execute_update_company_diagnosis_draft(
    draft_path: str,
    company_info: Dict[str, Any] = None,
    employees_financials: Dict[str, Any] = None,
    investment: Dict[str, Any] = None,
    kpi: Dict[str, Any] = None,
    checklist_answers: List[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """진단시트 드래프트(JSON)에 사용자 응답을 반영"""
    is_valid, error = _validate_file_path(draft_path, allowed_extensions=[".json"], require_temp_dir=True)
    if not is_valid:
        return {"success": False, "error": error}

    try:
        with open(draft_path, "r", encoding="utf-8") as f:
            draft = json.load(f)
    except Exception as e:
        return {"success": False, "error": f"드래프트 로드 실패: {str(e)}"}

    template_version = draft.get("template_version") or "2025"
    try:
        template = _load_company_diagnosis_template(template_version)
    except Exception as e:
        return {"success": False, "error": f"템플릿 로드 실패: {str(e)}"}

    updated = False

    def merge_section(section_key: str, payload: Dict[str, Any]):
        nonlocal updated
        if not payload:
            return
        section = draft.get(section_key)
        if not isinstance(section, dict):
            section = {}
            draft[section_key] = section
        for k, v in payload.items():
            if v is None:
                continue
            section[k] = v
            updated = True

    merge_section("company_info", company_info or {})
    merge_section("employees_financials", employees_financials or {})
    merge_section("investment", investment or {})

    # KPI 섹션은 nested
    if kpi:
        kpi_section = draft.get("kpi")
        if not isinstance(kpi_section, dict):
            kpi_section = {"kpis": [], "milestone": {}}
            draft["kpi"] = kpi_section

        for k, v in (kpi or {}).items():
            if v is None:
                continue
            if k == "kpis" and isinstance(v, list):
                # 최대 5개만 유지, name 없는 항목 제거
                cleaned = []
                for item in v:
                    if not isinstance(item, dict):
                        continue
                    name = _normalize_optional_text(item.get("name"))
                    if not name:
                        continue
                    cleaned.append(
                        {
                            "name": name,
                            "current": _normalize_optional_text(item.get("current")),
                            "target": _normalize_optional_text(item.get("target")),
                        }
                    )
                    if len(cleaned) >= 5:
                        break
                kpi_section["kpis"] = cleaned
                updated = True
            elif k == "milestone" and isinstance(v, dict):
                milestone_section = kpi_section.get("milestone")
                if not isinstance(milestone_section, dict):
                    milestone_section = {}
                    kpi_section["milestone"] = milestone_section
                for mk, mv in v.items():
                    if mv is None:
                        continue
                    milestone_section[mk] = mv
                    updated = True
            else:
                kpi_section[k] = v
                updated = True

    # 체크리스트 응답
    if checklist_answers:
        if not isinstance(draft.get("checklist_answers"), dict):
            draft["checklist_answers"] = {}
        answers = draft["checklist_answers"]
        template_ids = {it.get("id") for it in (template.get("checklist_items") or []) if it.get("id")}

        for ans in checklist_answers:
            if not isinstance(ans, dict):
                continue
            item_id = _normalize_optional_text(ans.get("id"))
            status = _normalize_optional_text(ans.get("status"))
            if not item_id or item_id not in template_ids:
                continue
            if status not in ["예", "아니오"]:
                continue
            answers[item_id] = {
                "status": status,
                "detail": _normalize_optional_text(ans.get("detail")),
            }
            updated = True

    if not updated:
        progress = _diagnosis_draft_progress(draft, template)
        return {
            "success": True,
            "draft_path": str(Path(draft_path)),
            "progress": progress,
            "message": "변경사항이 없어 드래프트는 유지되었습니다.",
        }

    from datetime import datetime

    draft["updated_at"] = datetime.now().isoformat()
    try:
        with open(draft_path, "w", encoding="utf-8") as f:
            json.dump(draft, f, ensure_ascii=False, indent=2)
    except Exception as e:
        return {"success": False, "error": f"드래프트 저장 실패: {str(e)}"}

    progress = _diagnosis_draft_progress(draft, template)
    return {
        "success": True,
        "draft_path": str(Path(draft_path)),
        "progress": progress,
        "message": "드래프트를 업데이트했습니다.",
    }


def execute_generate_company_diagnosis_sheet_from_draft(draft_path: str, output_filename: str = None) -> Dict[str, Any]:
    """드래프트(JSON) 기반 기업현황 진단시트 엑셀 생성"""
    is_valid, error = _validate_file_path(draft_path, allowed_extensions=[".json"], require_temp_dir=True)
    if not is_valid:
        return {"success": False, "error": error}

    try:
        with open(draft_path, "r", encoding="utf-8") as f:
            draft = json.load(f)
    except Exception as e:
        return {"success": False, "error": f"드래프트 로드 실패: {str(e)}"}

    template_version = draft.get("template_version") or "2025"
    try:
        template = _load_company_diagnosis_template(template_version)
    except Exception as e:
        return {"success": False, "error": f"템플릿 로드 실패: {str(e)}"}

    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    user_id = _extract_user_id_from_temp_path(draft_path)
    output_dir = PROJECT_ROOT / "temp" / user_id
    output_dir.mkdir(parents=True, exist_ok=True)

    if not output_filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"diagnosis_sheet_{timestamp}.xlsx"
    else:
        output_filename = _sanitize_filename(output_filename)
        if not output_filename.lower().endswith(".xlsx"):
            output_filename += ".xlsx"

    output_path = output_dir / output_filename

    wb = Workbook()
    # 기본 시트 제거 후 필요한 시트 생성
    default_ws = wb.active
    wb.remove(default_ws)

    ws_summary = wb.create_sheet(DIAG_SHEET_EXPORT_SUMMARY)
    ws_exit = wb.create_sheet(DIAG_SHEET_EXIT_CHECKLIST)
    ws_info = wb.create_sheet(DIAG_SHEET_INFO)
    ws_check = wb.create_sheet(DIAG_SHEET_CHECKLIST)
    ws_kpi = wb.create_sheet(DIAG_SHEET_KPI)
    ws_report = wb.create_sheet(DIAG_SHEET_REPORT)

    # 공통 스타일
    wrap_top = Alignment(wrap_text=True, vertical="top")
    bold = Font(bold=True)

    # ========================================
    # (기업용) 1. 기업정보
    # ========================================
    ws_info["B4"].value = "기업명"
    ws_info["C4"].value = "대표자"
    ws_info["F4"].value = "법인설립일자"
    ws_info["G4"].value = "사업자등록번호"
    ws_info["H4"].value = "주업태"
    for cell in ["B4", "C4", "F4", "G4", "H4"]:
        ws_info[cell].font = bold

    ws_info["C5"].value = "성명"
    ws_info["D5"].value = "이메일"
    ws_info["E5"].value = "전화번호"
    for cell in ["C5", "D5", "E5"]:
        ws_info[cell].font = bold

    ci = draft.get("company_info") or {}
    ws_info["B6"].value = ci.get("company_name")
    ws_info["C6"].value = ci.get("representative_name")
    ws_info["D6"].value = ci.get("email")
    ws_info["E6"].value = ci.get("phone")
    ws_info["F6"].value = ci.get("incorporation_date")
    ws_info["G6"].value = ci.get("business_registration_number")
    ws_info["H6"].value = ci.get("business_type")

    ws_info["B8"].value = "본점 소재지"
    ws_info["G8"].value = "지점 또는 연구소 소재지"
    ws_info["B8"].font = bold
    ws_info["G8"].font = bold

    ws_info["B9"].value = ci.get("hq_address")
    ws_info["G9"].value = ci.get("branch_address")
    ws_info["B9"].alignment = wrap_top
    ws_info["G9"].alignment = wrap_top

    ws_info["B11"].value = "종업원수(명)"
    ws_info["D11"].value = "매출액(원)"
    ws_info["F11"].value = "자본총계(원)"
    ws_info["G11"].value = "인증/지정여부"
    for cell in ["B11", "D11", "F11", "G11"]:
        ws_info[cell].font = bold

    ws_info["B12"].value = "정규직"
    ws_info["C12"].value = "계약직"
    ws_info["D12"].value = "2024년"
    ws_info["E12"].value = "2023년"
    for cell in ["B12", "C12", "D12", "E12"]:
        ws_info[cell].font = bold

    ef = draft.get("employees_financials") or {}
    ws_info["B13"].value = ef.get("employees_full_time")
    ws_info["C13"].value = ef.get("employees_contract")
    ws_info["D13"].value = ef.get("revenue_2024")
    ws_info["E13"].value = ef.get("revenue_2023")
    ws_info["F13"].value = ef.get("equity_total")
    ws_info["G13"].value = ef.get("certification")

    ws_info["B15"].value = "투자이력"
    ws_info["F15"].value = "2025년 내 희망 투자액"
    ws_info["H15"].value = "투자전 희망 기업가치(Pre-Valuation)"
    for cell in ["B15", "F15", "H15"]:
        ws_info[cell].font = bold

    inv = draft.get("investment") or {}
    ws_info["B16"].value = inv.get("investment_history")
    ws_info["F16"].value = inv.get("desired_investment_amount")
    ws_info["H16"].value = inv.get("pre_money_valuation")
    ws_info["B16"].alignment = wrap_top

    ws_info.freeze_panes = "A6"
    ws_info.column_dimensions["B"].width = 22
    ws_info.column_dimensions["C"].width = 16
    ws_info.column_dimensions["D"].width = 24
    ws_info.column_dimensions["E"].width = 18
    ws_info.column_dimensions["F"].width = 16
    ws_info.column_dimensions["G"].width = 20
    ws_info.column_dimensions["H"].width = 20

    # ========================================
    # (기업용) 2. 체크리스트
    # ========================================
    ws_check["B4"].value = "모듈"
    ws_check["C4"].value = "No"
    ws_check["D4"].value = "항목"
    ws_check["E4"].value = "세부항목"
    ws_check["F4"].value = "질문"
    ws_check["G4"].value = "현황"
    ws_check["H4"].value = "근거/요청"
    for cell in ["B4", "C4", "D4", "E4", "F4", "G4", "H4"]:
        ws_check[cell].font = bold

    answers = draft.get("checklist_answers") or {}
    row = 5
    for it in (template.get("checklist_items") or []):
        module = it.get("module")
        it_id = it.get("id")
        ws_check.cell(row, 2).value = module  # B
        ws_check.cell(row, 3).value = it.get("no")  # C
        ws_check.cell(row, 4).value = it.get("item")  # D
        ws_check.cell(row, 5).value = it.get("sub_item")  # E
        ws_check.cell(row, 6).value = it.get("question")  # F

        ans = (answers or {}).get(it_id) if isinstance(answers, dict) else None
        if isinstance(ans, dict):
            ws_check.cell(row, 7).value = ans.get("status")  # G
            ws_check.cell(row, 8).value = ans.get("detail")  # H

        ws_check.cell(row, 6).alignment = wrap_top
        ws_check.cell(row, 8).alignment = wrap_top
        row += 1

    ws_check.freeze_panes = "A5"
    ws_check.column_dimensions["B"].width = 12
    ws_check.column_dimensions["C"].width = 6
    ws_check.column_dimensions["D"].width = 18
    ws_check.column_dimensions["E"].width = 18
    ws_check.column_dimensions["F"].width = 60
    ws_check.column_dimensions["G"].width = 10
    ws_check.column_dimensions["H"].width = 40

    # ========================================
    # (기업용) 3. KPI기대사항
    # ========================================
    ws_kpi["B2"].value = "3. KPI 및 기대사항"
    ws_kpi["B2"].font = Font(bold=True, size=14)

    ws_kpi["B4"].value = "Business"
    ws_kpi["C4"].value = "서비스/제품 소개"
    ws_kpi["D4"].value = "수익 구조"
    ws_kpi["E4"].value = "핵심 고객"
    ws_kpi["F4"].value = "KPI  (정량적 수치)"
    ws_kpi["G4"].value = "현황  (프로그램 시작 시점)"
    ws_kpi["H4"].value = "목표 수준  (프로그램 종료 시점)"
    for cell in ["B4", "C4", "D4", "E4", "F4", "G4", "H4"]:
        ws_kpi[cell].font = bold

    kpi_section = draft.get("kpi") or {}
    business_row = 5
    ws_kpi["C5"].value = kpi_section.get("service_intro")
    ws_kpi["D5"].value = kpi_section.get("revenue_model")
    ws_kpi["E5"].value = kpi_section.get("core_customer")

    kpi_items = kpi_section.get("kpis") if isinstance(kpi_section, dict) else []
    if not isinstance(kpi_items, list):
        kpi_items = []

    for idx, item in enumerate(kpi_items[:5]):
        r = business_row + idx
        ws_kpi.cell(r, 6).value = (item or {}).get("name")
        ws_kpi.cell(r, 7).value = (item or {}).get("current")
        ws_kpi.cell(r, 8).value = (item or {}).get("target")
        ws_kpi.cell(r, 7).alignment = wrap_top
        ws_kpi.cell(r, 8).alignment = wrap_top

    milestone_header_row = 10
    ws_kpi[f"B{milestone_header_row}"].value = "Milestone"
    ws_kpi[f"C{milestone_header_row}"].value = "국내 사업 계획(2025)"
    ws_kpi[f"D{milestone_header_row}"].value = "글로벌 확장 계획(2025)"
    ws_kpi[f"E{milestone_header_row}"].value = "장기 목표 (3년 내)"
    ws_kpi[f"F{milestone_header_row}"].value = "프로그램 목표/기대사항"
    ws_kpi[f"G{milestone_header_row}"].value = "올해 성장/성과 기대&목표"
    ws_kpi[f"H{milestone_header_row}"].value = "우려/리스크"
    for cell in [
        f"B{milestone_header_row}",
        f"C{milestone_header_row}",
        f"D{milestone_header_row}",
        f"E{milestone_header_row}",
        f"F{milestone_header_row}",
        f"G{milestone_header_row}",
        f"H{milestone_header_row}",
    ]:
        ws_kpi[cell].font = bold

    ms = kpi_section.get("milestone") if isinstance(kpi_section, dict) else {}
    if not isinstance(ms, dict):
        ms = {}

    milestone_row = milestone_header_row + 1
    ws_kpi[f"C{milestone_row}"].value = ms.get("domestic_plan_2025")
    ws_kpi[f"D{milestone_row}"].value = ms.get("global_plan_2025")
    ws_kpi[f"E{milestone_row}"].value = ms.get("long_term_goal_3y")
    ws_kpi[f"F{milestone_row}"].value = ms.get("program_expectation")
    ws_kpi[f"G{milestone_row}"].value = ms.get("growth_goal")
    ws_kpi[f"H{milestone_row}"].value = ms.get("concerns")
    for col in ["C", "D", "E", "F", "G", "H"]:
        ws_kpi[f"{col}{milestone_row}"].alignment = wrap_top

    ws_kpi.freeze_panes = "A5"
    ws_kpi.column_dimensions["B"].width = 12
    ws_kpi.column_dimensions["C"].width = 28
    ws_kpi.column_dimensions["D"].width = 28
    ws_kpi.column_dimensions["E"].width = 22
    ws_kpi.column_dimensions["F"].width = 18
    ws_kpi.column_dimensions["G"].width = 26
    ws_kpi.column_dimensions["H"].width = 26

    # ========================================
    # (컨설턴트용) 분석보고서 (최소 구조)
    # ========================================
    ws_report["C5"].value = "기업명"
    ws_report["D5"].value = "작성일시"
    ws_report["C5"].font = bold
    ws_report["D5"].font = bold

    ws_report["D6"].value = (draft.get("company_info") or {}).get("company_name")
    ws_report["D7"].value = datetime.now().strftime("%Y-%m-%d %H:%M")

    categories = template.get("categories") or ["문제", "솔루션", "사업화", "자금조달", "팀/조직", "임팩트"]
    weights = template.get("weights") or {}
    # 헤더: C~H (row 9)
    for idx, cat in enumerate(categories[:6]):
        col = 3 + idx  # C=3
        ws_report.cell(9, col).value = cat
        ws_report.cell(9, col).font = bold
        ws_report.cell(10, col).value = weights.get(cat)

    ws_report["B15"].value = "기업 상황 요약(기업진단)"
    ws_report["B19"].value = "개선 필요사항"
    ws_report["B15"].font = bold
    ws_report["B19"].font = bold
    ws_report["C16"].alignment = wrap_top
    ws_report["C20"].alignment = wrap_top

    ws_report.column_dimensions["B"].width = 18
    ws_report.column_dimensions["C"].width = 45
    ws_report.column_dimensions["D"].width = 28

    try:
        wb.save(output_path)
    except Exception as e:
        return {"success": False, "error": f"엑셀 저장 실패: {str(e)}"}

    return {
        "success": True,
        "output_file": str(output_path),
        "message": f"기업현황 진단시트 생성 완료: {output_path.name}",
    }


def execute_analyze_company_diagnosis_sheet(excel_path: str) -> Dict[str, Any]:
    """기업현황 진단시트 분석"""
    # 입력 검증: 파일 경로 (temp 디렉토리 내부만 허용)
    is_valid, error = _validate_file_path(excel_path, allowed_extensions=[".xlsx", ".xls"], require_temp_dir=True)
    if not is_valid:
        return {"success": False, "error": error}

    try:
        from openpyxl import load_workbook

        wb = load_workbook(excel_path, data_only=False)

        company_info = _extract_diagnosis_company_info(wb)

        # 체크리스트 파싱
        checklist_items: List[Dict[str, Any]] = []
        checklist_summary: Dict[str, Any] = {}
        gaps: List[Dict[str, Any]] = []

        if DIAG_SHEET_CHECKLIST in wb.sheetnames:
            ws = wb[DIAG_SHEET_CHECKLIST]
            current_module_raw = None

            for row in range(5, ws.max_row + 1):
                module_raw = ws.cell(row, 2).value  # B: 모듈
                q_no = ws.cell(row, 3).value        # C: No
                item = ws.cell(row, 4).value        # D: 항목
                sub_item = ws.cell(row, 5).value    # E: 세부항목
                question = ws.cell(row, 6).value    # F: 질문
                status = ws.cell(row, 7).value      # G: 현황
                detail = ws.cell(row, 8).value      # H: 근거/요청

                if module_raw is not None:
                    current_module_raw = str(module_raw).strip()

                status_str = str(status).strip() if isinstance(status, str) else ""
                if status_str not in ("예", "아니오"):
                    continue

                module = _normalize_diagnosis_category(current_module_raw)
                entry = {
                    "row": row,
                    "module": module,
                    "module_raw": current_module_raw,
                    "no": q_no,
                    "item": item,
                    "sub_item": sub_item,
                    "question": question,
                    "status": status_str,
                    "detail": detail
                }
                checklist_items.append(entry)

                if module not in checklist_summary:
                    checklist_summary[module] = {"total": 0, "yes": 0, "no": 0}

                checklist_summary[module]["total"] += 1
                if status_str == "예":
                    checklist_summary[module]["yes"] += 1
                else:
                    checklist_summary[module]["no"] += 1
                    gaps.append({
                        "module": module,
                        "question": question,
                        "detail": detail,
                        "row": row
                    })

        # KPI 파싱 (요약 형태)
        kpi: Dict[str, Any] = {"business": [], "milestone": []}
        if DIAG_SHEET_KPI in wb.sheetnames:
            ws = wb[DIAG_SHEET_KPI]
            current_section = None
            for row in range(5, min(ws.max_row, 260) + 1):
                section = ws.cell(row, 2).value  # B
                if section is not None:
                    current_section = str(section).strip()

                if current_section == "Business":
                    kpi_row = {
                        "kpi": ws.cell(row, 6).value,      # F
                        "current": ws.cell(row, 7).value,  # G
                        "target": ws.cell(row, 8).value,   # H
                    }
                    # 첫 줄에는 소개/수익구조/고객도 포함
                    if row == 5:
                        kpi_row.update({
                            "service_intro": ws.cell(row, 3).value,
                            "revenue_model": ws.cell(row, 4).value,
                            "core_customer": ws.cell(row, 5).value,
                        })
                    if any(v is not None and str(v).strip() != "" for v in kpi_row.values()):
                        kpi["business"].append(kpi_row)

                elif current_section == "Milestone":
                    m_row = {
                        "domestic_plan": ws.cell(row, 3).value,
                        "global_plan": ws.cell(row, 4).value,
                        "long_term_goal": ws.cell(row, 5).value,
                        "program_expectation": ws.cell(row, 6).value,
                        "growth_goal": ws.cell(row, 7).value,
                        "concerns": ws.cell(row, 8).value,
                    }
                    if any(v is not None and str(v).strip() != "" for v in m_row.values()):
                        kpi["milestone"].append(m_row)

        # 분석보고서 시트 기준으로 가중치 로드 (없으면 기본값)
        default_weights = {
            "문제": 20,
            "솔루션": 20,
            "사업화": 20,
            "자금조달": 15,
            "팀/조직": 20,
            "임팩트": 5
        }

        weights = dict(default_weights)
        if DIAG_SHEET_REPORT in wb.sheetnames:
            ws = wb[DIAG_SHEET_REPORT]
            try:
                header_row = 9
                weight_row = 10
                # C~H
                for col in range(3, 9):
                    header = _normalize_diagnosis_category(ws.cell(header_row, col).value)
                    weight_val = ws.cell(weight_row, col).value
                    if header and isinstance(weight_val, (int, float)):
                        weights[header] = float(weight_val)
            except Exception:
                pass

        # 점수 계산 (자가진단 기반)
        scores: Dict[str, Any] = {}
        for category, weight in weights.items():
            stats = checklist_summary.get(category, {"total": 0, "yes": 0, "no": 0})
            total = stats.get("total", 0) or 0
            yes = stats.get("yes", 0) or 0
            no = stats.get("no", 0) or 0
            yes_rate = round((yes / total * 100), 1) if total else None
            score = round((yes / total * float(weight)), 1) if total else None
            scores[category] = {
                "weight": float(weight),
                "score": score,
                "yes": yes,
                "no": no,
                "total": total,
                "yes_rate_pct": yes_rate
            }

        return {
            "success": True,
            "excel_path": excel_path,
            "sheets": wb.sheetnames,
            "company_info": company_info,
            "checklist": {
                "items": checklist_items,
                "summary": checklist_summary,
                "gaps": gaps
            },
            "kpi": kpi,
            "scores": scores
        }

    except Exception as e:
        logger.error(f"Diagnosis sheet analysis failed: {e}", exc_info=True)
        return {"success": False, "error": f"진단시트 분석 실패: {str(e)}"}


def execute_write_company_diagnosis_report(
    excel_path: str,
    scores: Dict[str, float],
    summary_text: str,
    improvement_text: str,
    company_name: str = None,
    report_datetime: str = None,
    output_filename: str = None
) -> Dict[str, Any]:
    """컨설턴트용 분석보고서 시트에 보고서 반영 후 새 파일 생성"""
    is_valid, error = _validate_file_path(excel_path, allowed_extensions=[".xlsx", ".xls"], require_temp_dir=True)
    if not is_valid:
        return {"success": False, "error": error}

    try:
        from datetime import datetime
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment

        wb = load_workbook(excel_path, data_only=False)
        if DIAG_SHEET_REPORT not in wb.sheetnames:
            return {"success": False, "error": f"'{DIAG_SHEET_REPORT}' 시트를 찾을 수 없습니다"}

        # company_name fallback
        extracted_info = _extract_diagnosis_company_info(wb)
        final_company_name = (company_name or extracted_info.get("company_name") or "").strip()

        ws = wb[DIAG_SHEET_REPORT]

        if final_company_name:
            ws["D6"].value = final_company_name

        if not report_datetime:
            report_datetime = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws["D7"].value = report_datetime

        # 점수 반영: 헤더(9행)를 기준으로 매핑
        header_to_col: Dict[str, int] = {}
        for col in range(3, 9):  # C~H
            header = _normalize_diagnosis_category(ws.cell(9, col).value)
            if header:
                header_to_col[header] = col

        for category, val in (scores or {}).items():
            cat = _normalize_diagnosis_category(category)
            if not cat:
                continue
            col = header_to_col.get(cat)
            if not col:
                continue
            try:
                ws.cell(11, col).value = float(val)
            except (TypeError, ValueError):
                continue

        # 본문 반영
        ws["C16"].value = summary_text
        ws["C20"].value = improvement_text

        wrap = Alignment(wrap_text=True, vertical="top")
        ws["C16"].alignment = wrap
        ws["C20"].alignment = wrap

        # 출력 경로: temp/<user_id>/
        excel_path_obj = Path(excel_path)
        user_id = "cli_user"
        try:
            if "temp" in excel_path_obj.parts:
                temp_idx = excel_path_obj.parts.index("temp")
                if len(excel_path_obj.parts) > temp_idx + 1:
                    user_id = excel_path_obj.parts[temp_idx + 1]
        except (ValueError, IndexError):
            pass

        if not output_filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"diagnosis_report_{timestamp}.xlsx"
        else:
            output_filename = _sanitize_filename(output_filename)
            if not output_filename.endswith(".xlsx"):
                output_filename += ".xlsx"

        output_dir = PROJECT_ROOT / "temp" / user_id
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / output_filename

        wb.save(output_path)

        return {
            "success": True,
            "output_file": str(output_path),
            "message": f"진단 분석보고서 반영 완료: {output_path.name}"
        }

    except Exception as e:
        logger.error(f"Diagnosis report write failed: {e}", exc_info=True)
        return {"success": False, "error": f"진단 분석보고서 반영 실패: {str(e)}"}


# ========================================
# Underwriter opinion 데이터 도구 실행 함수
# ========================================

def execute_search_underwriter_opinion(
    jsonl_path: str = None,
    category: str = None,
    corp_name: str = None,
    query: str = None,
    max_results: int = 5,
    max_chars: int = 800,
    min_section_length: int = 0,
    return_patterns: bool = True,
    output_filename: str = None
) -> Dict[str, Any]:
    """인수인의견 JSONL에서 관련 문장 및 일반화 패턴 추출"""
    resolved_path, resolve_error = _resolve_underwriter_data_path(jsonl_path)
    if resolve_error:
        return {
            "success": False,
            "error": resolve_error,
            "suggested_action": "인수인의견 데이터 생성 스크립트를 실행하거나 UNDERWRITER_DATA_PATH를 설정하세요."
        }
    jsonl_path = resolved_path
    category = category.strip().lower() if isinstance(category, str) else category
    try:
        max_results = int(max_results) if max_results is not None else 5
    except (TypeError, ValueError):
        max_results = 5
    try:
        max_chars = int(max_chars) if max_chars is not None else 800
    except (TypeError, ValueError):
        max_chars = 800
    try:
        min_section_length = int(min_section_length) if min_section_length is not None else 0
    except (TypeError, ValueError):
        min_section_length = 0
    if isinstance(return_patterns, str):
        return_patterns = return_patterns.strip().lower() in ("true", "1", "yes", "y")

    is_valid, error = _validate_file_path(
        jsonl_path,
        allowed_extensions=[".jsonl"],
        require_temp_dir=True
    )
    if not is_valid:
        return {
            "success": False,
            "error": error,
            "suggested_action": "인수인의견 데이터 파일을 temp/ 하위로 이동해 주세요."
        }

    if not os.path.exists(jsonl_path):
        return {
            "success": False,
            "error": f"파일을 찾을 수 없습니다: {jsonl_path}",
            "suggested_action": "인수인의견 데이터 파일 경로를 확인하세요."
        }

    try:
        entries = _parse_underwriter_jsonl(jsonl_path)
    except Exception as e:
        logger.error(f"Underwriter JSONL parse failed: {e}", exc_info=True)
        return {"success": False, "error": "JSONL 파싱 실패"}

    results = []
    query_lower = query.lower() if query else None

    for entry in entries:
        section_text = entry.get("section_text") or ""
        section_title = entry.get("section_title") or ""
        combined_text = f"{section_title}\n{section_text}"

        if corp_name:
            if corp_name not in entry.get("corp_name", ""):
                continue

        if min_section_length and entry.get("section_length", 0) < min_section_length:
            continue

        matched_keywords = []
        if category:
            ok, matched_keywords = _match_underwriter_category(combined_text, category)
            if not ok:
                continue

        if query_lower:
            if query_lower not in combined_text.lower():
                continue
            matched_keywords.append(query_lower)

        if category == "market_size":
            snippet = _extract_market_size_sentences(section_text)
        else:
            snippet = ""

        if snippet and max_chars:
            snippet = snippet[:max_chars].strip()

        if not snippet:
            terms = matched_keywords or []
            snippet = _extract_snippet(section_text, terms, max_chars=max_chars)

        score = 0.0
        quality_score = entry.get("quality_score") or 0
        section_length = entry.get("section_length") or 0
        score += float(quality_score)
        score += float(section_length) * 0.0001
        score += len(matched_keywords) * 0.3

        results.append({
            "corp_name": entry.get("corp_name"),
            "rcept_no": entry.get("rcept_no"),
            "report_nm": entry.get("report_nm"),
            "rcept_dt": entry.get("rcept_dt"),
            "section_title": section_title,
            "section_length": section_length,
            "quality_score": quality_score,
            "matched_keywords": matched_keywords,
            "snippet": snippet,
            "score": round(score, 4)
        })

    results.sort(key=lambda x: x.get("score", 0), reverse=True)
    results = results[:max_results] if max_results else results

    patterns = []
    if return_patterns:
        seen = set()
        for item in results:
            template = _generalize_underwriter_text(item.get("snippet"), item.get("corp_name"))
            if template and template not in seen:
                patterns.append(template)
                seen.add(template)

    response = {
        "success": True,
        "source_path": jsonl_path,
        "filters": {
            "category": category,
            "corp_name": corp_name,
            "query": query,
            "max_results": max_results,
            "max_chars": max_chars,
            "min_section_length": min_section_length
        },
        "results": results,
        "patterns": patterns
    }

    if output_filename:
        sanitized_name = _sanitize_filename(output_filename)
        if not sanitized_name.endswith(".json"):
            sanitized_name += ".json"
        output_dir = PROJECT_ROOT / "temp"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / sanitized_name
        try:
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(response, f, ensure_ascii=False, indent=2)
            response["output_file"] = str(output_path)
        except OSError as e:
            logger.error(f"Failed to write underwriter output: {e}", exc_info=True)
            response["output_file_error"] = str(e)

    return response


def execute_search_underwriter_opinion_similar(
    query: str,
    jsonl_path: str = None,
    category: str = None,
    corp_name: str = None,
    top_k: int = 5,
    min_score: float = 0.05,
    max_chars: int = 800,
    min_section_length: int = 0,
    max_text_chars: int = 2000,
    return_patterns: bool = True,
    output_filename: str = None
) -> Dict[str, Any]:
    """인수인의견 JSONL에서 유사도 기반 문장 검색"""
    if not query or not str(query).strip():
        return {"success": False, "error": "query가 비어 있습니다"}

    resolved_path, resolve_error = _resolve_underwriter_data_path(jsonl_path)
    if resolve_error:
        return {
            "success": False,
            "error": resolve_error,
            "suggested_action": "인수인의견 데이터 생성 스크립트를 실행하거나 UNDERWRITER_DATA_PATH를 설정하세요."
        }
    jsonl_path = resolved_path
    category = category.strip().lower() if isinstance(category, str) else category

    try:
        top_k = int(top_k) if top_k is not None else 5
    except (TypeError, ValueError):
        top_k = 5
    try:
        max_chars = int(max_chars) if max_chars is not None else 800
    except (TypeError, ValueError):
        max_chars = 800
    try:
        min_section_length = int(min_section_length) if min_section_length is not None else 0
    except (TypeError, ValueError):
        min_section_length = 0
    try:
        max_text_chars = int(max_text_chars) if max_text_chars is not None else 2000
    except (TypeError, ValueError):
        max_text_chars = 2000
    try:
        min_score = float(min_score) if min_score is not None else 0.05
    except (TypeError, ValueError):
        min_score = 0.05
    if isinstance(return_patterns, str):
        return_patterns = return_patterns.strip().lower() in ("true", "1", "yes", "y")

    is_valid, error = _validate_file_path(
        jsonl_path,
        allowed_extensions=[".jsonl"],
        require_temp_dir=True
    )
    if not is_valid:
        return {
            "success": False,
            "error": error,
            "suggested_action": "인수인의견 데이터 파일을 temp/ 하위로 이동해 주세요."
        }

    if not os.path.exists(jsonl_path):
        return {
            "success": False,
            "error": f"파일을 찾을 수 없습니다: {jsonl_path}",
            "suggested_action": "인수인의견 데이터 파일 경로를 확인하세요."
        }

    index, index_error = _get_underwriter_tfidf_index(
        jsonl_path,
        min_n=3,
        max_n=5,
        max_text_chars=max_text_chars
    )
    if index_error:
        return {"success": False, "error": index_error}

    query_text = _normalize_similarity_text(str(query), max_chars=max_text_chars)
    query_vec, query_norm = _vectorize_query(query_text, index["idf"], min_n=3, max_n=5)

    results = []
    for entry, (vec, norm) in zip(index["entries"], index["vectors"]):
        section_text = entry.get("section_text") or ""
        section_title = entry.get("section_title") or ""
        combined_text = f"{section_title}\n{section_text}"

        if corp_name and corp_name not in entry.get("corp_name", ""):
            continue

        if min_section_length and entry.get("section_length", 0) < min_section_length:
            continue

        if category:
            ok, _ = _match_underwriter_category(combined_text, category)
            if not ok:
                continue

        score = _cosine_similarity(query_vec, query_norm, vec, norm)
        if score < min_score:
            continue

        snippet = _extract_snippet(section_text, [], max_chars=max_chars)
        results.append({
            "corp_name": entry.get("corp_name"),
            "rcept_no": entry.get("rcept_no"),
            "report_nm": entry.get("report_nm"),
            "rcept_dt": entry.get("rcept_dt"),
            "section_title": section_title,
            "section_length": entry.get("section_length"),
            "quality_score": entry.get("quality_score"),
            "snippet": snippet,
            "score": round(score, 6)
        })

    results.sort(key=lambda x: x.get("score", 0), reverse=True)
    if top_k:
        results = results[:top_k]

    patterns = []
    if return_patterns:
        seen = set()
        for item in results:
            template = _generalize_underwriter_text(item.get("snippet"), item.get("corp_name"))
            if template and template not in seen:
                patterns.append(template)
                seen.add(template)

    response = {
        "success": True,
        "source_path": jsonl_path,
        "filters": {
            "query": query,
            "category": category,
            "corp_name": corp_name,
            "top_k": top_k,
            "min_score": min_score,
            "max_chars": max_chars,
            "min_section_length": min_section_length,
            "max_text_chars": max_text_chars
        },
        "results": results,
        "patterns": patterns
    }

    if output_filename:
        sanitized_name = _sanitize_filename(output_filename)
        if not sanitized_name.endswith(".json"):
            sanitized_name += ".json"
        output_dir = PROJECT_ROOT / "temp"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / sanitized_name
        try:
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(response, f, ensure_ascii=False, indent=2)
            response["output_file"] = str(output_path)
        except OSError as e:
            logger.error(f"Failed to write underwriter similarity output: {e}", exc_info=True)
            response["output_file_error"] = str(e)

    return response


# ========================================
# PDF 시장규모 근거 추출 도구 실행 함수
# ========================================

def execute_extract_pdf_market_evidence(
    pdf_path: str,
    max_pages: int = 30,
    max_results: int = 20,
    keywords: List[str] = None
) -> Dict[str, Any]:
    """PDF에서 시장규모 근거 문장을 추출"""
    is_valid, error = _validate_file_path(pdf_path, allowed_extensions=['.pdf'], require_temp_dir=True)
    if not is_valid:
        return {"success": False, "error": error}

    if not os.path.exists(pdf_path):
        return {"success": False, "error": f"파일을 찾을 수 없습니다: {pdf_path}"}

    try:
        max_pages = int(max_pages) if max_pages is not None else 30
    except (TypeError, ValueError):
        max_pages = 30
    try:
        max_results = int(max_results) if max_results is not None else 20
    except (TypeError, ValueError):
        max_results = 20

    base_keywords = [
        "시장", "규모", "전망", "성장", "성장률", "CAGR",
        "TAM", "SAM", "SOM", "수요", "가격", "매출", "톤", "kg", "달러"
    ]
    if keywords:
        base_keywords.extend([str(k) for k in keywords if k])

    import fitz  # PyMuPDF

    doc = None
    try:
        try:
            file_hash = compute_file_hash(Path(pdf_path))
            payload = {
                "version": CACHE_VERSION,
                "file_hash": file_hash,
                "max_pages": max_pages,
                "max_results": max_results,
                "keywords": base_keywords,
                "tool": "extract_pdf_market_evidence",
            }
            cache_key = compute_payload_hash(payload)
            cache_dir = get_cache_dir("market_evidence", "shared")
            cache_path = cache_dir / f"{cache_key}.json"
            cached = load_json(cache_path)
            if cached:
                cached["cache_hit"] = True
                return cached
        except Exception:
            cache_path = None

        doc = fitz.open(pdf_path)
        total_pages = len(doc)
        pages_to_read = min(total_pages, max_pages)

        evidence = []
        seen = set()

        for page_idx in range(pages_to_read):
            page = doc[page_idx]
            text = page.get_text()
            for line in text.splitlines():
                line_text = _normalize_text(line)
                if not line_text:
                    continue

                if not any(keyword in line_text for keyword in base_keywords):
                    continue

                if not re.search(r"\d", line_text):
                    continue

                numbers = _extract_numeric_phrases(line_text)
                if not numbers:
                    continue

                key = (page_idx + 1, line_text)
                if key in seen:
                    continue
                seen.add(key)

                matched = [kw for kw in base_keywords if kw in line_text]
                evidence.append({
                    "page": page_idx + 1,
                    "text": line_text,
                    "numbers": numbers,
                    "matched_keywords": matched[:10],
                    "source": f"PDF p.{page_idx + 1}"
                })

                if len(evidence) >= max_results:
                    break
            if len(evidence) >= max_results:
                break

        warnings = []
        if not evidence:
            warnings.append("시장규모 근거 문장을 찾지 못했습니다. 키워드나 페이지 범위를 확장하세요.")

        result = {
            "success": True,
            "file_path": pdf_path,
            "total_pages": total_pages,
            "pages_read": pages_to_read,
            "evidence": evidence,
            "evidence_count": len(evidence),
            "warnings": warnings,
            "cache_hit": False,
            "cached_at": datetime.utcnow().isoformat()
        }
        try:
            file_hash = compute_file_hash(Path(pdf_path))
            payload = {
                "version": CACHE_VERSION,
                "file_hash": file_hash,
                "max_pages": max_pages,
                "max_results": max_results,
                "keywords": base_keywords,
                "tool": "extract_pdf_market_evidence",
            }
            cache_key = compute_payload_hash(payload)
            cache_dir = get_cache_dir("market_evidence", "shared")
            cache_path = cache_dir / f"{cache_key}.json"
            save_json(cache_path, result)
        except Exception:
            pass

        return result
    except Exception as e:
        logger.error(f"PDF market evidence extraction failed: {e}", exc_info=True)
        return {"success": False, "error": f"시장규모 근거 추출 실패: {str(e)}"}
    finally:
        if doc:
            doc.close()


def execute_fetch_underwriter_opinion_data(
    start_date: str = None,
    end_date: str = None,
    lookback_days: int = 365,
    out_dir: str = None,
    max_items: int = 200,
    last_only: bool = True,
    include_corrections: bool = False,
    min_length: int = 600,
    min_score: float = 3.0,
    api_key: str = None
) -> Dict[str, Any]:
    """인수인의견 원천 데이터 수집 및 JSONL 생성"""
    api_key = api_key or os.getenv("DART_API_KEY")
    if not api_key:
        return {
            "success": False,
            "error": "API 키가 필요합니다. tool 입력 또는 환경변수로 설정하세요."
        }

    today = datetime.now().date()
    if end_date:
        end_str = end_date.strip()
    else:
        end_str = today.strftime("%Y%m%d")

    if start_date:
        start_str = start_date.strip()
    else:
        try:
            lookback_days = int(lookback_days) if lookback_days is not None else 365
        except (TypeError, ValueError):
            lookback_days = 365
        start_str = (today - timedelta(days=lookback_days)).strftime("%Y%m%d")

    if not re.match(r"^\d{8}$", start_str) or not re.match(r"^\d{8}$", end_str):
        return {"success": False, "error": "start_date/end_date 형식은 YYYYMMDD 이어야 합니다."}

    if out_dir:
        out_path = Path(out_dir)
        if not out_path.is_absolute():
            out_path = (PROJECT_ROOT / out_path).resolve()
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = (PROJECT_ROOT / "temp" / f"underwriter_opinion_{timestamp}").resolve()

    # temp 하위로 제한
    temp_root = (PROJECT_ROOT / "temp").resolve()
    try:
        out_path.relative_to(temp_root)
    except ValueError:
        return {"success": False, "error": "출력 디렉토리는 temp/ 하위여야 합니다."}

    out_path.mkdir(parents=True, exist_ok=True)

    try:
        max_items = int(max_items) if max_items is not None else 200
    except (TypeError, ValueError):
        max_items = 200
    try:
        min_length = int(min_length) if min_length is not None else 600
    except (TypeError, ValueError):
        min_length = 600
    try:
        min_score = float(min_score) if min_score is not None else 3.0
    except (TypeError, ValueError):
        min_score = 3.0

    cmd = [
        sys.executable,
        str(PROJECT_ROOT / "scripts" / "dart_extract_underwriter_opinion.py"),
        "--start", start_str,
        "--end", end_str,
        "--out", str(out_path),
        "--api-key", api_key,
        "--max-items", str(max_items),
        "--min-length", str(min_length),
        "--min-score", str(min_score),
    ]

    if last_only:
        cmd.append("--last-only")
    if include_corrections:
        cmd.append("--include-corrections")

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            check=False
        )
    except OSError as e:
        return {"success": False, "error": f"데이터 수집 실행 실패: {str(e)}"}

    if result.returncode != 0:
        err_tail = (result.stderr or "").strip().splitlines()[-5:]
        return {
            "success": False,
            "error": "데이터 수집 실패",
            "details": "\n".join(err_tail)
        }

    output_file = out_path / "underwriter_opinion.jsonl"
    if not output_file.exists():
        return {"success": False, "error": "결과 JSONL 파일을 찾을 수 없습니다."}

    return {
        "success": True,
        "output_dir": str(out_path),
        "output_file": str(output_file),
        "start_date": start_str,
        "end_date": end_str,
        "max_items": max_items
    }


# ========================================
# Peer PER 분석 도구 실행 함수
# ========================================

def execute_read_pdf_as_text(
    pdf_path: str,
    max_pages: int = 30,
    output_mode: str = "structured",
    extract_financial_tables: bool = True
) -> Dict[str, Any]:
    """PDF 파일을 Dolphin AI 모델로 파싱하여 읽기

    Dolphin 사용 불가 시 PyMuPDF로 자동 폴백합니다.
    """
    # 입력 검증: 파일 경로 (temp 디렉토리 내부만 허용)
    is_valid, error = _validate_file_path(pdf_path, allowed_extensions=['.pdf'], require_temp_dir=True)
    if not is_valid:
        return {"success": False, "error": error}

    # 파일 존재 확인
    if not os.path.exists(pdf_path):
        return {"success": False, "error": f"파일을 찾을 수 없습니다: {pdf_path}"}

    # 캐시 확인
    try:
        file_hash = compute_file_hash(Path(pdf_path))
        payload = {
            "version": CACHE_VERSION,
            "file_hash": file_hash,
            "max_pages": max_pages,
            "output_mode": output_mode,
            "extract_financial_tables": extract_financial_tables,
            "tool": "read_pdf_as_text_dolphin",
        }
        cache_key = compute_payload_hash(payload)
        cache_dir = get_cache_dir("dolphin_pdf", "shared")
        cache_path = cache_dir / f"{cache_key}.json"
        cached = load_json(cache_path)
        if cached:
            cached["cache_hit"] = True
            logger.info(f"Cache hit for PDF: {pdf_path}")
            return cached
    except Exception:
        cache_path = None

    # Claude Vision으로 처리 시도
    try:
        from dolphin_service.processor import ClaudeVisionProcessor

        processor = ClaudeVisionProcessor()
        result = processor.process_pdf(
            pdf_path=pdf_path,
            max_pages=max_pages,
            output_mode=output_mode,
        )

        # Claude가 이미 financial_tables를 추출하므로 별도 처리 불필요
        # 캐시 저장
        if cache_path and result.get("success"):
            save_json(cache_path, result)

        logger.info(f"PDF processed with Claude Vision: {pdf_path}")
        return result

    except ImportError as e:
        logger.warning(f"Claude Vision 모듈 로드 실패, PyMuPDF로 폴백: {e}")
        return _execute_read_pdf_as_text_pymupdf(pdf_path, max_pages, cache_path)

    except Exception as e:
        logger.warning(f"Claude Vision 처리 실패, PyMuPDF로 폴백: {e}")
        return _execute_read_pdf_as_text_pymupdf(pdf_path, max_pages, cache_path)


def _execute_read_pdf_as_text_pymupdf(
    pdf_path: str,
    max_pages: int = 30,
    cache_path: Path = None
) -> Dict[str, Any]:
    """PyMuPDF를 사용한 기존 PDF 텍스트 추출 (폴백용)"""
    import fitz  # PyMuPDF

    doc = None
    try:
        doc = fitz.open(pdf_path)
        total_pages = len(doc)
        pages_to_read = min(total_pages, max_pages)

        text_content = []

        for page_num in range(pages_to_read):
            page = doc[page_num]
            text = page.get_text()

            if text.strip():
                text_content.append(f"\n{'='*60}")
                text_content.append(f"페이지 {page_num + 1}")
                text_content.append(f"{'='*60}")
                text_content.append(text)

        full_text = "\n".join(text_content)

        logger.info(f"PDF read with PyMuPDF (fallback): {pdf_path} ({pages_to_read}/{total_pages} pages)")
        result = {
            "success": True,
            "file_path": pdf_path,
            "total_pages": total_pages,
            "pages_read": pages_to_read,
            "content": full_text,
            "char_count": len(full_text),
            "processing_method": "pymupdf_fallback",
            "fallback_used": True,
            "cache_hit": False,
            "cached_at": datetime.utcnow().isoformat()
        }
        if cache_path:
            save_json(cache_path, result)
        return result

    except FileNotFoundError:
        return {"success": False, "error": f"파일을 찾을 수 없습니다: {pdf_path}"}
    except PermissionError:
        return {"success": False, "error": f"파일 접근 권한이 없습니다: {pdf_path}"}
    except Exception as e:
        logger.error(f"Failed to read PDF {pdf_path}: {e}", exc_info=True)
        return {
            "success": False,
            "error": f"PDF 파일 읽기 실패: {str(e)}"
        }
    finally:
        if doc is not None:
            doc.close()


def execute_parse_pdf_dolphin(
    pdf_path: str,
    max_pages: int = 30,
    output_mode: str = "structured",
    extract_financial_tables: bool = True
) -> Dict[str, Any]:
    """Dolphin AI 모델로 PDF 파싱 (전용 도구)"""
    return execute_read_pdf_as_text(
        pdf_path=pdf_path,
        max_pages=max_pages,
        output_mode=output_mode,
        extract_financial_tables=extract_financial_tables
    )


def execute_extract_pdf_tables(
    pdf_path: str,
    max_pages: int = 50
) -> Dict[str, Any]:
    """PDF에서 테이블만 추출"""
    result = execute_read_pdf_as_text(
        pdf_path=pdf_path,
        max_pages=max_pages,
        output_mode="tables_only",
        extract_financial_tables=True
    )

    if not result.get("success"):
        return result

    # 테이블만 추출하여 반환
    return {
        "success": True,
        "file_path": pdf_path,
        "total_pages": result.get("total_pages", 0),
        "financial_tables": result.get("financial_tables", {}),
        "structured_content": result.get("structured_content", {}),
        "processing_method": result.get("processing_method", "unknown"),
        "cache_hit": result.get("cache_hit", False),
    }


# ========================================
# 대화형 투자 분석 도구 실행 함수
# ========================================

def execute_start_analysis_session(
    initial_pdf_path: str = None,
    max_pages: int = 30
) -> Dict[str, Any]:
    """대화형 투자 분석 세션 시작"""
    try:
        from dolphin_service.processor import get_or_create_session

        session = get_or_create_session()

        result = {
            "success": True,
            "session_id": session.session_id,
            "message": "새 분석 세션이 시작되었습니다.",
        }

        # 초기 PDF가 있으면 분석
        if initial_pdf_path:
            status = session.add_pdf(initial_pdf_path, max_pages)
            # 필수/선택 누락 데이터 합치기
            all_missing = status.get("critical_missing", []) + status.get("optional_missing", [])
            result.update({
                "initial_file_analyzed": True,
                "status": status.get("status"),
                "message": status.get("message"),
                "collected_data": status.get("accumulated_data", {}),
                "missing_data": all_missing,
            })

            # 부족한 데이터가 있으면 안내
            if all_missing:
                result["next_steps"] = [
                    f"- {item['name']}: {item['suggestion']}"
                    for item in all_missing
                ]
        else:
            result["next_steps"] = [
                "PDF 파일을 업로드하거나 add_supplementary_data 도구로 데이터를 추가하세요.",
                "예: 재무 데이터, Cap Table, 투자 조건 등"
            ]

        return result

    except Exception as e:
        logger.exception("Analysis session start failed")
        return {
            "success": False,
            "error": f"세션 시작 실패: {str(e)}"
        }


def execute_add_supplementary_data(
    session_id: str,
    pdf_path: str = None,
    text_input: str = None,
    data_type: str = "general",
    max_pages: int = 30
) -> Dict[str, Any]:
    """분석 세션에 추가 데이터 입력"""
    try:
        from dolphin_service.processor import get_or_create_session

        session = get_or_create_session(session_id)

        if session.session_id != session_id:
            return {
                "success": False,
                "error": f"세션 '{session_id}'를 찾을 수 없습니다. start_analysis_session으로 새 세션을 시작하세요."
            }

        if not pdf_path and not text_input:
            return {
                "success": False,
                "error": "pdf_path 또는 text_input 중 하나는 필수입니다."
            }

        result = {
            "success": True,
            "session_id": session_id,
        }

        # PDF 파일 추가
        if pdf_path:
            status = session.add_pdf(pdf_path, max_pages)
            result.update({
                "file_added": pdf_path,
                "status": status.get("status"),
                "message": status.get("message"),
            })

        # 텍스트 입력 추가
        if text_input:
            status = session.add_text_input(text_input, data_type)
            result.update({
                "text_added": True,
                "data_type": data_type,
                "status": status.get("status"),
                "message": status.get("message"),
            })

        # 현재 상태 업데이트
        current_status = session._get_status()
        all_missing = current_status.get("critical_missing", []) + current_status.get("optional_missing", [])
        result.update({
            "collected_data": current_status.get("accumulated_data", {}),
            "missing_data": all_missing,
        })

        if all_missing:
            result["next_steps"] = [
                f"- {item['name']}: {item['suggestion']}"
                for item in all_missing
            ]
        else:
            result["next_steps"] = ["모든 필수 데이터가 수집되었습니다. complete_analysis를 호출하세요."]

        return result

    except Exception as e:
        logger.exception("Supplementary data add failed")
        return {
            "success": False,
            "error": f"데이터 추가 실패: {str(e)}"
        }


def execute_get_analysis_status(session_id: str) -> Dict[str, Any]:
    """분석 세션 상태 확인"""
    try:
        from dolphin_service.processor import get_or_create_session

        session = get_or_create_session(session_id)

        if session.session_id != session_id:
            return {
                "success": False,
                "error": f"세션 '{session_id}'를 찾을 수 없습니다."
            }

        status = session._get_status()

        return {
            "success": True,
            "session_id": session_id,
            "status": status.get("status"),
            "message": status.get("message"),
            "collected_data": status.get("accumulated_data", {}),
            "source_files": session.accumulated_data.get("source_files", []),
            "text_inputs_count": len(session.accumulated_data.get("text_inputs", [])),
            "critical_missing": status.get("critical_missing", []),
            "optional_missing": status.get("optional_missing", []),
        }

    except Exception as e:
        logger.exception("Analysis status check failed")
        return {
            "success": False,
            "error": f"상태 확인 실패: {str(e)}"
        }


def execute_complete_analysis(session_id: str) -> Dict[str, Any]:
    """분석 세션 완료 및 최종 결과 반환"""
    try:
        from dolphin_service.processor import get_or_create_session

        session = get_or_create_session(session_id)

        if session.session_id != session_id:
            return {
                "success": False,
                "error": f"세션 '{session_id}'를 찾을 수 없습니다."
            }

        return session.get_final_analysis()

    except Exception as e:
        logger.exception("Analysis completion failed")
        return {
            "success": False,
            "error": f"분석 완료 실패: {str(e)}"
        }


def _fetch_stock_info(ticker: str) -> dict:
    """yfinance에서 주식 정보 조회 (Rate Limit 대응)"""
    import yfinance as yf
    import random

    # Rate Limit 방지를 위한 딜레이 (5~10초)
    delay = random.uniform(5.0, 10.0)
    logger.info(f"Waiting {delay:.1f}s before fetching {ticker}...")
    time.sleep(delay)

    max_retries = 3
    for attempt in range(max_retries):
        try:
            stock = yf.Ticker(ticker)
            info = stock.info

            # Rate Limit 응답 체크 (빈 dict 또는 에러 메시지)
            if not info or (isinstance(info, dict) and info.get("error")):
                if attempt < max_retries - 1:
                    # 재시도 시 30초 대기
                    retry_delay = 30 + random.uniform(0, 10)
                    logger.warning(f"Rate limit detected for {ticker}, retrying in {retry_delay:.1f}s (attempt {attempt+1}/{max_retries})...")
                    time.sleep(retry_delay)
                    continue
            return info

        except Exception as e:
            if attempt < max_retries - 1:
                # 에러 시 30초 대기
                retry_delay = 30 + random.uniform(0, 10)
                logger.warning(f"Error fetching {ticker}: {e}, retrying in {retry_delay:.1f}s (attempt {attempt+1}/{max_retries})...")
                time.sleep(retry_delay)
            else:
                raise

    return {}


def execute_get_stock_financials(ticker: str) -> Dict[str, Any]:
    """yfinance로 상장 기업 재무 지표 조회"""

    try:
        info = _fetch_stock_info(ticker)

        # 기본 정보가 없으면 에러
        if not info or info.get("regularMarketPrice") is None:
            return {
                "success": False,
                "error": f"티커 '{ticker}'를 찾을 수 없습니다. 티커 형식을 확인하세요. (미국: AAPL, 한국 KOSPI: 005930.KS, KOSDAQ: 035720.KQ)"
            }

        # 재무 지표 추출
        result = {
            "success": True,
            "ticker": ticker,
            "company_name": info.get("longName") or info.get("shortName", "N/A"),
            "sector": info.get("sector", "N/A"),
            "industry": info.get("industry", "N/A"),
            "country": info.get("country", "N/A"),
            "currency": info.get("currency", "USD"),

            # 시가총액
            "market_cap": info.get("marketCap"),
            "market_cap_formatted": _format_large_number(info.get("marketCap")),

            # 밸류에이션 지표
            "trailing_per": info.get("trailingPE"),
            "forward_per": info.get("forwardPE"),
            "psr": info.get("priceToSalesTrailing12Months"),
            "pbr": info.get("priceToBook"),
            "ev_ebitda": info.get("enterpriseToEbitda"),
            "ev_revenue": info.get("enterpriseToRevenue"),

            # 수익성 지표
            "revenue": info.get("totalRevenue"),
            "revenue_formatted": _format_large_number(info.get("totalRevenue")),
            "net_income": info.get("netIncomeToCommon"),
            "net_income_formatted": _format_large_number(info.get("netIncomeToCommon")),
            "operating_margin": info.get("operatingMargins"),
            "profit_margin": info.get("profitMargins"),
            "gross_margin": info.get("grossMargins"),

            # 성장률
            "revenue_growth": info.get("revenueGrowth"),
            "earnings_growth": info.get("earningsGrowth"),

            # 기타
            "current_price": info.get("regularMarketPrice"),
            "52_week_high": info.get("fiftyTwoWeekHigh"),
            "52_week_low": info.get("fiftyTwoWeekLow")
        }

        return result

    except ImportError:
        return {
            "success": False,
            "error": "yfinance가 설치되지 않았습니다. pip install yfinance를 실행하세요."
        }
    except Exception as e:
        return {
            "success": False,
            "error": f"주식 정보 조회 실패: {str(e)}"
        }


def execute_analyze_peer_per(
    tickers: List[str],
    include_forward_per: bool = True
) -> Dict[str, Any]:
    """여러 Peer 기업 PER 일괄 조회 및 비교 분석"""

    try:
        import statistics
        now_ts = time.time()
        try:
            payload = {
                "version": CACHE_VERSION,
                "tickers": tickers,
                "include_forward_per": include_forward_per,
                "tool": "analyze_peer_per",
            }
            cache_key = compute_payload_hash(payload)
            cache_dir = get_cache_dir("peer_per", "shared")
            cache_path = cache_dir / f"{cache_key}.json"
            cached = load_json(cache_path)
            if cached:
                cached_ts = cached.get("cached_at_ts")
                if isinstance(cached_ts, (int, float)) and now_ts - cached_ts < CACHE_TTL_SECONDS:
                    cached["cache_hit"] = True
                    return cached
        except Exception:
            cache_path = None

        peer_data = []
        failed_tickers = []
        total = len(tickers)

        logger.info(f"[Peer PER 분석] 총 {total}개 기업 조회 시작 (예상 소요: {total * 8}~{total * 12}초)")

        for idx, ticker in enumerate(tickers, 1):
            logger.info(f"[{idx}/{total}] {ticker} 조회 중...")

            try:
                # 재시도 지원 헬퍼 함수 사용
                info = _fetch_stock_info(ticker)

                if not info or info.get("regularMarketPrice") is None:
                    logger.warning(f"[{idx}/{total}] {ticker} 조회 실패 - 데이터 없음")
                    failed_tickers.append(ticker)
                    continue

                company_name = info.get("longName") or info.get("shortName", "N/A")
                logger.info(f"[{idx}/{total}] {ticker} 완료 - {company_name}")

                data = {
                    "ticker": ticker,
                    "company_name": company_name,
                    "sector": info.get("sector", "N/A"),
                    "industry": info.get("industry", "N/A"),
                    "market_cap": info.get("marketCap"),
                    "market_cap_formatted": _format_large_number(info.get("marketCap")),
                    "trailing_per": info.get("trailingPE"),
                    "forward_per": info.get("forwardPE") if include_forward_per else None,
                    "revenue": info.get("totalRevenue"),
                    "revenue_formatted": _format_large_number(info.get("totalRevenue")),
                    "operating_margin": info.get("operatingMargins"),
                    "profit_margin": info.get("profitMargins"),
                    "revenue_growth": info.get("revenueGrowth")
                }

                peer_data.append(data)

            except Exception as e:
                logger.warning(f"[{idx}/{total}] {ticker} 조회 실패 - {e}")
                failed_tickers.append(ticker)

        logger.info(f"[Peer PER 분석] 완료 - 성공: {len(peer_data)}개, 실패: {len(failed_tickers)}개")

        if not peer_data:
            return {
                "success": False,
                "error": "유효한 티커가 없습니다.",
                "failed_tickers": failed_tickers
            }

        # 통계 계산
        trailing_pers = [d["trailing_per"] for d in peer_data if d["trailing_per"] is not None]
        forward_pers = [d["forward_per"] for d in peer_data if d.get("forward_per") is not None]
        operating_margins = [d["operating_margin"] for d in peer_data if d["operating_margin"] is not None]

        stats = {}

        if trailing_pers:
            stats["trailing_per"] = {
                "mean": round(statistics.mean(trailing_pers), 2),
                "median": round(statistics.median(trailing_pers), 2),
                "min": round(min(trailing_pers), 2),
                "max": round(max(trailing_pers), 2),
                "count": len(trailing_pers)
            }

        if forward_pers:
            stats["forward_per"] = {
                "mean": round(statistics.mean(forward_pers), 2),
                "median": round(statistics.median(forward_pers), 2),
                "min": round(min(forward_pers), 2),
                "max": round(max(forward_pers), 2),
                "count": len(forward_pers)
            }

        if operating_margins:
            stats["operating_margin"] = {
                "mean": round(statistics.mean(operating_margins) * 100, 2),
                "median": round(statistics.median(operating_margins) * 100, 2),
                "min": round(min(operating_margins) * 100, 2),
                "max": round(max(operating_margins) * 100, 2),
                "count": len(operating_margins)
            }

        warnings = []
        outliers = {}
        trailing_pairs = [(d["ticker"], d["trailing_per"]) for d in peer_data if d["trailing_per"] is not None]
        if len(trailing_pairs) >= 4:
            values = [val for _, val in trailing_pairs]
            q1, q3 = statistics.quantiles(values, n=4)[0], statistics.quantiles(values, n=4)[2]
            iqr = q3 - q1
            lower = q1 - 1.5 * iqr
            upper = q3 + 1.5 * iqr
            outlier_tickers = [ticker for ticker, val in trailing_pairs if val < lower or val > upper]
            if outlier_tickers:
                outliers["trailing_per"] = outlier_tickers
                warnings.append(f"Trailing PER 이상치 후보: {', '.join(outlier_tickers)}")

        missing_per = [d["ticker"] for d in peer_data if d["trailing_per"] is None]
        if missing_per:
            warnings.append(f"Trailing PER 미확인: {', '.join(missing_per)}")

        result = {
            "success": True,
            "peer_count": len(peer_data),
            "peers": peer_data,
            "statistics": stats,
            "failed_tickers": failed_tickers,
            "summary": _generate_per_summary(stats),
            "warnings": warnings,
            "outliers": outliers,
            "cache_hit": False,
            "cached_at_ts": now_ts
        }
        if cache_path:
            save_json(cache_path, result)
        return result

    except ImportError:
        return {
            "success": False,
            "error": "yfinance가 설치되지 않았습니다. pip install yfinance를 실행하세요."
        }
    except Exception as e:
        return {
            "success": False,
            "error": f"Peer PER 분석 실패: {str(e)}"
        }


def execute_query_investment_portfolio(
    query: str = None,
    filters: Dict[str, Any] = None,
    limit: int = None,
    sort_by: str = None,
    sort_order: str = "desc",
) -> Dict[str, Any]:
    """투자기업 CSV 기반 조회"""

    try:
        records = search_portfolio_records(
            query=query,
            filters=filters or {},
            limit=limit,
            sort_by=sort_by,
            sort_order=sort_order or "desc",
        )
        summary = summarize_portfolio_records(records, query=query or "", filters=filters or {})
        return {
            "success": True,
            "summary": summary,
            "records": records,
        }
    except FileNotFoundError as exc:
        return {"success": False, "error": str(exc)}
    except Exception as exc:
        return {"success": False, "error": f"투자기업 조회 실패: {str(exc)}"}


def _format_large_number(value) -> str:
    """큰 숫자를 읽기 쉬운 형식으로 변환"""
    if value is None:
        return "N/A"

    if abs(value) >= 1_000_000_000_000:  # 1조 이상
        return f"{value / 1_000_000_000_000:.2f}조"
    elif abs(value) >= 1_000_000_000:  # 10억 이상
        return f"{value / 1_000_000_000:.2f}B"
    elif abs(value) >= 1_000_000:  # 100만 이상
        return f"{value / 1_000_000:.2f}M"
    else:
        return f"{value:,.0f}"


def _generate_per_summary(stats: Dict[str, Any]) -> str:
    """PER 분석 요약 텍스트 생성"""
    summary_parts = []

    if "trailing_per" in stats:
        tp = stats["trailing_per"]
        summary_parts.append(f"Trailing PER: 평균 {tp['mean']}x, 중간값 {tp['median']}x (범위: {tp['min']}x ~ {tp['max']}x)")

    if "forward_per" in stats:
        fp = stats["forward_per"]
        summary_parts.append(f"Forward PER: 평균 {fp['mean']}x, 중간값 {fp['median']}x (범위: {fp['min']}x ~ {fp['max']}x)")

    if "operating_margin" in stats:
        om = stats["operating_margin"]
        summary_parts.append(f"영업이익률: 평균 {om['mean']}%, 중간값 {om['median']}% (범위: {om['min']}% ~ {om['max']}%)")

    return "\n".join(summary_parts)


# Tool 실행 함수 매핑
TOOL_EXECUTORS = {
    # Exit 프로젝션 도구
    "read_excel_as_text": execute_read_excel_as_text,
    "analyze_excel": execute_analyze_excel,
    "analyze_and_generate_projection": execute_analyze_and_generate_projection,
    # 기업현황 진단시트 도구
    "analyze_company_diagnosis_sheet": execute_analyze_company_diagnosis_sheet,
    "create_company_diagnosis_draft": execute_create_company_diagnosis_draft,
    "update_company_diagnosis_draft": execute_update_company_diagnosis_draft,
    "generate_company_diagnosis_sheet_from_draft": execute_generate_company_diagnosis_sheet_from_draft,
    "write_company_diagnosis_report": execute_write_company_diagnosis_report,
    "calculate_valuation": execute_calculate_valuation,
    "calculate_dilution": execute_calculate_dilution,
    "calculate_irr": execute_calculate_irr,
    "generate_exit_projection": execute_generate_exit_projection,
    # Peer PER 분석 도구
    "search_underwriter_opinion": execute_search_underwriter_opinion,
    "read_pdf_as_text": execute_read_pdf_as_text,
    "get_stock_financials": execute_get_stock_financials,
    "analyze_peer_per": execute_analyze_peer_per,
    "query_investment_portfolio": execute_query_investment_portfolio,
    "search_underwriter_opinion_similar": execute_search_underwriter_opinion_similar,
    "extract_pdf_market_evidence": execute_extract_pdf_market_evidence,
    "fetch_underwriter_opinion_data": execute_fetch_underwriter_opinion_data,
    # Dolphin PDF 파싱 도구
    "parse_pdf_dolphin": execute_parse_pdf_dolphin,
    "extract_pdf_tables": execute_extract_pdf_tables,
    # 대화형 투자 분석 도구
    "start_analysis_session": execute_start_analysis_session,
    "add_supplementary_data": execute_add_supplementary_data,
    "get_analysis_status": execute_get_analysis_status,
    "complete_analysis": execute_complete_analysis,
    # 스타트업 발굴 지원 도구는 파일 끝에서 동적 등록 (함수 정의 후)
}


def execute_tool(tool_name: str, tool_input: Dict[str, Any]) -> Dict[str, Any]:
    """도구 실행 디스패처"""

    executor = TOOL_EXECUTORS.get(tool_name)

    if not executor:
        return {
            "success": False,
            "error": f"Unknown tool: {tool_name}"
        }

    try:
        return executor(**tool_input)
    except Exception as e:
        return {
            "success": False,
            "error": f"Tool execution error: {str(e)}"
        }


# ========================================
# 스타트업 발굴 지원 도구
# ========================================

def execute_analyze_government_policy(
    pdf_paths: List[str] = None,
    text_content: str = None,
    focus_keywords: List[str] = None,
    max_pages_per_pdf: int = 30,
    api_key: str = None
) -> Dict[str, Any]:
    """정부 정책 PDF/텍스트 분석 실행"""
    try:
        from discovery_service import PolicyAnalyzer

        # PDF 경로 검증
        validated_paths = []
        if pdf_paths:
            for pdf_path in pdf_paths:
                is_valid, error = _validate_file_path(pdf_path, require_temp_dir=True)
                if not is_valid:
                    return {"success": False, "error": f"경로 검증 실패: {pdf_path} - {error}"}
                validated_paths.append(pdf_path)

        # PDF도 텍스트도 없으면 오류
        if not validated_paths and not text_content:
            return {"success": False, "error": "PDF 파일 또는 텍스트 콘텐츠가 필요합니다"}

        analyzer = PolicyAnalyzer(api_key=api_key)
        result = analyzer.analyze_content(
            pdf_paths=validated_paths if validated_paths else None,
            text_content=text_content,
            focus_keywords=focus_keywords,
            max_pages=max_pages_per_pdf
        )

        return result

    except ImportError as e:
        return {
            "success": False,
            "error": f"discovery_service 모듈을 찾을 수 없습니다: {str(e)}"
        }
    except Exception as e:
        return {
            "success": False,
            "error": f"정책 분석 실패: {str(e)}"
        }


def execute_search_iris_plus_metrics(
    query: str,
    category: str = None,
    sdg_filter: List[int] = None,
    top_k: int = 10
) -> Dict[str, Any]:
    """IRIS+ 메트릭 검색 실행"""
    try:
        from discovery_service import IRISMapper

        mapper = IRISMapper()
        result = mapper.search_metrics(
            query=query,
            category=category,
            sdg_filter=sdg_filter,
            top_k=top_k
        )

        return result

    except FileNotFoundError as e:
        return {
            "success": False,
            "error": f"IRIS+ 카탈로그를 찾을 수 없습니다: {str(e)}"
        }
    except Exception as e:
        return {
            "success": False,
            "error": f"IRIS+ 검색 실패: {str(e)}"
        }


def execute_map_policy_to_iris(
    policy_themes: List[str],
    target_industries: List[str] = None,
    min_relevance_score: float = 0.3
) -> Dict[str, Any]:
    """정책 테마를 IRIS+ 메트릭에 매핑"""
    try:
        from discovery_service import IRISMapper

        mapper = IRISMapper()
        result = mapper.map_themes_to_iris(
            themes=policy_themes,
            industries=target_industries,
            min_score=min_relevance_score
        )

        return result

    except Exception as e:
        return {
            "success": False,
            "error": f"IRIS+ 매핑 실패: {str(e)}"
        }


def execute_generate_industry_recommendation(
    policy_analysis: Dict[str, Any],
    iris_mapping: Dict[str, Any],
    interest_areas: List[str] = None,
    recommendation_count: int = 5,
    api_key: str = None,
    document_weight: float = None
) -> Dict[str, Any]:
    """유망 산업 추천 생성"""
    try:
        from discovery_service import IndustryRecommender

        recommender = IndustryRecommender(api_key=api_key)
        result = recommender.generate_recommendations(
            policy_analysis=policy_analysis,
            iris_mapping=iris_mapping,
            interest_areas=interest_areas,
            top_k=recommendation_count,
            document_weight=document_weight if document_weight is not None else 0.7
        )

        return result

    except Exception as e:
        try:
            from discovery_service import IndustryRecommender

            recommender = IndustryRecommender(api_key=None)
            fallback = recommender.quick_recommend(
                themes=policy_analysis.get("policy_themes", []),
                industries=policy_analysis.get("target_industries", []),
                interest_areas=interest_areas,
                top_k=recommendation_count,
            )
            def _placeholder_evidence(industry: str) -> List[str]:
                label = industry or "해당 산업"
                return [
                    f"[ASSUMPTION] {label} 산업은 정책 우선순위가 될 가능성이 있음",
                    f"[ASSUMPTION] {label} 수요가 중기적으로 증가할 가능성이 있음",
                ]
            doc_weight_value = float(document_weight if document_weight is not None else 0.7)
            if not interest_areas:
                doc_weight_value = 1.0
            interest_weight_value = (1 - doc_weight_value) if interest_areas else 0.0
            return {
                "success": True,
                "recommendations": [
                    {
                        "rank": item.get("rank"),
                        "industry": item.get("industry"),
                        "total_score": item.get("score", 0),
                        "policy_score": item.get("score", 0),
                        "impact_score": 0.0,
                        "interest_match": False,
                        "rationale": "로컬 폴백 추천",
                        "evidence": _placeholder_evidence(item.get("industry")),
                        "sources": ["미제공(가정)"],
                        "assumptions": ["근거 데이터 제한"],
                        "uncertainties": ["정책 원문 근거 부족"],
                        "evidence_markers": [
                            {
                                "marker": "[ASSUMPTION]",
                                "statement": f"{item.get('industry', '해당 산업')} 수요 확대 가정",
                                "source": "미제공(가정)",
                                "effect_size": "",
                            }
                        ],
                        "iris_codes": [],
                        "sdgs": [],
                        "startup_examples": [],
                        "cautions": ["근거 확보 전에는 투자 결정 보류"],
                    }
                    for item in fallback
                ],
                "emerging_areas": [],
                "caution_areas": [],
                "summary": "API 오류로 로컬 폴백 추천",
                "weighting": {
                    "document_weight": round(doc_weight_value, 2),
                    "interest_weight": round(interest_weight_value, 2),
                    "policy_weight": 0.6,
                    "impact_weight": 0.4,
                },
                "fallback_used": True,
                "error": str(e),
            }
        except Exception:
            pass
        return {
            "success": False,
            "error": f"산업 추천 생성 실패: {str(e)}"
        }


# ========================================
# 스타트업 발굴 도구 동적 등록
# (함수 정의 후에 TOOL_EXECUTORS에 추가)
# ========================================
TOOL_EXECUTORS["analyze_government_policy"] = execute_analyze_government_policy
TOOL_EXECUTORS["search_iris_plus_metrics"] = execute_search_iris_plus_metrics
TOOL_EXECUTORS["map_policy_to_iris"] = execute_map_policy_to_iris
TOOL_EXECUTORS["generate_industry_recommendation"] = execute_generate_industry_recommendation
