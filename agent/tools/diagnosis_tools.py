"""
Company diagnosis sheet tools.

MYSC-style diagnosis sheet lifecycle: create draft, update, generate Excel,
analyze existing sheets, and write consultant reports.
"""

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from ._common import (
    PROJECT_ROOT,
    _normalize_text,
    _sanitize_filename,
    _validate_file_path,
    logger,
)

# ========================================
# Sheet name constants
# ========================================

DIAG_SHEET_INFO = "(기업용) 1. 기업정보"
DIAG_SHEET_CHECKLIST = "(기업용) 2. 체크리스트"
DIAG_SHEET_KPI = "(기업용) 3. KPI기대사항"
DIAG_SHEET_EXIT_CHECKLIST = "(기업&컨설턴트용) EXIT 체크리스트"
DIAG_SHEET_REPORT = "(컨설턴트용) 분석보고서"
DIAG_SHEET_EXPORT_SUMMARY = "내보내기 요약"

# ========================================
# Tool definitions
# ========================================

TOOLS = [
    {
        "name": "analyze_company_diagnosis_sheet",
        "description": "MYSC 기업현황 진단시트(xlsx)를 분석하여 회사 기본정보, 체크리스트 응답(예/아니오), KPI/마일스톤, 항목별 점수(자가진단 기반)를 구조화된 형태로 반환합니다. 컨설턴트용 분석보고서 작성 전에 반드시 이 도구로 진단시트를 파악하세요.",
        "input_schema": {
            "type": "object",
            "properties": {
                "excel_path": {
                    "type": "string",
                    "description": "분석할 기업현황 진단시트 엑셀 파일 경로",
                }
            },
            "required": ["excel_path"],
        },
    },
    {
        "name": "create_company_diagnosis_draft",
        "description": "템플릿 업로드 없이도 대화를 통해 기업현황 진단시트를 작성할 수 있도록, 사용자별 임시 드래프트(JSON)를 생성합니다. 이후 update_company_diagnosis_draft로 내용을 채우고 generate_company_diagnosis_sheet_from_draft로 엑셀을 생성하세요.",
        "input_schema": {
            "type": "object",
            "properties": {
                "user_id": {
                    "type": "string",
                    "description": "사용자 고유 ID (temp/<user_id>/에 드래프트 생성)",
                },
                "template_version": {
                    "type": "string",
                    "description": "템플릿 버전 (기본: 2025)",
                    "default": "2025",
                },
            },
            "required": ["user_id"],
        },
    },
    {
        "name": "update_company_diagnosis_draft",
        "description": "기업현황 진단시트 드래프트(JSON)에 사용자 응답을 반영하고, 다음으로 질문해야 할 항목(필드/체크리스트 배치)과 진행률을 반환합니다. PII를 포함할 수 있으므로 필요한 값만 최소한으로 전달하세요.",
        "input_schema": {
            "type": "object",
            "properties": {
                "draft_path": {
                    "type": "string",
                    "description": "create_company_diagnosis_draft로 생성된 드래프트 경로 (temp 내부 .json)",
                },
                "company_info": {
                    "type": "object",
                    "description": "기업 기본정보 (선택)",
                    "properties": {
                        "company_name": {"type": "string"},
                        "representative_name": {"type": "string"},
                        "email": {"type": "string"},
                        "phone": {"type": "string"},
                        "incorporation_date": {"type": ["string", "number"]},
                        "business_registration_number": {"type": "string"},
                        "business_type": {"type": "string"},
                        "hq_address": {"type": "string"},
                        "branch_address": {"type": "string"},
                    },
                },
                "employees_financials": {
                    "type": "object",
                    "description": "인력/재무 정보 (선택)",
                    "properties": {
                        "employees_full_time": {"type": ["string", "number"]},
                        "employees_contract": {"type": ["string", "number"]},
                        "revenue_2024": {"type": ["string", "number"]},
                        "revenue_2023": {"type": ["string", "number"]},
                        "equity_total": {"type": ["string", "number"]},
                        "certification": {"type": "string"},
                    },
                },
                "investment": {
                    "type": "object",
                    "description": "투자 정보 (선택)",
                    "properties": {
                        "investment_history": {"type": "string"},
                        "desired_investment_amount": {"type": "string"},
                        "pre_money_valuation": {"type": "string"},
                    },
                },
                "kpi": {
                    "type": "object",
                    "description": "KPI/마일스톤 (선택)",
                    "properties": {
                        "service_intro": {"type": "string"},
                        "revenue_model": {"type": "string"},
                        "core_customer": {"type": "string"},
                        "kpis": {
                            "type": "array",
                            "description": "정량 KPI (최대 5개 권장)",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "name": {"type": "string"},
                                    "current": {"type": "string"},
                                    "target": {"type": "string"},
                                },
                                "required": ["name"],
                            },
                        },
                        "milestone": {
                            "type": "object",
                            "properties": {
                                "domestic_plan_2025": {"type": "string"},
                                "global_plan_2025": {"type": "string"},
                                "long_term_goal_3y": {"type": "string"},
                                "program_expectation": {"type": "string"},
                                "growth_goal": {"type": "string"},
                                "concerns": {"type": "string"},
                            },
                        },
                    },
                },
                "checklist_answers": {
                    "type": "array",
                    "description": "체크리스트 응답 (선택) - 예/아니오 + 근거/요청(선택)",
                    "items": {
                        "type": "object",
                        "properties": {
                            "id": {"type": "string"},
                            "status": {"type": "string", "enum": ["예", "아니오"]},
                            "detail": {"type": "string"},
                        },
                        "required": ["id", "status"],
                    },
                },
            },
            "required": ["draft_path"],
        },
    },
    {
        "name": "generate_company_diagnosis_sheet_from_draft",
        "description": "대화로 수집한 드래프트(JSON)를 기반으로 기업현황 진단시트 엑셀(xlsx)을 생성합니다. 생성된 엑셀은 temp 디렉토리에 저장되며, 이후 analyze_company_diagnosis_sheet / write_company_diagnosis_report로 확장할 수 있습니다.",
        "input_schema": {
            "type": "object",
            "properties": {
                "draft_path": {
                    "type": "string",
                    "description": "드래프트 JSON 경로 (temp 내부)",
                },
                "output_filename": {
                    "type": "string",
                    "description": "출력 파일명 (선택, 기본: diagnosis_sheet_YYYYMMDD_HHMMSS.xlsx)",
                },
            },
            "required": ["draft_path"],
        },
    },
    {
        "name": "write_company_diagnosis_report",
        "description": "기업현황 진단시트의 '(컨설턴트용) 분석보고서' 시트에 컨설턴트 요약/개선사항/점수를 반영한 새 엑셀 파일을 생성합니다. 원본 파일은 수정하지 않고 temp 디렉토리에 결과 파일을 저장합니다.",
        "input_schema": {
            "type": "object",
            "properties": {
                "excel_path": {
                    "type": "string",
                    "description": "원본 기업현황 진단시트 엑셀 파일 경로 (temp 내부)",
                },
                "company_name": {
                    "type": "string",
                    "description": "기업명 (선택사항, 비어있으면 원본에서 추출)",
                },
                "report_datetime": {
                    "type": "string",
                    "description": "작성일시 문자열 (선택사항, 비어있으면 현재 시각)",
                },
                "scores": {
                    "type": "object",
                    "description": "항목별 점수 (문제/솔루션/사업화/자금조달/팀/조직/임팩트)",
                    "properties": {
                        "문제": {"type": "number"},
                        "솔루션": {"type": "number"},
                        "사업화": {"type": "number"},
                        "자금조달": {"type": "number"},
                        "팀/조직": {"type": "number"},
                        "임팩트": {"type": "number"},
                    },
                },
                "summary_text": {
                    "type": "string",
                    "description": "기업 상황 요약/기업진단 내용 (분석보고서 본문)",
                },
                "improvement_text": {
                    "type": "string",
                    "description": "개선 필요사항 (분석보고서 본문)",
                },
                "output_filename": {
                    "type": "string",
                    "description": "출력 파일명 (선택사항, 기본값: diagnosis_report_YYYYMMDD_HHMMSS.xlsx)",
                },
            },
            "required": ["excel_path", "scores", "summary_text", "improvement_text"],
        },
    },
]

# ========================================
# Helper functions
# ========================================


def _load_company_diagnosis_template(template_version: str = "2025") -> Dict[str, Any]:
    """로컬 리소스에서 기업현황 진단시트 템플릿 정의를 로드"""
    version = (template_version or "2025").strip()
    if version != "2025":
        raise ValueError(f"지원하지 않는 template_version: {version}")

    template_path = PROJECT_ROOT / "shared" / "resources" / "company_diagnosis_template_2025.json"
    with open(template_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _sanitize_user_id(user_id: str, max_length: int = 80) -> str:
    """temp/<user_id>/ 경로용 user_id 정화"""
    if not user_id:
        return "anonymous"
    s = str(user_id).strip()
    s = s.replace("/", "_").replace("\\", "_").replace("..", "_")
    s = re.sub(r"[^\w\-]", "_", s, flags=re.UNICODE)
    s = re.sub(r"_+", "_", s).strip("_")
    if len(s) > max_length:
        s = s[:max_length]
    return s or "anonymous"


def _extract_user_id_from_temp_path(path_str: str, default: str = "cli_user") -> str:
    try:
        p = Path(path_str)
        if "temp" in p.parts:
            temp_idx = p.parts.index("temp")
            if len(p.parts) > temp_idx + 1:
                return p.parts[temp_idx + 1]
    except Exception:
        pass
    return default


def _normalize_optional_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return str(value).strip() or None


def _normalize_diagnosis_category(value: Any) -> str:
    """체크리스트 모듈/분석보고서 헤더를 표준 카테고리로 정규화"""
    if value is None:
        return ""
    s = str(value).strip().replace("\n", " ")
    s = s.split("(")[0].strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _extract_diagnosis_company_info(wb) -> Dict[str, Any]:
    """기업정보 시트에서 기본 정보 추출 (가능한 경우)"""
    info: Dict[str, Any] = {}

    if DIAG_SHEET_INFO in wb.sheetnames:
        ws = wb[DIAG_SHEET_INFO]
        info["company_name"] = ws["B6"].value
        info["representative_name"] = ws["C6"].value
        info["email"] = ws["D6"].value
        info["phone"] = ws["E6"].value
        info["incorporation_date"] = ws["F6"].value
        info["business_registration_number"] = ws["G6"].value
        info["business_type"] = ws["H6"].value
        info["hq_address"] = ws["B9"].value
        info["branch_address"] = ws["G9"].value

    if not info.get("company_name") and DIAG_SHEET_EXIT_CHECKLIST in wb.sheetnames:
        ws = wb[DIAG_SHEET_EXIT_CHECKLIST]
        maybe = ws["C2"].value
        if maybe:
            info["company_name"] = maybe

    if not info.get("company_name") and DIAG_SHEET_REPORT in wb.sheetnames:
        ws = wb[DIAG_SHEET_REPORT]
        maybe = ws["D6"].value
        if maybe:
            info["company_name"] = maybe

    for k, v in list(info.items()):
        if isinstance(v, str):
            info[k] = v.strip()

    return info


def _diagnosis_draft_progress(draft: Dict[str, Any], template: Dict[str, Any]) -> Dict[str, Any]:
    """드래프트 진행률/다음 질문 계산 (PII를 반환하지 않음)"""
    company_fields = template.get("company_info_fields") or []
    employees_fields = template.get("employees_financial_fields") or []
    investment_fields = template.get("investment_fields") or []
    kpi_fields = (template.get("kpi_fields") or {}).get("business") or []
    milestone_fields = (template.get("kpi_fields") or {}).get("milestone") or []
    checklist_items = template.get("checklist_items") or []
    weights = template.get("weights") or {}

    company_info = draft.get("company_info") or {}
    employees_financials = draft.get("employees_financials") or {}
    investment = draft.get("investment") or {}
    kpi = draft.get("kpi") or {}
    milestone = (kpi.get("milestone") or {}) if isinstance(kpi, dict) else {}
    kpis = (kpi.get("kpis") or []) if isinstance(kpi, dict) else []
    checklist_answers = draft.get("checklist_answers") or {}

    missing: List[Dict[str, Any]] = []

    def add_missing(section: str, field: Dict[str, Any]):
        missing.append(
            {
                "type": "field",
                "section": section,
                "key": field.get("key"),
                "label": field.get("label"),
            }
        )

    for f in company_fields:
        key = f.get("key")
        if f.get("optional") is True:
            continue
        if key and not _normalize_optional_text(company_info.get(key)):
            add_missing("company_info", f)

    for f in employees_fields:
        key = f.get("key")
        if f.get("optional") is True:
            continue
        if key and not _normalize_optional_text(employees_financials.get(key)):
            add_missing("employees_financials", f)

    for f in investment_fields:
        key = f.get("key")
        if f.get("optional") is True:
            continue
        if key and not _normalize_optional_text(investment.get(key)):
            add_missing("investment", f)

    for f in kpi_fields:
        key = f.get("key")
        if f.get("optional") is True:
            continue
        if key and not _normalize_optional_text(kpi.get(key)):
            add_missing("kpi", f)

    if not isinstance(kpis, list) or len([x for x in kpis if (x or {}).get("name")]) == 0:
        missing.append({"type": "kpi_items", "section": "kpi", "key": "kpis", "label": "정량 KPI"})

    for f in milestone_fields:
        key = f.get("key")
        if f.get("optional") is True:
            continue
        if key and not _normalize_optional_text(milestone.get(key)):
            add_missing("kpi.milestone", f)

    answered_ids = set(checklist_answers.keys()) if isinstance(checklist_answers, dict) else set()
    remaining_items = [it for it in checklist_items if it.get("id") and it.get("id") not in answered_ids]

    total_fields = (
        len(company_fields)
        + len(employees_fields)
        + len(investment_fields)
        + len(kpi_fields)
        + 1  # kpi list
        + len(milestone_fields)
        + len(checklist_items)
    )

    answered_fields = 0
    for f in company_fields:
        k = f.get("key")
        if k and _normalize_optional_text(company_info.get(k)):
            answered_fields += 1
    for f in employees_fields:
        k = f.get("key")
        if k and _normalize_optional_text(employees_financials.get(k)):
            answered_fields += 1
    for f in investment_fields:
        k = f.get("key")
        if k and _normalize_optional_text(investment.get(k)):
            answered_fields += 1
    for f in kpi_fields:
        k = f.get("key")
        if k and _normalize_optional_text(kpi.get(k)):
            answered_fields += 1
    if isinstance(kpis, list) and len([x for x in kpis if (x or {}).get("name")]) > 0:
        answered_fields += 1
    for f in milestone_fields:
        k = f.get("key")
        if k and _normalize_optional_text(milestone.get(k)):
            answered_fields += 1

    answered_checklist = len(checklist_items) - len(remaining_items)
    answered_total = answered_fields + answered_checklist
    completion_pct = round((answered_total / total_fields * 100), 1) if total_fields else 0.0

    checklist_by_module: Dict[str, Dict[str, int]] = {}
    for it in checklist_items:
        it_id = it.get("id")
        module = _normalize_diagnosis_category(it.get("module"))
        if not it_id or not module:
            continue
        checklist_by_module.setdefault(module, {"total": 0, "yes": 0, "no": 0})
        checklist_by_module[module]["total"] += 1
        ans = (checklist_answers or {}).get(it_id) if isinstance(checklist_answers, dict) else None
        status = (ans or {}).get("status") if isinstance(ans, dict) else None
        if status == "예":
            checklist_by_module[module]["yes"] += 1
        elif status == "아니오":
            checklist_by_module[module]["no"] += 1

    scores: Dict[str, Any] = {}
    for category, weight in (weights or {}).items():
        stats = checklist_by_module.get(category, {"total": 0, "yes": 0, "no": 0})
        total = stats.get("total", 0) or 0
        yes = stats.get("yes", 0) or 0
        yes_rate = round((yes / total * 100), 1) if total else None
        score = round((yes / total * float(weight)), 1) if total else None
        scores[category] = {
            "weight": float(weight),
            "score": score,
            "yes": yes,
            "no": stats.get("no", 0) or 0,
            "total": total,
            "yes_rate_pct": yes_rate,
        }

    next_step: Dict[str, Any]
    if missing:
        m0 = missing[0]
        if m0["type"] == "field":
            next_step = {
                "type": "field",
                "section": m0.get("section"),
                "key": m0.get("key"),
                "label": m0.get("label"),
                "prompt": f"{m0.get('label')}을(를) 알려주세요.",
            }
        elif m0["type"] == "kpi_items":
            next_step = {
                "type": "kpi_items",
                "section": "kpi",
                "key": "kpis",
                "label": "정량 KPI",
                "prompt": "정량 KPI를 1~5개까지 알려주세요. 예: '매출: 현재 월 6,700만원 → 목표 월 1억; 고용: 현재 7명 → 목표 9명'",
            }
        else:
            next_step = {"type": "field", "prompt": "다음 정보를 알려주세요."}
    elif remaining_items:
        first = remaining_items[0]
        module = _normalize_diagnosis_category(first.get("module"))
        batch_size = 6
        batch: List[Dict[str, Any]] = []
        for it in remaining_items:
            if _normalize_diagnosis_category(it.get("module")) != module:
                break
            batch.append({"id": it.get("id"), "question": it.get("question")})
            if len(batch) >= batch_size:
                break
        next_step = {
            "type": "checklist_batch",
            "module": module,
            "items": batch,
            "prompt": "다음 체크리스트에 대해 각 항목을 '예/아니오'로 답해주세요. 형식 예: '문제_01 예, 문제_02 아니오(사유...)'",
        }
    else:
        next_step = {"type": "complete", "prompt": "모든 항목이 채워졌습니다. 엑셀로 저장할까요?"}

    return {
        "template_version": template.get("version"),
        "completion_pct": completion_pct,
        "answered": {
            "fields": answered_fields,
            "checklist": answered_checklist,
            "total": answered_total,
        },
        "total": {
            "fields": len(company_fields) + len(employees_fields) + len(investment_fields) + len(kpi_fields) + 1 + len(milestone_fields),
            "checklist": len(checklist_items),
            "overall": total_fields,
        },
        "scores": scores,
        "next": next_step,
    }


# ========================================
# Executor functions
# ========================================


def execute_create_company_diagnosis_draft(user_id: str, template_version: str = "2025") -> Dict[str, Any]:
    """템플릿 없이 대화로 작성할 수 있는 진단시트 드래프트(JSON) 생성"""
    try:
        template = _load_company_diagnosis_template(template_version)
    except Exception as e:
        return {"success": False, "error": f"템플릿 로드 실패: {str(e)}"}

    safe_user_id = _sanitize_user_id(user_id)
    output_dir = PROJECT_ROOT / "temp" / safe_user_id
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    draft_path = output_dir / f"diagnosis_draft_{timestamp}.json"

    draft = {
        "template_version": template.get("version", template_version),
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat(),
        "company_info": {},
        "employees_financials": {},
        "investment": {},
        "kpi": {"kpis": [], "milestone": {}},
        "checklist_answers": {},
    }

    try:
        with open(draft_path, "w", encoding="utf-8") as f:
            json.dump(draft, f, ensure_ascii=False, indent=2)
    except Exception as e:
        return {"success": False, "error": f"드래프트 저장 실패: {str(e)}"}

    progress = _diagnosis_draft_progress(draft, template)
    return {
        "success": True,
        "draft_path": str(draft_path),
        "progress": progress,
        "message": "드래프트를 생성했습니다.",
    }


def execute_update_company_diagnosis_draft(
    draft_path: str,
    company_info: Dict[str, Any] = None,
    employees_financials: Dict[str, Any] = None,
    investment: Dict[str, Any] = None,
    kpi: Dict[str, Any] = None,
    checklist_answers: List[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """진단시트 드래프트(JSON)에 사용자 응답을 반영"""
    is_valid, error = _validate_file_path(draft_path, allowed_extensions=[".json"], require_temp_dir=True)
    if not is_valid:
        return {"success": False, "error": error}

    try:
        with open(draft_path, "r", encoding="utf-8") as f:
            draft = json.load(f)
    except Exception as e:
        return {"success": False, "error": f"드래프트 로드 실패: {str(e)}"}

    template_version = draft.get("template_version") or "2025"
    try:
        template = _load_company_diagnosis_template(template_version)
    except Exception as e:
        return {"success": False, "error": f"템플릿 로드 실패: {str(e)}"}

    updated = False

    def merge_section(section_key: str, payload: Dict[str, Any]):
        nonlocal updated
        if not payload:
            return
        section = draft.get(section_key)
        if not isinstance(section, dict):
            section = {}
            draft[section_key] = section
        for k, v in payload.items():
            if v is None:
                continue
            section[k] = v
            updated = True

    merge_section("company_info", company_info or {})
    merge_section("employees_financials", employees_financials or {})
    merge_section("investment", investment or {})

    if kpi:
        kpi_section = draft.get("kpi")
        if not isinstance(kpi_section, dict):
            kpi_section = {"kpis": [], "milestone": {}}
            draft["kpi"] = kpi_section

        for k, v in (kpi or {}).items():
            if v is None:
                continue
            if k == "kpis" and isinstance(v, list):
                cleaned = []
                for item in v:
                    if not isinstance(item, dict):
                        continue
                    name = _normalize_optional_text(item.get("name"))
                    if not name:
                        continue
                    cleaned.append(
                        {
                            "name": name,
                            "current": _normalize_optional_text(item.get("current")),
                            "target": _normalize_optional_text(item.get("target")),
                        }
                    )
                    if len(cleaned) >= 5:
                        break
                kpi_section["kpis"] = cleaned
                updated = True
            elif k == "milestone" and isinstance(v, dict):
                milestone_section = kpi_section.get("milestone")
                if not isinstance(milestone_section, dict):
                    milestone_section = {}
                    kpi_section["milestone"] = milestone_section
                for mk, mv in v.items():
                    if mv is None:
                        continue
                    milestone_section[mk] = mv
                    updated = True
            else:
                kpi_section[k] = v
                updated = True

    if checklist_answers:
        if not isinstance(draft.get("checklist_answers"), dict):
            draft["checklist_answers"] = {}
        answers = draft["checklist_answers"]
        template_ids = {it.get("id") for it in (template.get("checklist_items") or []) if it.get("id")}

        for ans in checklist_answers:
            if not isinstance(ans, dict):
                continue
            item_id = _normalize_optional_text(ans.get("id"))
            status = _normalize_optional_text(ans.get("status"))
            if not item_id or item_id not in template_ids:
                continue
            if status not in ["예", "아니오"]:
                continue
            answers[item_id] = {
                "status": status,
                "detail": _normalize_optional_text(ans.get("detail")),
            }
            updated = True

    if not updated:
        progress = _diagnosis_draft_progress(draft, template)
        return {
            "success": True,
            "draft_path": str(Path(draft_path)),
            "progress": progress,
            "message": "변경사항이 없어 드래프트는 유지되었습니다.",
        }

    draft["updated_at"] = datetime.now().isoformat()
    try:
        with open(draft_path, "w", encoding="utf-8") as f:
            json.dump(draft, f, ensure_ascii=False, indent=2)
    except Exception as e:
        return {"success": False, "error": f"드래프트 저장 실패: {str(e)}"}

    progress = _diagnosis_draft_progress(draft, template)
    return {
        "success": True,
        "draft_path": str(Path(draft_path)),
        "progress": progress,
        "message": "드래프트를 업데이트했습니다.",
    }


def execute_generate_company_diagnosis_sheet_from_draft(draft_path: str, output_filename: str = None) -> Dict[str, Any]:
    """드래프트(JSON) 기반 기업현황 진단시트 엑셀 생성"""
    is_valid, error = _validate_file_path(draft_path, allowed_extensions=[".json"], require_temp_dir=True)
    if not is_valid:
        return {"success": False, "error": error}

    try:
        with open(draft_path, "r", encoding="utf-8") as f:
            draft = json.load(f)
    except Exception as e:
        return {"success": False, "error": f"드래프트 로드 실패: {str(e)}"}

    template_version = draft.get("template_version") or "2025"
    try:
        template = _load_company_diagnosis_template(template_version)
    except Exception as e:
        return {"success": False, "error": f"템플릿 로드 실패: {str(e)}"}

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    user_id = _extract_user_id_from_temp_path(draft_path)
    output_dir = PROJECT_ROOT / "temp" / user_id
    output_dir.mkdir(parents=True, exist_ok=True)

    if not output_filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"diagnosis_sheet_{timestamp}.xlsx"
    else:
        output_filename = _sanitize_filename(output_filename)
        if not output_filename.lower().endswith(".xlsx"):
            output_filename += ".xlsx"

    output_path = output_dir / output_filename

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws_summary = wb.create_sheet(DIAG_SHEET_EXPORT_SUMMARY)
    ws_exit = wb.create_sheet(DIAG_SHEET_EXIT_CHECKLIST)
    ws_info = wb.create_sheet(DIAG_SHEET_INFO)
    ws_check = wb.create_sheet(DIAG_SHEET_CHECKLIST)
    ws_kpi = wb.create_sheet(DIAG_SHEET_KPI)
    ws_report = wb.create_sheet(DIAG_SHEET_REPORT)

    wrap_top = Alignment(wrap_text=True, vertical="top")
    bold = Font(bold=True)

    # (기업용) 1. 기업정보
    ws_info["B4"].value = "기업명"
    ws_info["C4"].value = "대표자"
    ws_info["F4"].value = "법인설립일자"
    ws_info["G4"].value = "사업자등록번호"
    ws_info["H4"].value = "주업태"
    for cell in ["B4", "C4", "F4", "G4", "H4"]:
        ws_info[cell].font = bold

    ws_info["C5"].value = "성명"
    ws_info["D5"].value = "이메일"
    ws_info["E5"].value = "전화번호"
    for cell in ["C5", "D5", "E5"]:
        ws_info[cell].font = bold

    ci = draft.get("company_info") or {}
    ws_info["B6"].value = ci.get("company_name")
    ws_info["C6"].value = ci.get("representative_name")
    ws_info["D6"].value = ci.get("email")
    ws_info["E6"].value = ci.get("phone")
    ws_info["F6"].value = ci.get("incorporation_date")
    ws_info["G6"].value = ci.get("business_registration_number")
    ws_info["H6"].value = ci.get("business_type")

    ws_info["B8"].value = "본점 소재지"
    ws_info["G8"].value = "지점 또는 연구소 소재지"
    ws_info["B8"].font = bold
    ws_info["G8"].font = bold

    ws_info["B9"].value = ci.get("hq_address")
    ws_info["G9"].value = ci.get("branch_address")
    ws_info["B9"].alignment = wrap_top
    ws_info["G9"].alignment = wrap_top

    ws_info["B11"].value = "종업원수(명)"
    ws_info["D11"].value = "매출액(원)"
    ws_info["F11"].value = "자본총계(원)"
    ws_info["G11"].value = "인증/지정여부"
    for cell in ["B11", "D11", "F11", "G11"]:
        ws_info[cell].font = bold

    ws_info["B12"].value = "정규직"
    ws_info["C12"].value = "계약직"
    ws_info["D12"].value = "2024년"
    ws_info["E12"].value = "2023년"
    for cell in ["B12", "C12", "D12", "E12"]:
        ws_info[cell].font = bold

    ef = draft.get("employees_financials") or {}
    ws_info["B13"].value = ef.get("employees_full_time")
    ws_info["C13"].value = ef.get("employees_contract")
    ws_info["D13"].value = ef.get("revenue_2024")
    ws_info["E13"].value = ef.get("revenue_2023")
    ws_info["F13"].value = ef.get("equity_total")
    ws_info["G13"].value = ef.get("certification")

    ws_info["B15"].value = "투자이력"
    ws_info["F15"].value = "2025년 내 희망 투자액"
    ws_info["H15"].value = "투자전 희망 기업가치(Pre-Valuation)"
    for cell in ["B15", "F15", "H15"]:
        ws_info[cell].font = bold

    inv = draft.get("investment") or {}
    ws_info["B16"].value = inv.get("investment_history")
    ws_info["F16"].value = inv.get("desired_investment_amount")
    ws_info["H16"].value = inv.get("pre_money_valuation")
    ws_info["B16"].alignment = wrap_top

    ws_info.freeze_panes = "A6"
    ws_info.column_dimensions["B"].width = 22
    ws_info.column_dimensions["C"].width = 16
    ws_info.column_dimensions["D"].width = 24
    ws_info.column_dimensions["E"].width = 18
    ws_info.column_dimensions["F"].width = 16
    ws_info.column_dimensions["G"].width = 20
    ws_info.column_dimensions["H"].width = 20

    # (기업용) 2. 체크리스트
    ws_check["B4"].value = "모듈"
    ws_check["C4"].value = "No"
    ws_check["D4"].value = "항목"
    ws_check["E4"].value = "세부항목"
    ws_check["F4"].value = "질문"
    ws_check["G4"].value = "현황"
    ws_check["H4"].value = "근거/요청"
    for cell in ["B4", "C4", "D4", "E4", "F4", "G4", "H4"]:
        ws_check[cell].font = bold

    answers = draft.get("checklist_answers") or {}
    row = 5
    for it in template.get("checklist_items") or []:
        module = it.get("module")
        it_id = it.get("id")
        ws_check.cell(row, 2).value = module
        ws_check.cell(row, 3).value = it.get("no")
        ws_check.cell(row, 4).value = it.get("item")
        ws_check.cell(row, 5).value = it.get("sub_item")
        ws_check.cell(row, 6).value = it.get("question")

        ans = (answers or {}).get(it_id) if isinstance(answers, dict) else None
        if isinstance(ans, dict):
            ws_check.cell(row, 7).value = ans.get("status")
            ws_check.cell(row, 8).value = ans.get("detail")

        ws_check.cell(row, 6).alignment = wrap_top
        ws_check.cell(row, 8).alignment = wrap_top
        row += 1

    ws_check.freeze_panes = "A5"
    ws_check.column_dimensions["B"].width = 12
    ws_check.column_dimensions["C"].width = 6
    ws_check.column_dimensions["D"].width = 18
    ws_check.column_dimensions["E"].width = 18
    ws_check.column_dimensions["F"].width = 60
    ws_check.column_dimensions["G"].width = 10
    ws_check.column_dimensions["H"].width = 40

    # (기업용) 3. KPI기대사항
    ws_kpi["B2"].value = "3. KPI 및 기대사항"
    ws_kpi["B2"].font = Font(bold=True, size=14)

    ws_kpi["B4"].value = "Business"
    ws_kpi["C4"].value = "서비스/제품 소개"
    ws_kpi["D4"].value = "수익 구조"
    ws_kpi["E4"].value = "핵심 고객"
    ws_kpi["F4"].value = "KPI  (정량적 수치)"
    ws_kpi["G4"].value = "현황  (프로그램 시작 시점)"
    ws_kpi["H4"].value = "목표 수준  (프로그램 종료 시점)"
    for cell in ["B4", "C4", "D4", "E4", "F4", "G4", "H4"]:
        ws_kpi[cell].font = bold

    kpi_section = draft.get("kpi") or {}
    business_row = 5
    ws_kpi["C5"].value = kpi_section.get("service_intro")
    ws_kpi["D5"].value = kpi_section.get("revenue_model")
    ws_kpi["E5"].value = kpi_section.get("core_customer")

    kpi_items = kpi_section.get("kpis") if isinstance(kpi_section, dict) else []
    if not isinstance(kpi_items, list):
        kpi_items = []

    for idx, item in enumerate(kpi_items[:5]):
        r = business_row + idx
        ws_kpi.cell(r, 6).value = (item or {}).get("name")
        ws_kpi.cell(r, 7).value = (item or {}).get("current")
        ws_kpi.cell(r, 8).value = (item or {}).get("target")
        ws_kpi.cell(r, 7).alignment = wrap_top
        ws_kpi.cell(r, 8).alignment = wrap_top

    milestone_header_row = 10
    ws_kpi[f"B{milestone_header_row}"].value = "Milestone"
    ws_kpi[f"C{milestone_header_row}"].value = "국내 사업 계획(2025)"
    ws_kpi[f"D{milestone_header_row}"].value = "글로벌 확장 계획(2025)"
    ws_kpi[f"E{milestone_header_row}"].value = "장기 목표 (3년 내)"
    ws_kpi[f"F{milestone_header_row}"].value = "프로그램 목표/기대사항"
    ws_kpi[f"G{milestone_header_row}"].value = "올해 성장/성과 기대&목표"
    ws_kpi[f"H{milestone_header_row}"].value = "우려/리스크"
    for cell in [
        f"B{milestone_header_row}",
        f"C{milestone_header_row}",
        f"D{milestone_header_row}",
        f"E{milestone_header_row}",
        f"F{milestone_header_row}",
        f"G{milestone_header_row}",
        f"H{milestone_header_row}",
    ]:
        ws_kpi[cell].font = bold

    ms = kpi_section.get("milestone") if isinstance(kpi_section, dict) else {}
    if not isinstance(ms, dict):
        ms = {}

    milestone_row = milestone_header_row + 1
    ws_kpi[f"C{milestone_row}"].value = ms.get("domestic_plan_2025")
    ws_kpi[f"D{milestone_row}"].value = ms.get("global_plan_2025")
    ws_kpi[f"E{milestone_row}"].value = ms.get("long_term_goal_3y")
    ws_kpi[f"F{milestone_row}"].value = ms.get("program_expectation")
    ws_kpi[f"G{milestone_row}"].value = ms.get("growth_goal")
    ws_kpi[f"H{milestone_row}"].value = ms.get("concerns")
    for col in ["C", "D", "E", "F", "G", "H"]:
        ws_kpi[f"{col}{milestone_row}"].alignment = wrap_top

    ws_kpi.freeze_panes = "A5"
    ws_kpi.column_dimensions["B"].width = 12
    ws_kpi.column_dimensions["C"].width = 28
    ws_kpi.column_dimensions["D"].width = 28
    ws_kpi.column_dimensions["E"].width = 22
    ws_kpi.column_dimensions["F"].width = 18
    ws_kpi.column_dimensions["G"].width = 26
    ws_kpi.column_dimensions["H"].width = 26

    # (컨설턴트용) 분석보고서
    ws_report["C5"].value = "기업명"
    ws_report["D5"].value = "작성일시"
    ws_report["C5"].font = bold
    ws_report["D5"].font = bold

    ws_report["D6"].value = (draft.get("company_info") or {}).get("company_name")
    ws_report["D7"].value = datetime.now().strftime("%Y-%m-%d %H:%M")

    categories = template.get("categories") or ["문제", "솔루션", "사업화", "자금조달", "팀/조직", "임팩트"]
    weights = template.get("weights") or {}
    for idx, cat in enumerate(categories[:6]):
        col = 3 + idx
        ws_report.cell(9, col).value = cat
        ws_report.cell(9, col).font = bold
        ws_report.cell(10, col).value = weights.get(cat)

    ws_report["B15"].value = "기업 상황 요약(기업진단)"
    ws_report["B19"].value = "개선 필요사항"
    ws_report["B15"].font = bold
    ws_report["B19"].font = bold
    ws_report["C16"].alignment = wrap_top
    ws_report["C20"].alignment = wrap_top

    ws_report.column_dimensions["B"].width = 18
    ws_report.column_dimensions["C"].width = 45
    ws_report.column_dimensions["D"].width = 28

    try:
        wb.save(output_path)
    except Exception as e:
        return {"success": False, "error": f"엑셀 저장 실패: {str(e)}"}

    return {
        "success": True,
        "output_file": str(output_path),
        "message": f"기업현황 진단시트 생성 완료: {output_path.name}",
    }


def execute_analyze_company_diagnosis_sheet(excel_path: str) -> Dict[str, Any]:
    """기업현황 진단시트 분석"""
    is_valid, error = _validate_file_path(excel_path, allowed_extensions=[".xlsx", ".xls"], require_temp_dir=True)
    if not is_valid:
        return {"success": False, "error": error}

    try:
        from openpyxl import load_workbook

        wb = load_workbook(excel_path, data_only=False)

        company_info = _extract_diagnosis_company_info(wb)

        checklist_items: List[Dict[str, Any]] = []
        checklist_summary: Dict[str, Any] = {}
        gaps: List[Dict[str, Any]] = []

        if DIAG_SHEET_CHECKLIST in wb.sheetnames:
            ws = wb[DIAG_SHEET_CHECKLIST]
            current_module_raw = None

            for row in range(5, ws.max_row + 1):
                module_raw = ws.cell(row, 2).value
                q_no = ws.cell(row, 3).value
                item = ws.cell(row, 4).value
                sub_item = ws.cell(row, 5).value
                question = ws.cell(row, 6).value
                status = ws.cell(row, 7).value
                detail = ws.cell(row, 8).value

                if module_raw is not None:
                    current_module_raw = str(module_raw).strip()

                status_str = str(status).strip() if isinstance(status, str) else ""
                if status_str not in ("예", "아니오"):
                    continue

                module = _normalize_diagnosis_category(current_module_raw)
                entry = {
                    "row": row,
                    "module": module,
                    "module_raw": current_module_raw,
                    "no": q_no,
                    "item": item,
                    "sub_item": sub_item,
                    "question": question,
                    "status": status_str,
                    "detail": detail,
                }
                checklist_items.append(entry)

                if module not in checklist_summary:
                    checklist_summary[module] = {"total": 0, "yes": 0, "no": 0}

                checklist_summary[module]["total"] += 1
                if status_str == "예":
                    checklist_summary[module]["yes"] += 1
                else:
                    checklist_summary[module]["no"] += 1
                    gaps.append(
                        {
                            "module": module,
                            "question": question,
                            "detail": detail,
                            "row": row,
                        }
                    )

        kpi: Dict[str, Any] = {"business": [], "milestone": []}
        if DIAG_SHEET_KPI in wb.sheetnames:
            ws = wb[DIAG_SHEET_KPI]
            current_section = None
            for row in range(5, min(ws.max_row, 260) + 1):
                section = ws.cell(row, 2).value
                if section is not None:
                    current_section = str(section).strip()

                if current_section == "Business":
                    kpi_row = {
                        "kpi": ws.cell(row, 6).value,
                        "current": ws.cell(row, 7).value,
                        "target": ws.cell(row, 8).value,
                    }
                    if row == 5:
                        kpi_row.update(
                            {
                                "service_intro": ws.cell(row, 3).value,
                                "revenue_model": ws.cell(row, 4).value,
                                "core_customer": ws.cell(row, 5).value,
                            }
                        )
                    if any(v is not None and str(v).strip() != "" for v in kpi_row.values()):
                        kpi["business"].append(kpi_row)

                elif current_section == "Milestone":
                    m_row = {
                        "domestic_plan": ws.cell(row, 3).value,
                        "global_plan": ws.cell(row, 4).value,
                        "long_term_goal": ws.cell(row, 5).value,
                        "program_expectation": ws.cell(row, 6).value,
                        "growth_goal": ws.cell(row, 7).value,
                        "concerns": ws.cell(row, 8).value,
                    }
                    if any(v is not None and str(v).strip() != "" for v in m_row.values()):
                        kpi["milestone"].append(m_row)

        default_weights = {
            "문제": 20,
            "솔루션": 20,
            "사업화": 20,
            "자금조달": 15,
            "팀/조직": 20,
            "임팩트": 5,
        }

        weights = dict(default_weights)
        if DIAG_SHEET_REPORT in wb.sheetnames:
            ws = wb[DIAG_SHEET_REPORT]
            try:
                header_row = 9
                weight_row = 10
                for col in range(3, 9):
                    header = _normalize_diagnosis_category(ws.cell(header_row, col).value)
                    weight_val = ws.cell(weight_row, col).value
                    if header and isinstance(weight_val, (int, float)):
                        weights[header] = float(weight_val)
            except Exception:
                pass

        scores: Dict[str, Any] = {}
        for category, weight in weights.items():
            stats = checklist_summary.get(category, {"total": 0, "yes": 0, "no": 0})
            total = stats.get("total", 0) or 0
            yes = stats.get("yes", 0) or 0
            no = stats.get("no", 0) or 0
            yes_rate = round((yes / total * 100), 1) if total else None
            score = round((yes / total * float(weight)), 1) if total else None
            scores[category] = {
                "weight": float(weight),
                "score": score,
                "yes": yes,
                "no": no,
                "total": total,
                "yes_rate_pct": yes_rate,
            }

        return {
            "success": True,
            "excel_path": excel_path,
            "sheets": wb.sheetnames,
            "company_info": company_info,
            "checklist": {
                "items": checklist_items,
                "summary": checklist_summary,
                "gaps": gaps,
            },
            "kpi": kpi,
            "scores": scores,
        }

    except Exception as e:
        logger.error(f"Diagnosis sheet analysis failed: {e}", exc_info=True)
        return {"success": False, "error": f"진단시트 분석 실패: {str(e)}"}


def execute_write_company_diagnosis_report(
    excel_path: str,
    scores: Dict[str, float],
    summary_text: str,
    improvement_text: str,
    company_name: str = None,
    report_datetime: str = None,
    output_filename: str = None,
) -> Dict[str, Any]:
    """컨설턴트용 분석보고서 시트에 보고서 반영 후 새 파일 생성"""
    is_valid, error = _validate_file_path(excel_path, allowed_extensions=[".xlsx", ".xls"], require_temp_dir=True)
    if not is_valid:
        return {"success": False, "error": error}

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment

        wb = load_workbook(excel_path, data_only=False)
        if DIAG_SHEET_REPORT not in wb.sheetnames:
            return {"success": False, "error": f"'{DIAG_SHEET_REPORT}' 시트를 찾을 수 없습니다"}

        extracted_info = _extract_diagnosis_company_info(wb)
        final_company_name = (company_name or extracted_info.get("company_name") or "").strip()

        ws = wb[DIAG_SHEET_REPORT]

        if final_company_name:
            ws["D6"].value = final_company_name

        if not report_datetime:
            report_datetime = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws["D7"].value = report_datetime

        header_to_col: Dict[str, int] = {}
        for col in range(3, 9):
            header = _normalize_diagnosis_category(ws.cell(9, col).value)
            if header:
                header_to_col[header] = col

        for category, val in (scores or {}).items():
            cat = _normalize_diagnosis_category(category)
            if not cat:
                continue
            col = header_to_col.get(cat)
            if not col:
                continue
            try:
                ws.cell(11, col).value = float(val)
            except (TypeError, ValueError):
                continue

        ws["C16"].value = summary_text
        ws["C20"].value = improvement_text

        wrap = Alignment(wrap_text=True, vertical="top")
        ws["C16"].alignment = wrap
        ws["C20"].alignment = wrap

        excel_path_obj = Path(excel_path)
        user_id = "cli_user"
        try:
            if "temp" in excel_path_obj.parts:
                temp_idx = excel_path_obj.parts.index("temp")
                if len(excel_path_obj.parts) > temp_idx + 1:
                    user_id = excel_path_obj.parts[temp_idx + 1]
        except (ValueError, IndexError):
            pass

        if not output_filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"diagnosis_report_{timestamp}.xlsx"
        else:
            output_filename = _sanitize_filename(output_filename)
            if not output_filename.endswith(".xlsx"):
                output_filename += ".xlsx"

        output_dir = PROJECT_ROOT / "temp" / user_id
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / output_filename

        wb.save(output_path)

        return {
            "success": True,
            "output_file": str(output_path),
            "message": f"진단 분석보고서 반영 완료: {output_path.name}",
        }

    except Exception as e:
        logger.error(f"Diagnosis report write failed: {e}", exc_info=True)
        return {"success": False, "error": f"진단 분석보고서 반영 실패: {str(e)}"}


EXECUTORS = {
    "create_company_diagnosis_draft": execute_create_company_diagnosis_draft,
    "update_company_diagnosis_draft": execute_update_company_diagnosis_draft,
    "generate_company_diagnosis_sheet_from_draft": execute_generate_company_diagnosis_sheet_from_draft,
    "analyze_company_diagnosis_sheet": execute_analyze_company_diagnosis_sheet,
    "write_company_diagnosis_report": execute_write_company_diagnosis_report,
}
