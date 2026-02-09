"""
Excel extraction and analysis tools.

Read, analyze, and generate projections from investment Excel files.
"""

import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from ._common import (
    PROJECT_ROOT,
    _sanitize_filename,
    _validate_file_path,
    _validate_numeric_param,
    logger,
)

TOOLS = [
    {
        "name": "read_excel_as_text",
        "description": "엑셀 파일을 텍스트로 변환하여 읽습니다. 모든 시트의 내용을 텍스트 형식으로 반환하므로, 엑셀 구조가 다양해도 유연하게 대응할 수 있습니다. 이 도구로 먼저 엑셀 내용을 읽은 후, 필요한 정보를 파악하세요.",
        "input_schema": {
            "type": "object",
            "properties": {
                "excel_path": {
                    "type": "string",
                    "description": "읽을 엑셀 파일 경로",
                },
                "sheet_names": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "읽을 시트 이름 리스트 (선택사항, 없으면 모든 시트)",
                },
                "max_rows": {
                    "type": "integer",
                    "description": "각 시트에서 읽을 최대 행 수 (기본값: 50)",
                },
            },
            "required": ["excel_path"],
        },
    },
    {
        "name": "analyze_excel",
        "description": "투자 검토 엑셀 파일을 자동으로 분석하여 투자조건, IS요약(연도별 당기순이익), Cap Table(총발행주식수)을 추출합니다. 일반적인 엑셀 구조에서 작동하지만, 구조가 특이하면 read_excel_as_text를 사용하세요.",
        "input_schema": {
            "type": "object",
            "properties": {
                "excel_path": {
                    "type": "string",
                    "description": "분석할 엑셀 파일 경로",
                }
            },
            "required": ["excel_path"],
        },
    },
    {
        "name": "analyze_and_generate_projection",
        "description": "엑셀 파일을 분석하고 즉시 Exit 프로젝션을 생성합니다. 파일에서 투자 조건과 재무 데이터를 자동으로 추출한 후, 지정된 연도와 PER 배수로 Exit 시나리오를 계산하여 새로운 엑셀 파일을 생성합니다.",
        "input_schema": {
            "type": "object",
            "properties": {
                "excel_path": {
                    "type": "string",
                    "description": "분석할 투자검토 엑셀 파일 경로",
                },
                "target_year": {
                    "type": "integer",
                    "description": "Exit 목표 연도 (예: 2028, 2030)",
                },
                "per_multiples": {
                    "type": "array",
                    "items": {"type": "number"},
                    "description": "PER 배수 리스트 (예: [10, 20, 30])",
                },
                "company_name": {
                    "type": "string",
                    "description": "회사명 (선택사항)",
                },
                "output_filename": {
                    "type": "string",
                    "description": "출력 파일명 (선택사항, 기본값: exit_projection_YYYYMMDD_HHMMSS.xlsx)",
                },
            },
            "required": ["excel_path", "target_year", "per_multiples"],
        },
    },
]


def execute_read_excel_as_text(
    excel_path: str, sheet_names: List[str] = None, max_rows: int = 50
) -> Dict[str, Any]:
    """엑셀 파일을 텍스트로 변환하여 읽기"""
    is_valid, error = _validate_file_path(
        excel_path, allowed_extensions=[".xlsx", ".xls"], require_temp_dir=True
    )
    if not is_valid:
        return {"success": False, "error": error}

    wb = None
    try:
        from openpyxl import load_workbook

        if not os.path.exists(excel_path):
            return {"success": False, "error": f"파일을 찾을 수 없습니다: {excel_path}"}

        wb = load_workbook(excel_path, data_only=True)

        sheets_data = {}
        target_sheets = sheet_names if sheet_names else wb.sheetnames

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            sheet_text = []

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx > max_rows:
                    break

                row_values = [str(cell) if cell is not None else "" for cell in row]

                if not any(val.strip() for val in row_values):
                    continue

                row_text = " | ".join(row_values[:15])
                sheet_text.append(f"Row {row_idx}: {row_text}")

            sheets_data[sheet_name] = "\n".join(sheet_text)

        result_text = ""
        for sheet_name, content in sheets_data.items():
            result_text += f"\n{'='*60}\n"
            result_text += f"시트: {sheet_name}\n"
            result_text += f"{'='*60}\n"
            result_text += content + "\n"

        logger.info(f"Excel read successfully: {excel_path} ({len(sheets_data)} sheets)")
        return {
            "success": True,
            "file_path": excel_path,
            "sheets": list(sheets_data.keys()),
            "content": result_text,
            "total_sheets": len(sheets_data),
        }

    except FileNotFoundError:
        return {"success": False, "error": f"파일을 찾을 수 없습니다: {excel_path}"}
    except PermissionError:
        return {"success": False, "error": f"파일 접근 권한이 없습니다: {excel_path}"}
    except Exception as e:
        logger.error(f"Failed to read excel {excel_path}: {e}", exc_info=True)
        return {"success": False, "error": f"엑셀 파일 읽기 실패: {str(e)}"}
    finally:
        if wb is not None:
            wb.close()


def execute_analyze_excel(excel_path: str) -> Dict[str, Any]:
    """엑셀 파일 분석 실행 - openpyxl로 직접 읽기"""
    is_valid, error = _validate_file_path(
        excel_path, allowed_extensions=[".xlsx", ".xls"], require_temp_dir=True
    )
    if not is_valid:
        return {"success": False, "error": error}

    from openpyxl import load_workbook

    wb = None
    try:
        if not os.path.exists(excel_path):
            return {"success": False, "error": f"파일을 찾을 수 없습니다: {excel_path}"}

        wb = load_workbook(excel_path, data_only=True)

        result = {
            "success": True,
            "file_path": excel_path,
            "sheets": wb.sheetnames,
            "investment_terms": {},
            "income_statement": {},
            "cap_table": {},
        }

        # IS요약 시트에서 순이익 데이터 추출
        is_sheet = None
        for sheet_name in wb.sheetnames:
            if "IS" in sheet_name or "손익" in sheet_name:
                is_sheet = wb[sheet_name]
                break

        if is_sheet:
            year_row_idx = None
            year_cols = {}

            for row_idx, row in enumerate(is_sheet.iter_rows(min_row=1, max_row=10), start=1):
                for col_idx, cell in enumerate(row):
                    if cell.value and isinstance(cell.value, str) and "년" in cell.value:
                        try:
                            year_val = int(cell.value.replace("년", "").replace(",", ""))
                            if 2020 <= year_val <= 2040:
                                year_row_idx = row_idx
                                year_cols[year_val] = col_idx
                        except ValueError:
                            pass

            net_income_data = {}
            if year_cols:
                for row in is_sheet.iter_rows(min_row=year_row_idx if year_row_idx else 1):
                    first_cell = row[1].value if len(row) > 1 else None
                    if first_cell and "당기순이익" in str(first_cell):
                        for year, col_idx in year_cols.items():
                            if col_idx < len(row):
                                value = row[col_idx].value
                                if value and isinstance(value, (int, float)):
                                    net_income_data[year] = int(value)
                        break

            result["income_statement"] = {
                "years": sorted(year_cols.keys()) if year_cols else [],
                "net_income": net_income_data,
            }

        # Cap Table에서 총 발행주식수 추출
        cap_sheet = None
        for sheet_name in wb.sheetnames:
            if "cap" in sheet_name.lower() or "주주" in sheet_name:
                cap_sheet = wb[sheet_name]
                break

        if cap_sheet:
            for row in cap_sheet.iter_rows():
                first_cell = row[0].value if row else None
                if first_cell and "합계" in str(first_cell):
                    if len(row) > 3 and row[3].value and isinstance(row[3].value, (int, float)):
                        incorporation_shares = int(row[3].value)
                        seed_shares = 0
                        if len(row) > 6 and row[6].value and isinstance(row[6].value, (int, float)):
                            seed_shares = int(row[6].value)

                        total_shares = incorporation_shares + seed_shares
                        result["cap_table"]["total_shares"] = total_shares
                        result["cap_table"]["incorporation_shares"] = incorporation_shares
                        result["cap_table"]["seed_shares"] = seed_shares
                        break

        # 투자조건 시트에서 투자 정보 추출
        invest_sheet = None
        for sheet_name in wb.sheetnames:
            if "투자조건" in sheet_name:
                invest_sheet = wb[sheet_name]
                break

        if invest_sheet:
            for row_idx, row in enumerate(invest_sheet.iter_rows(min_row=1, max_row=30)):
                if len(row) < 4:
                    continue

                second_cell = row[1].value if row[1] else None
                if not second_cell:
                    continue

                second_val = str(second_cell)

                if "투자금액" in second_val and "원" in second_val:
                    for cell in row[3:]:
                        if cell.value and isinstance(cell.value, (int, float)):
                            result["investment_terms"]["investment_amount"] = int(cell.value)
                            break

                if "투자단가" in second_val and "원" in second_val:
                    for cell in row[3:]:
                        if cell.value and isinstance(cell.value, (int, float)):
                            result["investment_terms"]["price_per_share"] = int(cell.value)
                            break

                if "투자주식수" in second_val:
                    for cell in row[3:]:
                        if cell.value and isinstance(cell.value, (int, float)):
                            result["investment_terms"]["shares"] = int(cell.value)
                            break

        logger.info(f"Excel analyzed successfully: {excel_path}")
        return result

    except FileNotFoundError:
        return {"success": False, "error": f"파일을 찾을 수 없습니다: {excel_path}"}
    except PermissionError:
        return {"success": False, "error": f"파일 접근 권한이 없습니다: {excel_path}"}
    except Exception as e:
        logger.error(f"Failed to analyze excel {excel_path}: {e}", exc_info=True)
        return {"success": False, "error": f"엑셀 파일 분석 실패: {str(e)}"}
    finally:
        if wb is not None:
            wb.close()


def execute_analyze_and_generate_projection(
    excel_path: str,
    target_year: int,
    per_multiples: List[float],
    company_name: str = None,
    output_filename: str = None,
) -> Dict[str, Any]:
    """엑셀 파일 분석 후 즉시 Exit 프로젝션 생성"""
    is_valid, error = _validate_file_path(excel_path, allowed_extensions=[".xlsx", ".xls"])
    if not is_valid:
        return {"success": False, "error": error}

    is_valid, year_val, error = _validate_numeric_param(
        target_year, "target_year", min_val=2020, max_val=2050
    )
    if not is_valid:
        return {"success": False, "error": error}
    target_year = int(year_val)

    if not per_multiples or not isinstance(per_multiples, list):
        return {"success": False, "error": "PER 멀티플 리스트가 필요합니다"}
    validated_multiples = []
    for m in per_multiples:
        is_valid, val, error = _validate_numeric_param(m, "PER multiple", min_val=0.1, max_val=1000)
        if not is_valid:
            return {"success": False, "error": error}
        validated_multiples.append(val)
    per_multiples = validated_multiples

    # 1단계: 엑셀 파일 분석
    analysis = execute_analyze_excel(excel_path)

    if not analysis["success"]:
        return analysis

    data = analysis
    income_statement = data.get("income_statement", {})
    net_income_data = income_statement.get("net_income", {})
    investment_terms = data.get("investment_terms", {})
    cap_table = data.get("cap_table", {})

    # 2단계: 필수 데이터 검증
    if target_year not in net_income_data:
        return {
            "success": False,
            "error": f"{target_year}년 순이익 데이터를 찾을 수 없습니다. 사용 가능한 연도: {list(net_income_data.keys())}",
        }

    net_income = net_income_data[target_year]
    investment_amount = investment_terms.get("investment_amount")
    price_per_share = investment_terms.get("price_per_share")
    shares = investment_terms.get("shares")
    total_shares = cap_table.get("total_shares")

    if not all([investment_amount, price_per_share, shares, total_shares]):
        return {
            "success": False,
            "error": "필수 투자 정보가 부족합니다",
            "found_data": {
                "investment_amount": investment_amount,
                "price_per_share": price_per_share,
                "shares": shares,
                "total_shares": total_shares,
            },
        }

    # 3단계: Exit 프로젝션 요약 계산
    investment_year = datetime.now().year
    holding_period_years = target_year - investment_year

    projection_summary: List[Dict[str, Any]] = []
    for per in per_multiples:
        enterprise_value = float(net_income) * float(per)
        proceeds = None
        multiple = None
        irr_pct = None

        try:
            per_share_value = enterprise_value / float(total_shares)
            proceeds = per_share_value * float(shares)
            multiple = proceeds / float(investment_amount)
            if holding_period_years > 0 and multiple > 0:
                irr_pct = (multiple ** (1 / holding_period_years) - 1) * 100
        except Exception:
            pass

        projection_summary.append(
            {"PER": float(per), "IRR": irr_pct, "Multiple": multiple}
        )

    # 4단계: Exit 프로젝션 생성
    excel_path_obj = Path(excel_path)
    user_id = "cli_user"
    try:
        if "temp" in excel_path_obj.parts:
            temp_idx = excel_path_obj.parts.index("temp")
            if len(excel_path_obj.parts) > temp_idx + 1:
                user_id = excel_path_obj.parts[temp_idx + 1]
    except (ValueError, IndexError):
        pass

    if not output_filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"exit_projection_{timestamp}.xlsx"
    else:
        output_filename = _sanitize_filename(output_filename)
        if not output_filename.endswith(".xlsx"):
            output_filename += ".xlsx"

    output_dir = PROJECT_ROOT / "temp" / user_id
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / output_filename

    if not company_name:
        company_name = _sanitize_filename(Path(excel_path).stem)
    else:
        company_name = _sanitize_filename(company_name)

    script_path = PROJECT_ROOT / "scripts" / "generate_exit_projection.py"

    cmd = [
        sys.executable,
        str(script_path),
        "--investment_amount",
        str(int(investment_amount)),
        "--price_per_share",
        str(int(price_per_share)),
        "--shares",
        str(int(shares)),
        "--total_shares",
        str(int(total_shares)),
        "--net_income_company",
        str(int(net_income)),
        "--net_income_reviewer",
        str(int(net_income)),
        "--target_year",
        str(target_year),
        "--company_name",
        company_name,
        "--per_multiples",
        ",".join(map(lambda x: str(int(x) if x == int(x) else x), per_multiples)),
        "--output",
        str(output_path),
    ]

    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, check=True, cwd=str(PROJECT_ROOT)
        )

        return {
            "success": True,
            "output_file": str(output_path),
            "assumptions": {
                "company_name": company_name,
                "target_year": target_year,
                "investment_year": investment_year,
                "holding_period_years": holding_period_years,
                "investment_amount": investment_amount,
                "shares": shares,
                "total_shares": total_shares,
                "net_income": net_income,
                "per_multiples": per_multiples,
            },
            "projection_summary": projection_summary,
            "analysis_data": {
                "target_year": target_year,
                "net_income": net_income,
                "investment_amount": investment_amount,
                "per_multiples": per_multiples,
                "company_name": company_name,
            },
            "message": f"Exit 프로젝션 생성 완료: {output_path.name}",
        }

    except subprocess.CalledProcessError as e:
        return {"success": False, "error": f"Exit 프로젝션 생성 실패: {e.stderr}"}


EXECUTORS = {
    "read_excel_as_text": execute_read_excel_as_text,
    "analyze_excel": execute_analyze_excel,
    "analyze_and_generate_projection": execute_analyze_and_generate_projection,
}
